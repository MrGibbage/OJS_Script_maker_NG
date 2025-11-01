# pip install pandas
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.cell import (
    column_index_from_string,
    get_column_letter,
    coordinate_from_string,
)
from openpyxl.worksheet.table import Table

# pip install pyyaml
import yaml
import warnings

import os, sys, re, glob
from pathlib import Path

# pip install Jinja2
from jinja2 import Environment, FileSystemLoader, select_autoescape

from typing import List, Dict

# pip install colorama
from colorama import Fore, Back, Style, init

# To create windows exe executable, run
# .venv\Scripts\pyinstaller.exe -F 2025\script-maker-mac-win.py -n script_maker-win
# in the project folder. The executable will be saved in the 'dist'
# folder. Just copy it up to the project folder.
# Double-click to run.


# Brief: Read an Excel table (openpyxl) and return (columns, data dict)
def read_excel_table(sheet, table_name):
    """
    This function will read an Excel table
    and return a tuple of columns and data

    This function assumes that tables have column headers
    :param sheet: the sheet
    :param table_name: the name of the table
    :return: columns (list) and data (dict)
    """
    table = sheet.tables[table_name]
    table_range = table.ref

    table_head = sheet[table_range][0]
    table_data = sheet[table_range][1:]

    columns = [column.value for column in table_head]
    data = {column: [] for column in columns}

    for row in table_data:
        row_val = [cell.value for cell in row]
        for key, val in zip(columns, row_val):
            data[key].append(val)

    return columns, data

# Brief: Check a worksheet range for any empty cells; returns True if any found
def check_range_for_empty_cells(worksheet, range_string):
    """
    Check if there are any empty cells in the specified range of an Excel worksheet.
    
    Args:
        worksheet: The Excel worksheet object
        range_string: String specifying the range (e.g. "A1:D10")
    
    Returns:
        bool: True if empty cells found, False if all cells have values
    """
    cells = worksheet[range_string]
    empty_cells = []
    
    # Handle both single row/column and rectangular ranges
    if not isinstance(cells[0], tuple):
        cells = [cells]
        
    for row_index, row in enumerate(cells):
        for col_index, cell in enumerate(row):
            if cell.value is None or cell.value == "":
                empty_cells.append(f"{cell.coordinate}")
    
    if empty_cells:
        warning_msg = f"Empty cells found in {range_string}: {', '.join(empty_cells)}"
        print(Fore.YELLOW + warning_msg)
        return True
        
    return False

# Brief: Verify numeric cells in range are within min_val..max_val (inclusive)
def check_range_for_valid_numbers(worksheet, range_string, min_val, max_val):
    """
    Check if all numeric values in the specified range fall between min and max values.
    
    Args:
        worksheet: The Excel worksheet object
        range_string: String specifying the range (e.g. "A1:D10")
        min_val: Minimum allowed value (inclusive)
        max_val: Maximum allowed value (inclusive)
    
    Returns:
        bool: True if invalid values found, False if all values are valid
    """
    cells = worksheet[range_string]
    invalid_cells = []
    
    # Handle both single row/column and rectangular ranges
    if not isinstance(cells[0], tuple):
        cells = [cells]
        
    for row in cells:
        for cell in row:
            # Skip empty cells
            if cell.value is None or cell.value == "":
                continue
                
            try:
                value = float(cell.value)
                if value < min_val or value > max_val:
                    invalid_cells.append(f"{cell.coordinate}={value}")
            except (ValueError, TypeError):
                invalid_cells.append(f"{cell.coordinate}=not a number")
    
    if invalid_cells:
        warning_msg = f"Invalid values found in {range_string} (must be between {min_val} and {max_val}): {', '.join(invalid_cells)}"
        print(Fore.YELLOW + warning_msg)
        return True
        
    return False

# Brief: Verify numeric cells in range are one of the allowed numeric values
def check_range_for_valid_values(worksheet, range_string, allowed_values):
    """
    Check if all numeric values in the specified range match the allowed values list.

    Args:
        worksheet: The Excel worksheet object
        range_string: String specifying the range (e.g. "A1:D10")
        allowed_values: List of allowed numeric values

    Returns:
        bool: True if invalid values found, False if all values are valid
    """
    cells = worksheet[range_string]
    invalid_cells = []

    # Handle both single row/column and rectangular ranges
    if not isinstance(cells[0], tuple):
        cells = [cells]
        
    for row in cells:
        for cell in row:
            # Skip empty cells
            if cell.value is None or cell.value == "":
                continue
                
            try:
                value = float(cell.value)
                if value not in allowed_values:
                    invalid_cells.append(f"{cell.coordinate}={value}")
            except (ValueError, TypeError):
                invalid_cells.append(f"{cell.coordinate}=not a number")
    
    if invalid_cells:
        warning_msg = f"Invalid values found in {range_string} (must be one of {allowed_values}): {', '.join(invalid_cells)}"
        print(Fore.YELLOW + warning_msg)
        return True
        
    return False

# ------------------------
# Helper utilities (added)
# ------------------------

# Brief: Print an error message and exit the program (centralized fatal handler)
def fatal(msg: str, pause: bool = True, code: int = 1) -> None:
    """Print fatal error and exit. Centralized to reduce repeated code."""
    print(Fore.RED + msg)
    if pause:
        try:
            input("Press enter to quit...")
        except Exception:
            pass
    sys.exit(code)


# Brief: Print a warning and prompt user to continue; allows abort via Ctrl-C
def warn_continue(msg: str) -> None:
    """Print a warning and allow the user to continue or abort with Ctrl-C."""
    print(Fore.YELLOW + msg)
    try:
        input("Press enter to continue. Press ctrl-c to quit...")
    except KeyboardInterrupt:
        fatal("\n\nStopped building the script. Please check that the OJS files are filled out correctly before trying to build the script again.", pause=False, code=0)


# Brief: Load an Excel workbook safely, fatal on error
def load_book(path):
    """Load a workbook, or fatal on error."""
    try:
        return load_workbook(path, data_only=True)
    except Exception as e:
        fatal(f"Fatal error. Could not open workbook {path}: {e}")


# Brief: Read a table from a workbook sheet and return as a pandas DataFrame
def df_from_table(book, sheet_name: str, table_name: str):
    """Read a table from a workbook sheet into a pandas DataFrame."""
    ws = book[sheet_name]
    cols, data = read_excel_table(ws, table_name)
    return pd.DataFrame(data=data, columns=cols)


# Brief: Run a sequence of validations (sheet, range, validator, args, err_msg)
def run_validations(book, validations):
    """Run a list of validations and fatal on the first failure.

    validations: iterable of (sheet_name, range_string, validator_fn, validator_args, err_msg)
    validator_fn should return True when an error condition is present (matching existing helpers).
    """
    for sheet_name, rng, fn, args, err_msg in validations:
        ws = book[sheet_name]
        if fn(ws, rng, *args):
            fatal(err_msg)


# Brief: Build a simple HTML <p> list of teams from the Rankings DataFrame
def build_team_list_html(dfRankings, div: int) -> str:
    """Return a single HTML string containing team list <p> entries for a division."""
    rows = []
    try:
        df_sorted = dfRankings.copy()
        df_sorted["_team_num_sort"] = pd.to_numeric(df_sorted["Team Number"], errors="coerce").fillna(0).astype(int)
        df_sorted = df_sorted.sort_values("_team_num_sort")
    except Exception:
        df_sorted = dfRankings
    for _, row in df_sorted.iterrows():
        try:
            team_num = str(int(row["Team Number"]))
        except Exception:
            team_num = ""
        team_name = row.get("Team Name", "") if hasattr(row, 'get') else row["Team Name"]
        rows.append(f"<p>(Div {div}) Team number {team_num}, {team_name}</p>")
    return "\n".join(rows) + ("\n" if rows else "")


# Brief: Build the Robot Game award HTML block from rankings and scores
def build_robot_game_html(dfRankings, div: int, count: int) -> str:
    """Build the Robot Game HTML block for a division, using awardCounts."""
    pieces = []
    for i in reversed(range(count)):
        rank = i + 1
        try:
            teamNum = int(dfRankings.loc[dfRankings["Robot Game Rank"] == rank, "Team Number"].values[0])
            teamName = dfRankings.loc[dfRankings["Robot Game Rank"] == rank, "Team Name"].values[0]
            score = int(dfRankings.loc[dfRankings["Robot Game Rank"] == rank, "Max Robot Game Score"].values[0])
        except Exception:
            fatal(
                "Fatal error. Some robot game scores are missing.\nAll robot game ranks must be properly visible on the Results and Rankings spreadsheet. Have you filled in the scores on the Robot Game Scores worksheet?"
            )
        pieces.append(f"<p>With a score of {score} points, the Division {div} {ordinals[i]} place award goes to team number {teamNum}, {teamName}</p>")
    return "\n".join(pieces) + ("\n" if pieces else "")


# Brief: Build HTML for other judged awards; warns if award not selected
def build_judged_awards_html(dfRankings, div: int, award: str, count: int) -> str:
    """Build HTML for other judged awards for a division."""
    pieces = []
    for i in reversed(range(count)):
        thisAward = award + " " + ordinals[i] + " Place"
        try:
            teamNum = dfRankings.loc[dfRankings["Award"] == thisAward, "Team Number"].values[0]
            teamName = dfRankings.loc[dfRankings["Award"] == thisAward, "Team Name"].values[0]
            pieces.append(f"<p>The Division {div} {ordinals[i]} place {award} award goes to team number {int(teamNum)}, {teamName}</p>")
        except Exception:
            warn_continue(Fore.YELLOW + f"{thisAward} award for Division {div} is missing. Have you selected it in the OJS?")

    return "\n".join(pieces) + ("\n" if pieces else "")

# Everything above here are just imports and function definitions.
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
# Everything below here is where the flow begins

# colorama init function
init()

print(Fore.LIGHTWHITE_EX+ "Building Closing Ceremony script")

if getattr(sys, "frozen", False):
    dir_path = os.path.dirname(sys.executable)
elif __file__:
    dir_path = os.path.dirname(__file__)

# need to account for how Windows vs. Unix/macOS deals with path names here -ac
if not os.path.exists(os.path.join(dir_path,"file_list.txt")):
    fatal(f"Fatal error. The file_list.txt file is missing in the tournament directory: {dir_path}")
    
# need to account for how Windows vs. Unix/macOS deals with path names here -ac
with open(os.path.join(dir_path, "file_list.txt"), "r") as file:
    for line in file:
        line = line.strip()  # Remove newline and extra spaces - changed path join next line -ac
        if not os.path.exists(os.path.join(dir_path, line)):
            fatal(f"Fatal error. The {line} file is missing in the tournament directory: {dir_path}")

templateLoader = FileSystemLoader(searchpath=dir_path)
templateEnv = Environment(loader=templateLoader)
TEMPLATE_FILE = "script_template.html.jinja"
try:
    template = templateEnv.get_template(TEMPLATE_FILE)
except Exception as e:
    fatal(f"Fatal error. Could not read the template file {TEMPLATE_FILE}.\nThe error was {e}")

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

ordinals = ["1st", "2nd", "3rd", "4th", "5th"]

# print("Opening yaml data file: " + yaml_data_file)
# with open(yaml_data_file) as f:
#     dict = yaml.load(f, Loader=yaml.FullLoader)

awards = {
    "Robot Game": "",
    "Champions": "",
    "Innovation Project": "",
    "Robot Design": "",
    "Core Values": "",
    "Judges": "",
}
rg_html = {}
dataframes = {}
# divAwards key will be 1 and/or 2, and the value will be the awardHtml dictionaries
# yes, it is a dictionary of dictionaries
divAwards: Dict[int, Dict[str, str]] = {}
divAwards[1] = {}
divAwards[2] = {}

teamList: List[str] = ["", "", ""]  # teamList[1], teamList[2] valid

# awardHtml key will be the award name, and the values will be the html
awardHtml = {}
advancingDf = {}
altAdvancingDf = {}
advancingHtml = {}
awardCounts = {}
judgesAwardDf = {}
judgesAwardHtml = {}
judgesAwardTotalCount = 0
divJudgesAwards = [0, 0]
allowedAdvancingCount = [0, 0]  # D1 is index 0
numTeams = [0, 0]

divisions = [1, 2]
for award in awards:
    for div in divisions:
        divAwards[div][award] = ""

removing = []
# path construction changed here for Unix/macOS systems, needs to be addressed for windows -ac
if len(glob.glob(f"{dir_path}/~*.xlsm")) > 0:
    fatal("Found temporary file(s) indicating that you have one or more spreadsheets open in Excel. Please close Excel and retry.")

# path construction changed here for Unix/macOS systems, needs to be addressed for windows -ac
directory_list: list[str] = glob.glob(f"{dir_path}/*div*.xlsm")
print(Fore.LIGHTWHITE_EX + f"Using this directory: {dir_path}")
print(Fore.LIGHTWHITE_EX + "Found these OJS files:" + Style.RESET_ALL)
print(directory_list)
if len(directory_list) == 0 or len(directory_list) > 2:
    fatal(f"Fatal error. There must be one or two OJS files in the directory. Found: {len(directory_list)}")

for award in awards:
    awardCounts[award] = 0

# Replace the long per-file processing loop with a streamlined implementation
# that uses the helpers above. This keeps the logic identical but centralizes
# validation, warnings, and HTML building.
for tourn_filename in directory_list:
    regex: str = r"([0-9]{4}-vadc-fll-challenge-.*)(-ojs-)(.*)-(div[1,2])(.xlsm)$"
    m0 = re.search(regex, Path(tourn_filename).name) if 'Path' in globals() else re.search(regex, os.path.basename(tourn_filename))
    if not m0:
        fatal(f"Unexpected OJS filename format: {tourn_filename}")

    div = int(m0.group(4)[-1])
    print(Fore.LIGHTWHITE_EX + f"Division {div}")

    book = load_book(tourn_filename)
    dfMeta = df_from_table(book, "Meta", "Meta")

    # safe extraction of awardCounts
    for award in awards:
        try:
            awardCounts[award] = dfMeta.loc[dfMeta["Key"] == award, "Value"].values[0]
        except Exception:
            awardCounts[award] = 0
    print("Award counts: " + str(awardCounts))

    dfRankings = df_from_table(book, "Results and Rankings", "TournamentData")
    numTeams = len(dfRankings)
    print(f"There are {numTeams} teams in Division {div}")
    print(Fore.LIGHTWHITE_EX + "Here is the Results and Rankings data")
    print(dfRankings)

    # Build and assign team list html
    teamList[div] = build_team_list_html(dfRankings, div)

    # validations consolidated
    validations = [
        ("Robot Game Scores", f"C2:E{numTeams+1}", check_range_for_empty_cells, (), f"There are missing scores on the Robot Game Scores worksheet for Div {div}"),
        ("Robot Game Scores", f"C2:E{numTeams+1}", check_range_for_valid_numbers, (0, 545), f"There are invalid scores on the Robot Game Scores worksheet for Div {div}"),
        ("Robot Design Input", f"D2:M{numTeams+1}", check_range_for_empty_cells, (), f"There are missing scores on the Robot Design Input worksheet for Div {div}"),
        ("Robot Design Input", f"D2:M{numTeams+1}", check_range_for_valid_values, ([0,1,2,3,4],), f"There are invalid values on the Robot Design Input worksheet for Div {div}"),
        ("Core Values Input", f"N2:P{numTeams+1}", check_range_for_empty_cells, (), f"There are missing scores on the Core Values Input worksheet for Div {div}"),
        ("Core Values Input", f"N2:P{numTeams+1}", check_range_for_valid_values, ([0,2,3,4],), f"There are invalid values on the Core Values Input worksheet for Div {div}"),
        ("Innovation Project Input", f"D2:M{numTeams+1}", check_range_for_empty_cells, (), f"There are missing scores on the Innovation Project Input worksheet for Div {div}"),
        ("Innovation Project Input", f"D2:M{numTeams+1}", check_range_for_valid_values, ([0,1,2,3,4],), f"There are invalid values on the Innovation Project Input worksheet for Div {div}"),
    ]
    run_validations(book, validations)
    print(Fore.LIGHTWHITE_EX + "It looks like all of the judging and robot game worksheets are filled in")

    # Robot Game HTML
    rg_html[div] = build_robot_game_html(dfRankings, div, awardCounts.get("Robot Game", 0))

    # Judged awards (all besides Robot Game and Judges)
    for award in awards:
        if award in ("Robot Game", "Judges"):
            continue
        divAwards[div][award] = build_judged_awards_html(dfRankings, div, award, awardCounts.get(award, 0))

    # Advancing
    advancingDf[div - 1] = dfRankings[dfRankings["Advance?"] == "Yes"]
    altAdvancingDf[div - 1] = dfRankings[dfRankings["Advance?"] == "Alt"]
    try:
        allowedAdvancingCount[div - 1] = int(dfMeta.loc[dfMeta["Key"] == "Advancing", "Value"].values[0])
    except Exception:
        allowedAdvancingCount[div - 1] = 0

    advancingHtml[div] = ""
    for index, row in advancingDf[div - 1].iterrows():
        try:
            teamNum = str(int(row["Team Number"]))
        except Exception:
            teamNum = ""
        teamName = row["Team Name"]
        advancingHtml[div] += f"<p>(Div {div}) Team number {teamNum}, {teamName}</p>\n"
    print(Fore.LIGHTWHITE_EX + f"Advancing: {advancingHtml[div]}")

    allowedJudgesAwardCount = 0
    try:
        allowedJudgesAwardCount = dfMeta.loc[dfMeta["Key"] == "Judges", "Value"].values[0]
    except Exception:
        allowedJudgesAwardCount = 0
    print(Fore.LIGHTWHITE_EX + f"Tournament has {allowedJudgesAwardCount} judges awards available across both divisions")

    judgesAwardDf[div] = dfRankings[(dfRankings["Award"].str.startswith("Judges", na=False))]
    judgesAwardHtml[div] = build_team_list_html(judgesAwardDf[div], div)
    judgesAwardTotalCount += len(judgesAwardDf[div])

    # Check for dupes in the awards
    filtered_df = dfRankings.dropna(subset=["Award"])[["Team Number", "Team Name", "Award"]]
    duplicate_rows = filtered_df[filtered_df.duplicated(subset=["Award"], keep=False)]
    if len(duplicate_rows) > 0:
        fatal("There are teams with duplicate awards\n" + str(duplicate_rows))

    print(Fore.GREEN + f"All done collecting data from the Div {div} OJS. Checking validity now.")
    try:
        divJudgesAwards[div - 1] = len(judgesAwardDf[div])
    except Exception:
        divJudgesAwards[div - 1] = 0

    print(Fore.LIGHTWHITE_EX + f"Total number of judges awards selected for Div {div} = {divJudgesAwards[div - 1]}")

    # Advancing checks
    try:
        divAdvancing = len(advancingDf[div - 1])
    except Exception:
        divAdvancing = 0

    try:
        divAltAdvancing = len(altAdvancingDf[div - 1])
    except Exception:
        divAltAdvancing = 0

    print(f"Total number of advancing teams selected for Div {div} = {divAdvancing}")
    if divAdvancing < allowedAdvancingCount[div - 1]:
        warn_continue(
            f"You have selected fewer advancing div {div} teams than allowed.\nYou are permitted a total of {allowedAdvancingCount[div - 1]} advancing teams for division {div}, but you have only selected {divAdvancing}.\n\nThis is not an error and you may continue building the script with fewer advancing teams than permitted."
        )
    if divAdvancing > allowedAdvancingCount[div - 1]:
        fatal(f"You have selected more advancing div {div} teams than allowed. You are permitted {allowedAdvancingCount[div - 1]}, but you have selected {divAdvancing}.")

    print(f"Total number of alt advancing teams selected for Div {div} = {divAltAdvancing}")
    if divAltAdvancing == 0:
        warn_continue(f"You have not selected one alternative advancing div {div} team.\nThis is not an error and you may continue building the script.")
    if divAltAdvancing > 1:
        warn_continue(f"You have selected more than one alternative advancing div {div} team.\nThis is not an error and you may continue building the script.")

# End of OJS loop
# # # # # # # # # # # # # # # # # # # # #

# Judges awards
judgesAwardTotalCount = divJudgesAwards[0] + divJudgesAwards[1]
if judgesAwardTotalCount > allowedJudgesAwardCount:
    fatal(f"You have selected too many judges awards: D1 = {divJudgesAwards[0]}; D2 = {divJudgesAwards[1]}. You are permitted a total of {allowedJudgesAwardCount} judges awards.\nIf your tournament has two divisions, that total number is across both divisions.\nIn other words, if you are allowed three judges awards, you could have three from\nDiv 1 or two from Div 1 and one from Div 2, etc.")

if judgesAwardTotalCount < allowedJudgesAwardCount:
    warn_continue(f"You have selected fewer judges awards than allowed. You are permitted a total of {allowedJudgesAwardCount} judges awards.\nIf your tournament has two divisions, that total number is across both divisions.\nFor example, if you are allowed three judges awards, you could have all three from Div 1.\nOr you could have two from Div 1 and one from Div 2, etc.\n\nThis is not an error and you may continue building the script with fewer judges awards than permitted.")

print(Fore.GREEN + 
    f"All checks look good."
)

print(Fore.LIGHTWHITE_EX + "Rendering the script")
out_text = template.render(
    tournament_name=dfMeta.loc[dfMeta["Key"] == "Tournament Long Name", "Value"].values[
        0
    ],
    div1_list = teamList[1],
    div2_list = teamList[2],
    rg_div1_list=rg_html[1],
    rg_div2_list=rg_html[2],
    rd_div1_list=divAwards[1]["Robot Design"],
    rd_div2_list=divAwards[2]["Robot Design"],
    rd_this_them="This team" if awardCounts["Robot Design"] == 1 else "These teams",
    ip_div1_list=divAwards[1]["Innovation Project"],
    ip_div2_list=divAwards[2]["Innovation Project"],
    ip_this_them=(
        "This team" if awardCounts["Innovation Project"] == 1 else "These teams"
    ),
    cv_div1_list=divAwards[1]["Core Values"],
    cv_div2_list=divAwards[2]["Core Values"],
    cv_this_them="This team" if awardCounts["Core Values"] == 1 else "These teams",
    ja_count=judgesAwardTotalCount,
    ja_list=judgesAwardHtml[1] + judgesAwardHtml[2],
    ja_go_goes=(
        "The Judges Awards go to teams"
        if judgesAwardTotalCount > 1
        else "The Judges Award goes to team"
    ),
    champ_div1_list=divAwards[1]["Champions"],
    champ_div2_list=divAwards[2]["Champions"],
    adv_div1_list=advancingHtml[1],
    adv_div2_list=advancingHtml[2],
)

# print(out_text)
# path construction changed here for absolute paths -ac
with open(
    os.path.join(dir_path,dfMeta.loc[dfMeta["Key"] == "Completed Script File", "Value"].values[0]), "w"
) as fh:
    fh.write(out_text)

print(Fore.GREEN + 
    "All done! The script hase been saved as "
    + dfMeta.loc[dfMeta["Key"] == "Completed Script File", "Value"].values[0]
    + "."
)
print(Fore.LIGHTWHITE_EX + "It is saved in the same folder with the other OJS files.")
print(Fore.LIGHTWHITE_EX + "Double-click the script file to view it on this computer,")
print(Fore.LIGHTWHITE_EX + "or email it to yourself and view it on a phone or tablet.")
input(Fore.LIGHTWHITE_EX + "Press enter to quit...")
