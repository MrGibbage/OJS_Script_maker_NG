# This program will run a pre-check on an OJS folder to make
# sure all of the files are set up correctly
# These are the checks the program will validate
import os, sys, re

# pip install pandas
import pandas as pd

from openpyxl.workbook.protection import WorkbookProtection
from openpyxl import load_workbook, Workbook

# pip install print-color
from print_color import print

dir_path = os.path.dirname(os.path.realpath(__file__))
dataframes = {}
base_file_name: str = ""

def open_dataframe(ojsfile: str, sheetname: str, division: str, cols: list[str], header_row: int):
    try:
        print(
            f"Opening dataframe. Now checking the {ojsfile}, {sheetname} worksheet ({division}). Looking for these columns: {cols}",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        dataframe = pd.read_excel(
            ojsfile,
            sheet_name=sheetname,
            header=header_row,
            usecols=cols,
        )
        print(dataframe)
        return dataframe
    except Exception as e:
        print(f"There was an error reading from the workbook {e}", tag=f'{tourn}', tag_color="red", color="red")
        input("Press enter to quit...")
        sys.exit(1)


def check_column_for_null_values(data: pd.DataFrame, division: str, cols: list[str]):
    print(f'Here are the columns (null): {cols}')
    try:
        for c in cols:
            if not(data[c].isnull().values.all()):
                print(
                    f"Found a non NaN value in the {division} {c} column.",
                    tag=f'{tourn}',
                    tag_color="red",
                    color="red",
                )
                input("Press enter to quit...")
                sys.exit(1)
            print(
                f"Checked for expected NaN values in the {division} {c} column -- OK (found NaN)",
                tag=f'{tourn}',
                tag_color="white",
                color="white",
            )
    except Exception as e:
        print(f"There was an error {e}", tag=f'{tourn}', tag_color="red", color="red")
        input("Press enter to quit...")
        sys.exit(1)


def check_column_for_valid_values(data: pd.DataFrame, division: str, v: int, cols: list[str]):
    print(f'Here are the columns (valid): {cols}')
    try:
        for c in cols:
            if (data[c] != v).all():
                print(
                    f"Found an unexpected value ({v}) in the {division}, {c} column.",
                    tag=f'{tourn}',
                    tag_color="red",
                    color="red",
                )
                input("Press enter to quit...")
                sys.exit(1)
            else:
                print(
                    f"Checked for {v} values in the {division}, {c} column -- OK",
                    tag=f'{tourn}',
                    tag_color="white",
                    color="white",
                )
    except Exception as e:
        print(f"There was an error {e}", tag=f'{tourn}', tag_color="red", color="red")
        input("Press enter to quit...")
        sys.exit(1)

def check_dataframe_for_valid_team_numbers_and_names(data: pd.DataFrame, division: str):
    # print("check_dataframe_for_valid_team_numbers_and_names")
    # print(data)
    for c in ["Team Number", "Team Name"]:
        try:
            if data[c].isnull().values.any():
                print(
                    f"Found a null in the {division} OJS Team Numbers\n{data[c]}",
                    tag=f'{tourn}',
                    tag_color="red",
                    color="red",
                )
                sys.exit(1)
            print(
                f"Checked for null values in the {division} {c} column -- OK",
                tag=f'{tourn}',
                tag_color="white",
                color="white",
            )
        except Exception as e:
            print(f"There was a team number/name error {e}", tag=f'{tourn}', tag_color="red", color="red")
            input("Press enter to quit...")
            sys.exit(1)


def run_checks(d):
    print("Getting a directory listing", tag=f'{tourn}', tag_color="white", color="white")
    try:
        directory_list: list[str] = os.listdir(d)
        print(
            f"Found {len(directory_list)} files in the folder.",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
    except Exception as e:
        print(
            f"*-*-* Could not get a directory list. We got this error: {e}",
            tag=f'{tourn}',
            tag_color="red",
            color="red",
        )
        input("Press enter to quit...")
        sys.exit(1)

    print("Looking for the OJS spreadsheets", tag=f'{tourn}', tag_color="white", color="white")
    xlsm_files: list[str] = [s for s in directory_list if s.endswith(".xlsm")]

    if len(xlsm_files) > 0:
        print(
            f"Found these spreadsheets: {xlsm_files}",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
    else:
        print(
            f"*-*-* Did not find any spreadsheets. Be sure to run this program from the same folder where the spreadsheets are saved.",
            tag=f'{tourn}',
            tag_color="red",
            color="red",
        )
        input("Press enter to quit...")
        sys.exit(1)

    if len(xlsm_files) > 2:
        print(
            f"*-*-* There should be only one or two spreadsheets in the folder. We found {len(xlsm_files)} spreadsheets.",
            tag=f'{tourn}',
            tag_color="red",
            color="red",
        )
        print(
            f"*-*-* Perhaps you have one or more of the spreadsheets open, "
            "which will add temprary files with .xlsm extensions.",
            tag=f'{tourn}',
            tag_color="red",
            color="red",
        )
        input("Press enter to quit...")
        sys.exit(1)

    regex: str = r"^([0-9]{4}-vadc-fll-challenge-.*)(-ojs-)(.*)-(div[1,2])(.xlsm)$"

    for f in xlsm_files:
        print(f"Checking {f}", tag=f'{tourn}', tag_color="white", color="white")
        m = re.search(regex, f)
        try:
            print(m.groups(), tag=f'{tourn}', tag_color="white", color="white")
            if not ((m.group(2)) == "-ojs-" and m.group(4).startswith("div")):
                print(
                    f"*-*-* {f} is not named correctly",
                    tag=f'{tourn}',
                    tag_color="red",
                    color="red",
                )
                input("Press enter to quit...")
                sys.exit(1)
        except Exception as e:
            print(
                f"*-*-* {f} is not named correctly",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            if (f[:1]=="~"):
                print(
                    f"It looks like {f} is a temporary file, which suggests you may have an OJS file open in Excel",
                    tag=f'{tourn}',
                    tag_color="red",
                    color="red",
                )
            print(
                "OJS files should be named with this pattern (all lowercase, "
                "no special characters or spaces)"
            )
            print("year-vadc-fll-challenge-season_name-ojs-tournament_name-div#.xlsm")
            print("For example, 2024-vadc-fll-challenge-submerged-ojs-norview-div1.xlsm")
            print("Where vadc-fll-challenge is always the same")
            input("Press enter to quit...")
            sys.exit(1)
    divlist: list[str] = ["div1", "div2"]

    div: list[str] = []
    if len(xlsm_files) == 2:
        print(
            "Checking to see if there is a div1 and a div2",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        m0 = re.search(regex, xlsm_files[0])
        m1 = re.search(regex, xlsm_files[1])
        div.append(m0.group(4))
        div.append(m1.group(4))
        print("Found", div[0], div[1], tag=f'{tourn}', tag_color="white", color="white")
        if "div1" in div and "div2" in div:
            print(
                "Good. There are two divisions: div1 and div2",
                tag=f'{tourn}',
                tag_color="white",
                color="white",
            )
        else:
            print(
                f"*-*-* There should be two different divisions",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)
        if m0.group(1) != m1.group(1):
            print(
                f"*-*-* {m0.group(1)} does not match {m1.group(1)}",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)
        if m0.group(3) != m1.group(3):
            print(
                f"*-*-* {m0.group(3)} does not match {m1.group(3)}",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)
        if m0.group(5) != m1.group(5):
            print(
                f"*-*-* {m0.group(5)} does not match {m1.group(5)}",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)
        base_file_name = m0.group(1) + m0.group(2) + m0.group(3) + "-"
        print(
            "Files appear to be named correctly.",
            tag="success",
            tag_color="green",
            color="white",
        )

    print(f"Base file name: {base_file_name}", tag=f'{tourn}', tag_color="white", color="white")

    if len(xlsm_files) == 1:
        print(
            "Checking to see if it is a div1 or div2",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        m = re.search(regex, xlsm_files[0])
        div = m.group(4)
        print("Found", div, tag=f'{tourn}', tag_color="white", color="white")
        if div in divlist:
            print(
                f"Good. Found {m.group(4)}", tag="success", tag_color="green", color="white"
            )
        else:
            print(
                f"*-*-* Neither div1 or div2; found {div}",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)
        divlist.remove("div1" if div == "div2" else "div2")
        print(
            "File appears to be named correctly.",
            tag="success",
            tag_color="green",
            color="white",
        )

    print(divlist, tag=f'{tourn}', tag_color="white", color="white")

    # Check to see if any of the spreadsheets are open in Excel
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        print(
            f"Checking {this_ojs_filename} to see if it is open",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        try:
            # https://stackoverflow.com/questions/6825994/check-if-a-file-is-open-in-python
            os.rename(this_ojs_filename, this_ojs_filename)
            print(
                f"{this_ojs_filename} is correctly closed.",
                tag=f'{tourn}',
                tag_color="white",
                color="white",
            )
        except:
            print(
                f"{this_ojs_filename} seems to be open. Do you have this file open in Excel?",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            sys.exit(1)

#### RESULTS AND RANKINGS
    print(
        f"Now checking the Results and Rankings worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "Results and Rankings", division, [
                "Team Number",
                "Team Name",
                "Max Robot Game Score",
                "Robot Game Rank",
                "Award",
                "Advance?",
            ], 1)
        print(
            "There should be no errors or warnings. All rows below should have team data.\n"
            "Max Robot game scores should be all 0\n"
            "Robot game ranks should be all 1\n"
            "Award and Advance? should be all NaN\n",
            "All team numbers should be integers and there should not be "
            "any team names NaN",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for Results and Rankings",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        print(dataframes[division])

    check_dataframe_for_valid_team_numbers_and_names(dataframes[division], division)
    check_column_for_valid_values(dataframes[division], division, 0, ["Max Robot Game Score"])
    check_column_for_valid_values(dataframes[division], division, 1, ["Robot Game Rank"])
    check_column_for_null_values(dataframes[division], division, ["Award", "Advance?"])


#### ROBOT GAME SCORES

    print(
        f"Now checking the Robot Game Scores worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "Robot Game Scores", division, [
                "Team Number",
                "Team Name",
                "Robot Game 1 Score",
                "Robot Game 2 Score",
                "Robot Game 3 Score",
                "Highest Robot Game Score",
            ], 0)

        print(
            "There should be no errors or warnings. All rows below should have team data.\n"
            "Robot game scores should be all NaN\n"
            "Award and Advance? should be all NaN\n",
            "All team numbers should be integers and there should not be "
            "any team names NaN\n"
            "Highest Robot Game Score should be all zero",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for Robot Game Scores",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        print(dataframes[division])

    check_dataframe_for_valid_team_numbers_and_names(dataframes[division], division)
    check_column_for_null_values(dataframes[division], division, ["Robot Game 1 Score", "Robot Game 2 Score", "Robot Game 3 Score"])
    check_column_for_valid_values(dataframes[division], division, 0, ["Highest Robot Game Score"])


#### INNOVATION PROJECT

    print(
        f"Now checking the Innovation Project Input worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "Innovation Project Input", division, [
                "Team Number",
                "Team Name",
                "Identify - Define",
                "Identify - Research (CV)",
                "Design - Plan",
                "Design - Teamwork (CV)",
                "Create - Innovation (CV)",
                "Create - Model",
                "Iterate - Sharing",
                "Iterate - Improvement",
                "Communicate - Impact (CV)",
                "Communicate - Fun (CV)",
                "Innovation Project Score",
                "Innovation Project Rank",
            ], 0)

        print(
            "There should be no errors or warnings. All rows below should have team data.\n"
            "Innovation Project scores should be all NaN\n"
            "All team numbers should be integers and there should not be "
            "any team names NaN\n",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for Innovation Project",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        print(dataframes[division])
        check_dataframe_for_valid_team_numbers_and_names(dataframes[division], division)
        check_column_for_null_values(dataframes[division], division, [
                "Identify - Define",
                "Identify - Research (CV)",
                "Design - Plan",
                "Design - Teamwork (CV)",
                "Create - Innovation (CV)",
                "Create - Model",
                "Iterate - Sharing",
                "Iterate - Improvement",
                "Communicate - Impact (CV)",
                "Communicate - Fun (CV)",
                "Innovation Project Rank",
            ])
        check_column_for_valid_values(dataframes[division], division, 0, ["Innovation Project Score"])



#### ROBOT DESIGN

    print(
        f"Now checking the Robot Design Input worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "Robot Design Input", division, [
                "Team Number",
                "Team Name",
                "Identify - Strategy",
                "Identify - Research (CV)",
                "Design - Ideas (CV)",
                "Design - Building/Coding",
                "Create - Attachments",
                "Create - Code/ Sensors",
                "Iterate - Testing",
                "Iterate - Improvements (CV)",
                "Communicate - Impact (CV)",
                "Communicate - Fun (CV)",
                "Robot Design Score",
                "Robot Design Rank",
            ], 0)

        print(
            "There should be no errors or warnings. All rows below should have team data.\n"
            "Robot Design scores should be all NaN\n"
            "All team numbers should be integers and there should not be "
            "any team names NaN\n",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for Robot Design",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        print(dataframes[division])
        check_dataframe_for_valid_team_numbers_and_names(dataframes[division], division)
        check_column_for_null_values(dataframes[division], division, [
                "Identify - Strategy",
                "Identify - Research (CV)",
                "Design - Ideas (CV)",
                "Design - Building/Coding",
                "Create - Attachments",
                "Create - Code/ Sensors",
                "Iterate - Testing",
                "Iterate - Improvements (CV)",
                "Communicate - Impact (CV)",
                "Communicate - Fun (CV)",
                "Robot Design Rank",
            ])
        check_column_for_valid_values(dataframes[division], division, 0, ["Robot Design Score"])

#### CORE VALUES

    print(
        f"Now checking the Core Values Input worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "Core Values Input", division, [
                "Team Number",
                "Team Name",
                "Identify - Research (CV-IP)",
                "Design - Teamwork (CV-IP)",
                "Create - Innovation (CV-IP)",
                "Communicate - Impact (CV-IP)",
                "Communicate - Fun (CV-IP)",
                "Identify - Research (CV-RD)",
                "Design - Ideas (CV-RD)",
                "Iterate - Improvements (CV-RD)",
                "Communicate - Impact (CV-RD)",
                "Communicate - Fun (CV-RD)",
                "Gracious Professionalism 1",
                "Gracious Professionalism 2",
                "Gracious Professionalism 3",
                "Gracious Professionalism Total",
                "Gracious Professionalism Score",
                "Core Values Score",
                "Core Values Rank",
            ], 0)

        print(
            "There should be no errors or warnings. All rows below should have team data.\n"
            "Core Values scores should be all NaN\n"
            "All team numbers should be integers and there should not be "
            "any team names NaN\n",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for Core Values",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        print(dataframes[division])
        check_dataframe_for_valid_team_numbers_and_names(dataframes[division], division)
        check_column_for_null_values(dataframes[division], division, [
                "Gracious Professionalism 1",
                "Gracious Professionalism 2",
                "Gracious Professionalism 3",
                "Gracious Professionalism Score",
                "Core Values Rank",
            ])
        check_column_for_valid_values(dataframes[division], division, 0, [
                "Identify - Research (CV-IP)",
                "Design - Teamwork (CV-IP)",
                "Create - Innovation (CV-IP)",
                "Communicate - Impact (CV-IP)",
                "Communicate - Fun (CV-IP)",
                "Identify - Research (CV-RD)",
                "Design - Ideas (CV-RD)",
                "Iterate - Improvements (CV-RD)",
                "Communicate - Impact (CV-RD)",
                "Communicate - Fun (CV-RD)",
                "Gracious Professionalism Total",
                "Core Values Score",
            ])

#### AWARD LIST

    print(
        f"Now checking the AwardList worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "AwardList", division, [
                "Award",
            ], 0)

        print(
            "There should be no errors or warnings. All judged awards for the tournament should be listed.",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for AwardList",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        print(dataframes[division])
        if len(dataframes[division]) < 4:
            print(
                f"{division} WARNING Award List is less than 4, indicating a possible mistake",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )

#### META

    print(
        f"Now checking the Meta worksheet",
        tag=f'{tourn}',
        tag_color="yellow",
        color="yellow",
    )
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        dataframes[division] = open_dataframe(this_ojs_filename, "Meta", division, [
                "Key", "Value"
            ], 0)

        print(
            "There should be no errors or warnings.",
            tag=f'{tourn}',
            tag_color="white",
            color="white",
        )
        print(
            f"{division} dataframe for Meta",
            tag=f'{tourn}',
            tag_color="yellow",
            color="yellow",
        )
        if len(dataframes[division]) != 13:
            print(
                f"{division} does not seem to have the right keys entered",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)

###### Workbook Protection
    for division in divlist:
        this_ojs_filename = d + "\\" + base_file_name + division + ".xlsm"
        workbook: Workbook = load_workbook(this_ojs_filename)
        print(f'{this_ojs_filename} protection')
        for ws in workbook.worksheets:
            if ws.protection.sheet:
                print(f'{ws} is protected')
            else:
                print(
                    f"{this_ojs_filename} {ws} is not protected",
                    tag='ERROR',
                    tag_color="red",
                    color="red",
                )
                input("Press enter to quit...")
                sys.exit(1)
        
        tapi_sheet = workbook["Team and Program Information"]
        if tapi_sheet.cell(row=1, column=9).value is not None:
            print(
                f"{division} Password entered on the Team and Program Information worksheet: {this_ojs_filename}",
                tag=f'{tourn}',
                tag_color="red",
                color="red",
            )
            input("Press enter to quit...")
            sys.exit(1)





# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

tourn = input("Which tournament to check? (press ENTER to check all) ")
if tourn == "":
    for directory in os.listdir(dir_path + "\\tournaments"):
        print(directory)
        tourn = directory
        run_checks(dir_path + "\\tournaments\\" + directory)
else:
    d = dir_path + "\\tournaments\\" + tourn
    run_checks(d)

#
# TODO
# check columns in tables/tabs
# check if all awards are set up correctly
# check if sheets are protected with a password
# check if password is entered
# check meta information
