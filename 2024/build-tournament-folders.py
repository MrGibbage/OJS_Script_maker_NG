# To create an executable file, run
# .venv\Scripts\pyinstaller.exe -F 2024\build-tournament-folders.py
# Then copy the build-tournament-folders.exe file from dist to 2024
#
import os, sys, re, time
import shutil
import warnings

# pip install openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import (
    column_index_from_string,
    get_column_letter,
    coordinate_from_string,
)
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule, CellIsRule, FormulaRule

# pip install print-color
from print_color import print

# pip install pywin32
import win32com.client

# pip install pandas
import pandas as pd

# pip install xlwings
import xlwings


# from https://stackoverflow.com/questions/56923379/how-to-read-an-existing-worksheet-table-with-openpyxl
def read_excel_table(sheet, table_name):
    """
    This function will read an Excel table
    and return a tuple of columns and data

    This function assumes that tables have column headers
    :param sheet: the sheet
    :param table_name: the name of the table
    :return: columns (list) and data (dict)
    """
    table = sheet.tables[table_name]
    table_range = table.ref

    table_head = sheet[table_range][0]
    table_data = sheet[table_range][1:]

    columns = [column.value for column in table_head]
    data = {column: [] for column in columns}

    for row in table_data:
        row_val = [cell.value for cell in row]
        for key, val in zip(columns, row_val):
            data[key].append(val)

    return columns, data


# creates the subfolders for each tournament
def create_folder(newpath):
    if os.path.exists(newpath):
        return
    try:
        os.makedirs(newpath)
        print(
            f"Created folder: {newpath}",
            tag="ok",
            tag_color="white",
            color="white",
        )
    except Exception as e:
        print(
            f"Could not create directory: {newpath}",
            tag="error",
            tag_color="red",
            color="red",
        )


# copies the files into the tournament folders
def copy_files(item: pd.Series):
    for filename in extrafilelist:
        try:
            newpath = dir_path + "\\tournaments\\" + item["Short Name"] + "\\"
            shutil.copy(filename, newpath)
        except Exception as e:
            print(
                f"Could not copy file: {filename} to {newpath}\n{e}",
                tag="error",
                tag_color="red",
                color="red",
            )
    # next copy the template files
    ojsfilelist.clear()
    if item["D1_OJS"] is not None:
        ojsfilelist.append(item["D1_OJS"])
    if item["D2_OJS"] is not None:
        ojsfilelist.append(item["D2_OJS"])
    for filename in ojsfilelist:
        try:
            shutil.copy(
                template_file,
                dir_path + "\\tournaments\\" + item["Short Name"] + "\\" + filename,
            )
        except Exception as e:
            print(
                f'Could not copy OJS file: {template_file} to {dir_path + "\\tournaments\\" + item["Short Name"] + "\\" + filename}',
                tag="error",
                tag_color="red",
                color="red",
            )


# edits the OJS spreadsheets with the correct tournament information
def set_up_tapi_worksheet(tournament: pd.Series):
    # open the OJS workbook
    for d in ["D1", "D2"]:
        print(
            f"Setting up Tournament Team and Program Information for {tournament["Short Name"]} {d}"
        )
        divassignees: pd.DataFrame = dfAssignments[
            (dfAssignments["Tournament"] == tournament["Short Name"])
            & (dfAssignments["Div"] == d)
        ]
        print(divassignees)
        print(
            f'There are {len(divassignees)} teams in this {d} {tournament["Short Name"]} tournament'
        )
        if len(divassignees.index) > 0:
            try:
                # print(divassignees)
                ojsfile = (
                    dir_path
                    + "\\tournaments\\"
                    + tournament["Short Name"]
                    + "\\"
                    + tournament[d + "_OJS"]
                )
                ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
                ws = ojs_book["Team and Program Information"]
                table: Table = ws.tables["OfficialTeamList"]
                table_range: str = table.ref  # should return a string C2:F3
                start_cell = table_range.split(":")[0]  # should return 'C2'
                coords = coordinate_from_string(start_cell)
                start_col = column_index_from_string(coords[0])
                start_index = divassignees.index[0]
                colonPos = (table.ref).find(":")
                print(
                    f'About to resize the table {table.ref}, {table.ref[:colonPos]}, {re.sub(r'\d', '', table.ref[colonPos + 1])}, {str(len(divassignees) + 2)}'
                )
                table.ref = (
                    table.ref[:colonPos]
                    + ":"
                    + re.sub(r"\d", "", table.ref[colonPos + 1])
                    + str(len(divassignees) + 2)
                )
                for i, row in divassignees.iterrows():
                    # cell = f'{get_column_letter(start_col)}{i + 3 - start_index}'
                    # print(f'Cell: {cell}, setting value {row['Team #']}')
                    # ws[cell] = row['Team #']
                    thisrow = i + 3 - start_index
                    ws.cell(row=thisrow, column=start_col).value = row["Team #"]
                    ws.cell(row=thisrow, column=start_col + 1).value = row["Team Name"]
                    ws.cell(row=thisrow, column=start_col + 2).value = row["Coach Name"]

                # Save the workbook
                print(f"Saving workbook. OfficialTeamList ref: {table.ref}")
                ojs_book.save(ojsfile)
                # ojs_book.close()
            except Exception as e:
                print(
                    f"There was an error: {e}",
                    tag="error",
                    tag_color="red",
                    color="red",
                )


def set_up_award_worksheet(tournament: pd.Series, judge_awards: int):
    for d in ["D1", "D2"]:
        print(f"Setting up Tournament Awards for {tournament["Short Name"]} {d}")
        divawards: pd.DataFrame = dfAwards[
            (dfAwards["Tournament"] == tournament["Short Name"])
            & (dfAwards["Div"] == d)
        ]
        divawards = divawards.transpose()
        divawards = divawards.reset_index()
        divawards.columns = ["Award", "Count"]
        divawards.drop(0, inplace=True)
        divawards.drop(1, inplace=True)
        divawards.drop(divawards[divawards["Award"] == "ADV"].index, inplace=True)
        divawards.drop(
            divawards[divawards["Award"] == "JudgedAwards"].index, inplace=True
        )
        divawards.drop(
            divawards[divawards["Award"] == "PerfAwards"].index, inplace=True
        )
        divawards.drop(divawards[divawards["Award"] == "AwardTot"].index, inplace=True)

        print(divawards)
        rg_awards: pd.DataFrame = divawards[
            (divawards["Award"].str.startswith("RG")) & (divawards["Count"] == 1)
        ]
        print(rg_awards)

        # Robot Game
        if tournament[d + "_OJS"] is None:
            print(f"Nothing for tournament[{d}]")
            continue

        if len(rg_awards.index) > 0:
            print("Looking for robot game awards")
            ojsfile = (
                dir_path
                + "\\tournaments\\"
                + tournament["Short Name"]
                + "\\"
                + tournament[d + "_OJS"]
            )
            ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
            ws = ojs_book["AwardList"]
            table: Table = ws.tables["RobotGameAwards"]
            table_range: str = table.ref
            print(f"table_range: {table_range}")
            start_cell = table_range.split(":")[0]
            coords = coordinate_from_string(start_cell)
            start_col = column_index_from_string(coords[0])
            # start_index = divassignees.index[0]
            table.ref = table.ref[:-1] + str(len(rg_awards) + 1)
            next_row = 2
            for i, row in rg_awards.iterrows():
                print(f"New row, looking for RG {row}")
                award = row["Award"].replace("RG", "Robot Game ")
                award = (
                    award.replace("1", "1st Place")
                    .replace("2", "2nd Place")
                    .replace("3", "3rd Place")
                )
                if award[:2] == "Ro" and row["Count"] == 1:
                    ws.cell(row=next_row, column=start_col).value = award
                    print(
                        f"Adding the robot game award: {award} to cell(row = {next_row}, column = {start_col})"
                    )
                    next_row = next_row + 1

        # Other judged awards
        divawards.drop(divawards[divawards["Award"] == "RG1"].index, inplace=True)
        divawards.drop(divawards[divawards["Award"] == "RG2"].index, inplace=True)
        divawards.drop(divawards[divawards["Award"] == "RG3"].index, inplace=True)
        divawards = divawards[(divawards["Count"] > 0)]
        divawards = divawards.reset_index(drop=True)
        print("Writing divawards to the AwardList table")
        print(divawards)

        if len(divawards.index) > 0:
            # ojsfile = dir_path + "\\tournaments\\" + tournament["Short Name"] + "\\" + tournament[d + "_OJS"]
            # ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
            # ws = ojs_book['AwardList']
            table: Table = ws.tables["AwardList"]
            table_range: str = table.ref  # should return a string A1:B2
            start_cell = table_range.split(":")[0]  # should return 'A1'
            coords = coordinate_from_string(start_cell)
            start_col = column_index_from_string(coords[0])
            # start_index = divassignees.index[0]
            table.ref = table.ref[:-1] + str(len(divawards) + 1 + judge_awards)
            for i, row in divawards.iterrows():
                award = (
                    row["Award"]
                    .replace("Champ", "Champions ")
                    .replace("RD", "Robot Design ")
                    .replace("CV", "Core Values ")
                    .replace("RG", "Robot Game ")
                    .replace("IP", "Innovation Project ")
                )
                award = (
                    award.replace("1", "1st Place")
                    .replace("2", "2nd Place")
                    .replace("3", "3rd Place")
                )
                ws.cell(row=i + 2, column=start_col).value = award
            next_row = i + 1
            for i in range(judge_awards):
                ws.cell(row=i + 2 + next_row, column=start_col).value = (
                    "Judges Award " + str(i + 1)
                )

            ojs_book.save(ojsfile)


def set_up_meta_worksheet(tournament: pd.Series, yr: int, seasonName: str):
    for d in ["D1", "D2"]:
        print(f"Setting up meta worksheet for {tournament["Short Name"]} {d}")
        print(f"Season year: {yr}; season name: {seasonName}")
        print(tournament)
        if tournament[d + "_OJS"] is not None:
            ojsfile = (
                dir_path
                + "\\tournaments\\"
                + tournament["Short Name"]
                + "\\"
                + tournament[d + "_OJS"]
            )
            print(f"Loading ojs workbook {ojsfile}")
            ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
            scriptfile = (
                str(yr)
                + "-"
                + seasonName
                + "-"
                + tournament["Short Name"]
                + "-script.html"
            )
            divawards: pd.DataFrame = dfAwards[
                (dfAwards["Tournament"] == tournament["Short Name"])
                & (dfAwards["Div"] == d)
            ]
            divawards = divawards.transpose()
            divawards = divawards.reset_index()
            divawards.columns = ["Award", "Count"]
            ws = ojs_book["Meta"]
            # Tournament Year
            ws.cell(row=2, column=2).value = yr
            # FLL Season Title
            ws.cell(row=3, column=2).value = seasonName
            # Tournament Short Name
            ws.cell(row=4, column=2).value = tournament["Short Name"]
            # Tournament Short Name
            ws.cell(row=5, column=2).value = tournament["Long Name"]
            # Completed Script File
            ws.cell(row=6, column=2).value = scriptfile
            # Division
            ws.cell(row=7, column=2).value = 1 if d == "D1" else 2
            # Robot Game
            ws.cell(row=8, column=2).value = divawards.loc[
                (divawards["Award"] == "RG1")
                | (divawards["Award"] == "RG2")
                | (divawards["Award"] == "RG3"),
                "Count",
            ].sum()
            # Innovation Project
            ws.cell(row=9, column=2).value = divawards.loc[
                (divawards["Award"] == "IP1")
                | (divawards["Award"] == "IP2")
                | (divawards["Award"] == "IP3"),
                "Count",
            ].sum()
            # Core Values
            ws.cell(row=10, column=2).value = divawards.loc[
                (divawards["Award"] == "CV1")
                | (divawards["Award"] == "CV2")
                | (divawards["Award"] == "CV3"),
                "Count",
            ].sum()
            # Robot Design
            ws.cell(row=11, column=2).value = divawards.loc[
                (divawards["Award"] == "RD1")
                | (divawards["Award"] == "RD2")
                | (divawards["Award"] == "RD3"),
                "Count",
            ].sum()
            # Champions
            ws.cell(row=12, column=2).value = divawards.loc[
                (divawards["Award"] == "Champ1")
                | (divawards["Award"] == "Champ2")
                | (divawards["Award"] == "Champ3"),
                "Count",
            ].sum()
            # Advancing
            ws.cell(row=13, column=2).value = divawards.loc[
                divawards["Award"] == "ADV", "Count"
            ].values[0]
            # Judges
            ws.cell(row=14, column=2).value = (
                tournament["Judges_1"]
                + tournament["Judges_2"]
                + tournament["Judges_3"]
                + tournament["Judges_4"]
                + tournament["Judges_5"]
                + tournament["Judges_6"]
            )
            ojs_book.save(ojsfile)


def copy_team_numbers(
    source_sheet: Worksheet, target_sheet: Worksheet, target_start_row: int
):
    source_start_row = 3

    column = "A"
    last_row = 0
    print(f"Copying team numbers to {target_sheet}")
    # Iterate through the rows in the specified column
    for row in range(1, source_sheet.max_row + 1):
        if source_sheet[f"{column}{row}"].value is not None:
            last_row = row
    team_count = last_row - source_start_row + 1
    col = 1  # Team number is always in column 1 ('A')
    # itterate over the source rows. The dest row may not always align with
    # the source row. Some sheets start on row 3, some start on 2
    current_target_row = target_start_row
    for row in range(source_start_row, source_start_row + team_count + 1):
        cell_value = source_sheet.cell(row=row, column=col).value
        target_sheet.cell(row=current_target_row, column=col).value = cell_value
        current_target_row += 1


def protect_worksheets(tournament: pd.Series):
    for d in ["D1", "D2"]:
        if tournament[d + "_OJS"] is None:
            print(f'*-*-*-* No division {d} to check for {tournament["Short Name"]}')
            continue
        print(f'Protecting {d} {tournament["Short Name"]}')
        ojsfile = (
            dir_path
            + "\\tournaments\\"
            + tournament["Short Name"]
            + "\\"
            + tournament[d + "_OJS"]
        )
        ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
        for ws in ojs_book.worksheets:
            print(f"{ws}")
            ws.protection.sheet = True

        ojs_book.save(ojsfile)


def resize_worksheets(tournament: pd.Series):
    worksheetNames = [
        "Robot Game Scores",
        "Innovation Project Input",
        "Robot Design Input",
        "Core Values Input",
        "Results and Rankings",
    ]
    worksheetTables = [
        "RobotGameScores",
        "InnovationProjectResults",
        "RobotDesignResults",
        "CoreValuesResults",
        "TournamentData",
    ]
    worksheet_start_row = [2, 2, 2, 2, 3]
    for d in ["D1", "D2"]:
        sheet_tables = zip(worksheetNames, worksheetTables, worksheet_start_row)
        if tournament[d + "_OJS"] is None:
            print(f'*-*-*-* No division {d} to check for {tournament["Short Name"]}')
            continue

        print(f'Resizing {d} {tournament["Short Name"]}')
        ojsfile = (
            dir_path
            + "\\tournaments\\"
            + tournament["Short Name"]
            + "\\"
            + tournament[d + "_OJS"]
        )
        ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
        divassignees: pd.DataFrame = dfAssignments[
            (dfAssignments["Tournament"] == tournament["Short Name"])
            & (dfAssignments["Div"] == d)
        ]

        # copy the team number data to each of the worksheets
        for s, t, r in sheet_tables:
            ws = ojs_book[s]
            tapi_sheet = ojs_book["Team and Program Information"]
            # first, copy the team numbers over
            copy_team_numbers(
                source_sheet=tapi_sheet, target_sheet=ws, target_start_row=r
            )

        # Resize the tables
        sheet_tables = zip(worksheetNames, worksheetTables, worksheet_start_row)
        for s, t, r in sheet_tables:
            print(f"{s}, {r}, {t}")
            ws = ojs_book[s]
            table: Table = ws.tables[t]
            table_range: str = table.ref
            start_cell = table_range.split(":")[0]
            start_row = int(re.findall(r"\d", start_cell)[0])
            colonPos = (table.ref).find(":")
            print(
                f'Resizing the {d} {t} table {table.ref}, {table.ref[:colonPos]}, {re.sub(r'\d', '', table.ref[colonPos + 1])}, {str(len(divassignees) + 2)}.'
            )
            table.ref = (
                table.ref[:colonPos]
                + ":"
                + re.sub(r"\d", "", table.ref[colonPos + 1])
                + str(start_row + len(divassignees))
            )
            # print(f'New table.ref = {table.ref}')
            ws.delete_rows(idx=start_row + len(divassignees) + 1, amount=200)
            # ws.protection.sheet = True

        ojs_book.save(ojsfile)


def add_conditional_formats(tournament: pd.Series):
    # Create fill
    greenAwardFill = PatternFill(
        start_color="00B050", end_color="00B050", fill_type="solid"
    )
    greenAdvFill = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    rgGoldFill = PatternFill(
        start_color="C9B037", end_color="C9B037", fill_type="solid"
    )
    rgSilverFill = PatternFill(
        start_color="D7D7D7", end_color="D7D7D7", fill_type="solid"
    )
    rgBronzeFill = PatternFill(
        start_color="AD8A56", end_color="AD8A56", fill_type="solid"
    )
    for d in ["D1", "D2"]:
        if tournament[d + "_OJS"] is None:
            continue
        print(f'Adding conditional formats to {d} {tournament["Short Name"]}')
        ojsfile = (
            dir_path
            + "\\tournaments\\"
            + tournament["Short Name"]
            + "\\"
            + tournament[d + "_OJS"]
        )
        ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)
        ws = ojs_book["Results and Rankings"]
        # Award
        ws.conditional_formatting.add(
            "O2",
            FormulaRule(
                formula=["COUNTA(AwardList!$A$2:$A$35)=COUNTA($O$3:$O$288)"],
                stopIfTrue=False,
                fill=greenAwardFill,
            ),
        )
        # Advancing
        ws.conditional_formatting.add(
            "P2",
            FormulaRule(
                formula=['COUNTIF($P:$P,"Yes")=Meta!$B$13'],
                stopIfTrue=False,
                fill=greenAdvFill,
            ),
        )
        # Robot Game Gold
        ws.conditional_formatting.add(
            "I1:I100",
            FormulaRule(
                formula=[
                    'AND(I1=1,IF(VLOOKUP("Robot Game 1st Place",AwardList!$C$2:$C$7,1,FALSE)="Robot Game 1st Place", TRUE, FALSE))'
                ],
                stopIfTrue=False,
                fill=rgGoldFill,
            ),
        )
        # Robot Game Silver
        ws.conditional_formatting.add(
            "I1:I100",
            FormulaRule(
                formula=[
                    'AND(I1=2,IF(VLOOKUP("Robot Game 2nd Place",AwardList!$C$2:$C$7,1,FALSE)="Robot Game 2nd Place", TRUE, FALSE))'
                ],
                stopIfTrue=False,
                fill=rgSilverFill,
            ),
        )
        # Robot Game Bronze
        ws.conditional_formatting.add(
            "I1:I100",
            FormulaRule(
                formula=[
                    'AND(I1=3,IF(VLOOKUP("Robot Game 3rd Place",AwardList!$C$2:$C$7,1,FALSE)="Robot Game 3rd Place", TRUE, FALSE))'
                ],
                stopIfTrue=False,
                fill=rgBronzeFill,
            ),
        )
        ojs_book.save(ojsfile)


def hide_worksheets(tournament: pd.Series):
    worksheetNames = ["Data Validation", "Meta", "AwardList"]
    for d in ["D1", "D2"]:
        if tournament[d + "_OJS"] is None:
            continue
        print(f'Hiding worksheets in {d} {tournament["Short Name"]}')
        ojsfile = (
            dir_path
            + "\\tournaments\\"
            + tournament["Short Name"]
            + "\\"
            + tournament[d + "_OJS"]
        )
        ojs_book = load_workbook(ojsfile, read_only=False, keep_vba=True)

        for sheetname in worksheetNames:
            ws = ojs_book[sheetname]
            ws.sheet_state = "hidden"

        ojs_book.save(ojsfile)


# # # # # # # # # # # # # # # # # # # # #

# If this line isn't here, we will get a UserWarning when we run the program, alerting us that the
# conditional formatting in the spreadsheets will not be preserved when we copy the data.
# We don't need conditional formatting within the data manipulation, so it isn't a big deal
# UserWarning: Data Validation extension is not supported and will be removed
# https://stackoverflow.com/questions/53965596/python-3-openpyxl-userwarning-data-validation-extension-not-supported
warnings.simplefilter(action="ignore", category=UserWarning)

cwd: str = os.getcwd()
if getattr(sys, "frozen", False):
    dir_path = os.path.dirname(sys.executable)
elif __file__:
    dir_path = os.path.dirname(__file__)

current_year: str = "2024"
tournament_file: str = dir_path + "\\2024-FLL-Qualifier-Tournaments.xlsx"
template_file: str = dir_path + "\\2024-Qualifier-Template.xlsm"

print("Checking to make sure *extra* files and folders are set up correctly")

# Any files that are to be copied directly into each torunament folder should be added to this list
extrafilelist: list[str] = [
    dir_path + "\\script_maker.exe",
    dir_path + "\\script_template.html.jinja",
]
ojsfilelist: list[str] = []

# Make sure the extra files exist
for filename in extrafilelist:
    try:
        if os.path.exists(filename):
            print(
                f"{filename}... CHECK!",
                tag="info",
                tag_color="green",
                color="green",
            )
    except Exception as e:
        print(
            f"Got an error checking for {filename}\n{e}",
            tag="error",
            tag_color="red",
            color="red",
        )
        input("Press enter to quit...")
        sys.exit(1)

# Read the state tournament workbook to get the details for all of the events
try:
    print(f"Getting tournaments from {dir_path}")
    book = load_workbook(tournament_file, data_only=True)
    ws = book["Tournaments"]
    columns, data = read_excel_table(ws, "TournamentList")
    dfTournaments = pd.DataFrame(data=data, columns=columns)
    ws = book["SeasonInfo"]
    columns, data = read_excel_table(ws, "SeasonInfo")
    seasonYear: int = ws["B2"].value
    seasonName: str = ws["B3"].value
    print(seasonYear, seasonName)

    print("Getting awards")
    ws = book["Awards"]
    columns, data = read_excel_table(ws, "AwardList")
    dfAwards = pd.DataFrame(data=data, columns=columns)
    print(dfAwards)

    print("Getting assignments")
    ws = book["Assignments"]
    columns, data = read_excel_table(ws, "Assignments")
    dfAssignments = pd.DataFrame(data=data, columns=columns)
    tourn_array: list[str] = []
    for index, row in dfTournaments.iterrows():
        tourn_array.append(row["Short Name"])

except Exception as e:
    print(
        f"Could not open the tournament file: {tournament_file}. Check to make sure it is not open in Excel.\n{e}",
        tag="error",
        tag_color="red",
        color="red",
    )
    input("Press enter to quit...")
    sys.exit(1)

# Are we building all of the tournaments, or just one?
tourn = input("Enter the tournament short name, or press ENTER for all tournaments: ")
if tourn != "":
    if tourn in tourn_array:
        dfTournaments = dfTournaments.loc[dfTournaments["Short Name"] == tourn]
    else:
        input(
            f"Tournament not found. The tournament name must come from this list: {tourn_array}\nPress enter to exit..."
        )
        sys.exit(1)

# Now that we have all of the info for the tournaments, loop through and
# start building the OJS files and folders
for index, row in dfTournaments.iterrows():
    newpath = dir_path + "\\tournaments\\" + row["Short Name"]
    judge_award_count = (
        row["Judges_1"]
        + row["Judges_2"]
        + row["Judges_3"]
        + row["Judges_4"]
        + row["Judges_5"]
        + row["Judges_6"]
    )
    create_folder(newpath)
    copy_files(row)
    set_up_tapi_worksheet(row)
    set_up_award_worksheet(row, judge_award_count)
    set_up_meta_worksheet(row, seasonYear, seasonName)
    add_conditional_formats(row)
    hide_worksheets(row)
    resize_worksheets(row)
    protect_worksheets(row)

input(
    f"All done. Created OJS workbooks for {len(dfTournaments)} tournament(s). Press enter to quit..."
)
sys.exit(1)
