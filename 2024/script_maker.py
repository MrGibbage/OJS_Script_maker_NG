# pip install pandas
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter, coordinate_from_string
from openpyxl.worksheet.table import Table

# pip install pyyaml
import yaml
import warnings

import os, sys, re, glob

# pip install Jinja2
from jinja2 import Environment, FileSystemLoader, select_autoescape

from typing import List, Dict
from print_color import print

# To create windows exe executable, run
# .venv\Scripts\pyinstaller.exe -F 2024\script_maker.py
# in the project folder. The executable will be saved in the 'dist'
# folder. Just copy it up to the project folder.
# Double-click to run.

def read_excel_table(sheet, table_name):
    """
    This function will read an Excel table
    and return a tuple of columns and data

    This function assumes that tables have column headers
    :param sheet: the sheet
    :param table_name: the name of the table
    :return: columns (list) and data (dict)
    """
    table = sheet.tables[table_name]
    table_range = table.ref

    table_head = sheet[table_range][0]
    table_data = sheet[table_range][1:]

    columns = [column.value for column in table_head]
    data = {column: [] for column in columns}

    for row in table_data:
        row_val = [cell.value for cell in row]
        for key, val in zip(columns, row_val):
            data[key].append(val)

    return columns, data

print("Building Closing Ceremony script")

if getattr(sys, 'frozen', False):
    dir_path = os.path.dirname(sys.executable)
elif __file__:
    dir_path = os.path.dirname(__file__)

templateLoader = FileSystemLoader(searchpath=dir_path)
templateEnv = Environment(loader=templateLoader)
TEMPLATE_FILE = "script_template.html.jinja"
try:
    template = templateEnv.get_template(TEMPLATE_FILE)
except Exception as e:
    print(
        f"Fatal error. Could not read the template file {TEMPLATE_FILE}.\nThe error was {e}",
        tag=f'error',
        tag_color="red",
        color="red",
    )
    input("Press enter to quit...")
    sys.exit(1)


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

ordinals = ["1st", "2nd", "3rd", "4th", "5th"]

# print("Opening yaml data file: " + yaml_data_file)
# with open(yaml_data_file) as f:
#     dict = yaml.load(f, Loader=yaml.FullLoader)

awards = {
    "Robot Game": "",
    "Champions": "",
    "Innovation Project": "",
    "Robot Design": "",
    "Core Values": "",
    "Judges": "",
}
rg_html = {}
dataframes = {}
# divAwards key will be 1 and/or 2, and the value will be the awardHtml dictionaries
# yes, it is a dictionary of dictionaries
divAwards: Dict[int, Dict[str, str]] = {}
divAwards[1] = {}
divAwards[2] = {} 
# awardHtml key will be the award name, and the values will be the html
awardHtml = {}
advancingDf = {}
advancingHtml = {}
awardCounts = {}
judgesAwardDf = {}
judgesAwardHtml = {}
judgesAwardTotalCount = 0
divJudgesAwards = [0, 0]
allowedAdvancingCount = [0, 0] #D1 is index 0

divisions = [1, 2]
for award in awards:
    for div in divisions:
        divAwards[div][award] = ''

removing = []
if len(glob.glob(f"{dir_path}\\~*.xlsm")) > 0:
    print("Found temporary file(s) indicating that you have one or more spreadsheets open in Excel. Please close Excel and retry.")
    input("Press enter to quit...")
    sys.exit(1)
directory_list: list[str] = glob.glob(f"{dir_path}\\*div*.xlsm")
print(f'Using this directory: {dir_path}')
print("Found these OJS files:")
print(directory_list)
if len(directory_list) ==0 or len(directory_list) > 2:
    print(
        f"Fatal error. There must be one or two OJS files in the directory. Found: {len(directory_list)}",
        tag=f'error',
        tag_color="red",
        color="red",
    )
    input("Press enter to quit...")
    sys.exit(1)

for award in awards:
    awardCounts[award] = 0

for tourn_filename in directory_list:
    regex: str = r"([0-9]{4}-vadc-fll-challenge-.*)(-ojs-)(.*)-(div[1,2])(.xlsm)$"
    # checking the file name for the division. The filename includes the path
    # so remove the path by using the length of dir_path
    m0 = re.search(regex, tourn_filename[len(dir_path) + 1:])
    div = int(m0.group(4)[-1])
    print(f"Division {div}")

    book = load_workbook(tourn_filename, data_only=True)
    ws = book['Meta']
    columns, data = read_excel_table(ws, 'Meta')
    dfMeta = pd.DataFrame(data=data, columns=["Key", "Value"])
    for award in awards:
        awardCounts[award] = dfMeta.loc[dfMeta['Key'] == award, 'Value'].values[0]
    print(awardCounts)
    ws = book["Results and Rankings"]
    columns, data = read_excel_table(ws, 'TournamentData')
    dfRankings = pd.DataFrame(data=data, columns=columns)
    print("Here is the Results and Rankings data")
    print(dfRankings)

    # Robot Game
    try:
        teamNum = dfRankings.loc[dfRankings['Robot Game Rank'] == 2, 'Team Number'].values[0]
        print(teamNum)
    except:
        print(
            f"Fatal error. I'm not seeing any scores. Have you filled out the OJS files and saved them to the right folder?",
            tag=f'error',
            tag_color="red",
            color="red",
        )
        input("Press enter to quit...")
        sys.exit(1)


    rg_html[div] = ""
    for i in reversed(range(awardCounts["Robot Game"])):
        try:
            teamNum = dfRankings.loc[dfRankings['Robot Game Rank'] == i + 1, 'Team Number'].values[0]
            teamName = dfRankings.loc[dfRankings['Robot Game Rank'] == i + 1, 'Team Name'].values[0]
            score = int(dfRankings.loc[dfRankings['Robot Game Rank'] == i + 1, 'Max Robot Game Score'].values[0])
        except:
            print(
                f"Fatal error. I'm having problems reading the robot game scores. Have you filled them in? Have you saved the OJS to the correct folder?",
                tag=f'error',
                tag_color="red",
                color="red",
            )
            print(dir_path)
            print("All robot game ranks must be properly visible on the Results and Rankings spreadsheet. Have you filled in the scores on the Robot Game Scores worksheet?")
            input("Press enter to quit...")
            sys.exit(1)
        rg_html[div] += (
            "<p>With a score of "
            + str(score)
            + " points, the Division " + str(div) + " "
            + ordinals[i]
            + " place award goes to team number "
            + str(int(teamNum))
            + ", "
            + teamName
            + "</p>\n"
        )
    
    # Judged awards
    for award in awards:
        print(award)
        if award == "Robot Game" or award == "Judges":
            continue
        for i in reversed(range(awardCounts[award])):
            thisAward = award + " " + ordinals[i] + " Place"
            print(f'Division {div}, {thisAward}')
            print(dfRankings.loc[dfRankings['Award'] == thisAward, 'Team Number'])
            try:
                teamNum = dfRankings.loc[dfRankings['Award'] == thisAward, 'Team Number'].values[0]
                teamName = dfRankings.loc[dfRankings['Award'] == thisAward, 'Team Name'].values[0]
            except:
                print(
                    f"I'm having problems reading the required {thisAward} award for Division {div}. Have you filled it in? Have you saved the OJS to the correct folder?",
                    tag=f'error',
                    tag_color="red",
                    color="red",
                )
                print(dir_path)
                print("All required awards must be properly selected on the Results and Rankings spreadsheet")
                input("Press enter to quit...")
                sys.exit(1)
            thisText = (
                "<p>The Division " + str(div) + " "
                + ordinals[i]
                + " place " + award + " award goes to team number "
                + str(int(teamNum))
                + ", "
                + teamName
                + "</p>\n"
            )
            divAwards[div][award] += thisText

    
    # Advancing
    advancingDf[div - 1] = dfRankings[dfRankings["Advance?"] == "Yes"]
    allowedAdvancingCount[div - 1] = dfMeta.loc[dfMeta['Key'] == "Advancing", 'Value'].values[0]
    advancingHtml[div] = ""
    # print(advancingDf[div])
    for index, row in advancingDf[div - 1].iterrows():
        try:
            teamNum = str(int(row['Team Number']))
        except:
            teamNum = ""
        teamName = row["Team Name"]
        advancingHtml[div] += "<p>(Div " + str(div) + ") Team number " + teamNum + ", " + teamName + "</p>\n"
    print(f'Advancing: {advancingHtml[div]}')

    print(dfRankings)
    # Judges Awards
    allowedJudgesAwardCount = dfMeta.loc[dfMeta['Key'] == "Judges", 'Value'].values[0]
    print(f'Tournament has {allowedJudgesAwardCount} judges awards available across both divisions')
    judgesAwardDf[div] = dfRankings[(dfRankings["Award"].str.startswith("Judges", na=False))]

    judgesAwardHtml[div] = ""
    # print(advancingDf[div])
    for index, row in judgesAwardDf[div].iterrows():
        try:
            teamNum = str(int(row['Team Number']))
        except:
            teamNum = ""
        teamName = row["Team Name"]
        judgesAwardHtml[div] += "<p>(Div " + str(div) + ") Team number " + teamNum + ", " + teamName + "</p>\n"
        judgesAwardTotalCount += 1

    # Check for dupes in the awards
    filtered_df = dfRankings.dropna(subset=['Award'])
    filtered_df = filtered_df[['Team Number', 'Team Name', 'Award']]
    duplicates = filtered_df.duplicated(subset=['Award'], keep=False)
    duplicate_rows = filtered_df[duplicates]
    if len(duplicate_rows) > 0:
        print(
            f'There are teams with duplicate awards',
            tag=f'error',
            tag_color="red",
            color="red",
        )
        print(duplicate_rows)
        input("Press enter to quit...")
        sys.exit(1)

    print(
        f'All done collecting data from the Div {div} OJS. Checking validity now.',
        tag=f'OK',
        tag_color="green",
        color="green",
    )
    try: 
        divJudgesAwards[div - 1] = len(judgesAwardDf[div])
    except:
        divJudgesAwards[div - 1] = 0

    print(f'Total number of judges awards selected for Div {div} = {divJudgesAwards[div - 1]}')

    # Advancing checks
    try: 
        divAdvancing = len(advancingDf[div - 1])
    except:
        divAdvancing = 0

    print(f'Total number of advancing teams selected for Div {div} = {divAdvancing}')
    if divAdvancing < allowedAdvancingCount[div - 1]:
        print(
            f'You have selected fewer advancing div {div} teams than allowed.\nYou are permitted a total of {allowedAdvancingCount[div - 1]} advancing teams for division {div}, but you have only selected {divAdvancing}.\n\nThis is not an error and you may continue building the script with fewer advancing teams than permitted.',
            tag=f'warning',
            tag_color="red",
            color="red",
        )
        try:
            input("Press enter to continue. Press ctrl-c to quit...")
        except:
            print("\n\nStopped building the script. Please check that the OJS files are filled out correctly before trying to build the script again.")
            sys.exit(0)

    if divAdvancing > allowedAdvancingCount[div - 1]:
        print(
            f'You have selected more advancing div {div} teams than allowed.\nYou are permitted a total of {allowedAdvancingCount[div - 1]} advancing teams for division {div}, but you have selected {divAdvancing}.\n\n',
            tag=f'error',
            tag_color="red",
            color="red",
        )
        input("Press enter to continue. Press ctrl-c to quit...")
        sys.exit(0)

# End of OJS loop
# # # # # # # # # # # # # # # # # # # # #

# Judges awards
judgesAwardTotalCount = divJudgesAwards[0] + divJudgesAwards[1]
if judgesAwardTotalCount > allowedJudgesAwardCount:
    print(
        f'You have selected too many judges awards: D1 = {divJudgesAwards[0]}; D2 = {divJudgesAwards[1]}. You are permitted a total of {allowedJudgesAwardCount} judges awards.\nIf your tournament has two divisions, that total number is across both divisions.\nIn other words, if you are allowed three judges awards, you could have three from\nDiv 1 or two from Div 1 and one from Div 2, etc.',
        tag=f'error',
        tag_color="red",
        color="red",
    )
    input("Press enter to quit...")
    sys.exit(1)

if judgesAwardTotalCount < allowedJudgesAwardCount:
    print(
        f'You have selected fewer judges awards than allowed. You are permitted a total of {allowedJudgesAwardCount} judges awards.\nIf your tournament has two divisions, that total number is across both divisions.\nFor example, if you are allowed three judges awards, you could have all three from Div 1.\nOr you could have two from Div 1 and one from Div 2, etc.\n\nThis is not an error and you may continue building the script with fewer judges awards than permitted.',
        tag=f'warning',
        tag_color="red",
        color="red",
    )
    try:
        input("Press enter to continue. Press ctrl-c to quit...")
    except:
        print("\n\nStopped building the script. Please check that the OJS files are filled out correctly before trying to build the script again.")
        sys.exit(0)

print(
    f'All checks look good.',
    tag=f'OK',
    tag_color="green",
    color="green",
)

print("Rendering the script")
out_text = template.render(
    tournament_name = dfMeta.loc[dfMeta['Key'] == "Tournament Long Name", 'Value'].values[0], 
    rg_div1_list = rg_html[1],
    rg_div2_list = rg_html[2],
    rd_div1_list = divAwards[1]["Robot Design"],
    rd_div2_list = divAwards[2]["Robot Design"],
    rd_this_them = "This team" if awardCounts["Robot Design"] == 1 else "These teams",
    ip_div1_list = divAwards[1]["Innovation Project"],
    ip_div2_list = divAwards[2]["Innovation Project"],
    ip_this_them = "This team" if awardCounts["Innovation Project"] == 1 else "These teams",
    cv_div1_list = divAwards[1]["Core Values"],
    cv_div2_list = divAwards[2]["Core Values"],
    cv_this_them = "This team" if awardCounts["Core Values"] == 1 else "These teams",
    ja_count = judgesAwardTotalCount,
    ja_list = judgesAwardHtml[1] + judgesAwardHtml[2],
    ja_go_goes = "The Judges Awards go to teams" if judgesAwardTotalCount > 1 else "The Judges Award goes to team",
    champ_div1_list = divAwards[1]["Champions"],
    champ_div2_list = divAwards[2]["Champions"],
    adv_div1_list = advancingHtml[1],
    adv_div2_list = advancingHtml[2],
)

# print(out_text)
with open(dfMeta.loc[dfMeta['Key'] == "Completed Script File", 'Value'].values[0], "w") as fh:
    fh.write(out_text)

print("All done! The script hase been saved as " + dfMeta.loc[dfMeta['Key'] == "Completed Script File", 'Value'].values[0] + ".")
print("It is saved in the same folder with the other OJS files.")
print("Double-click the script file to view it on this computer,")
print("or email it to yourself and view it on a phone or tablet.")
input("Press enter to quit...")