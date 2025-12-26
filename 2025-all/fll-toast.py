"""TOAST - Tournament OJS And Script Toolkit

Generate closing ceremony scripts from OJS files for FIRST LEGO League tournaments.

This script validates OJS data, collects award winners and team information,
and renders a closing ceremony script using a Jinja template.

Usage:
    python fll-toast.py [--verbose] [--debug]
    (Run from within a tournament folder containing tournament_config.json)
"""

import os
import sys
import json
import logging
import warnings
import argparse
from colorama import init, Fore, Style

# Suppress openpyxl warnings about conditional formatting
warnings.simplefilter(action="ignore", category=UserWarning)

# Add modules directory to path
sys.path.insert(0, os.path.dirname(__file__))

from modules.logger import setup_logger, print_error
from modules.ceremony_validator import OJSValidator
from modules.ceremony_data_collector import CeremonyDataCollector
from modules.ceremony_renderer import CeremonyRenderer

# Initialize colorama
init()


def print_splash():
    """Print TOAST splash screen."""
    print(f"\n{Fore.YELLOW}{'█' * 72}{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}{'  ' * 35}{Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}                           {Fore.CYAN}╔╦╗╔═╗╔═╗╔═╗╔╦╗{Style.RESET_ALL}                            {Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}                           {Fore.CYAN} ║ ║ ║╠═╣╚═╗ ║ {Style.RESET_ALL}                            {Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}                           {Fore.CYAN} ╩ ╚═╝╩ ╩╚═╝ ╩ {Style.RESET_ALL}                            {Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}{'  ' * 35}{Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}       {Fore.WHITE}Tournament OJS And Script Toolkit for FIRST LEGO League{Style.RESET_ALL}        {Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}█{Style.RESET_ALL}{'  ' * 35}{Fore.YELLOW}█{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}{'█' * 72}{Style.RESET_ALL}\n")


def print_header(text: str):
    """Print a formatted header."""
    print(f"\n{Fore.CYAN}{'═' * 70}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{text.center(70)}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'═' * 70}{Style.RESET_ALL}\n")


def print_success(text: str):
    """Print a success message."""
    print(f"{Fore.GREEN}✓ {text}{Style.RESET_ALL}")


def print_warning(text: str):
    """Print a warning message."""
    print(f"{Fore.YELLOW}⚠ {text}{Style.RESET_ALL}")


def print_error_msg(text: str):
    """Print an error message."""
    print(f"{Fore.RED}✗ {text}{Style.RESET_ALL}")


def load_config(config_path: str) -> dict:
    """Load tournament configuration file."""
    logger.info(f"Loading configuration from: {config_path}")
    
    if not os.path.exists(config_path):
        print_error(logger, f"Configuration file not found: {config_path}")
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        logger.info("✓ Configuration loaded successfully")
        return config
    except json.JSONDecodeError as e:
        print_error(logger, f"Invalid JSON in configuration file: {e}")
    except Exception as e:
        print_error(logger, f"Error loading configuration: {e}")


def generate_output_filename(ojs_filenames: list) -> str:
    """Generate output filename based on OJS filenames.
    
    Args:
        ojs_filenames: List of OJS filenames from config
        
    Returns:
        Output HTML filename
    """
    # Take first OJS filename and modify it
    base_name = ojs_filenames[0]
    
    # Remove -div1 or -div2 suffix and .xlsm extension
    base_name = base_name.replace('-div1.xlsm', '').replace('-div2.xlsm', '').replace('.xlsm', '')
    
    # Add closing-ceremony suffix
    output_name = f"{base_name}-closing-ceremony.html"
    
    logger.debug(f"Generated output filename: {output_name}")
    return output_name


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="TOAST - Tournament OJS And Script Toolkit: Generate closing ceremony scripts"
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging (INFO level)'
    )
    parser.add_argument(
        '--debug', '-d',
        action='store_true',
        help='Enable debug logging (DEBUG level, implies --verbose)'
    )
    
    return parser.parse_args()


def main():
    """Main execution function."""
    # Parse arguments first
    args = parse_arguments()
    
    # Determine logging level
    if args.debug:
        log_debug = True
    elif args.verbose:
        log_debug = False  # INFO level
    else:
        log_debug = False  # Default (WARNING level in setup_logger when debug=False)
    
    # Print splash screen
    print_splash()
    
    # Get directory where THIS script is located
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        script_dir = os.path.dirname(sys.executable)
    else:
        # Running as Python script
        script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Set up logger with appropriate level
    global logger
    logger = setup_logger("ceremony_generator", debug=log_debug, log_dir=script_dir)
    
    if args.debug:
        logger.info("Debug logging enabled")
    elif args.verbose:
        logger.info("Verbose logging enabled")
    
    logger.info(f"Script location: {script_dir}")
    
    # Load configuration
    config_path = os.path.join(script_dir, 'tournament_config.json')
    config = load_config(config_path)
    
    # Extract configuration
    info = config['INFO']
    using_divisions = info['using_divisions']
    ojs_filenames = info['ojs_filenames']
    
    # Read dual_emcee flag from OJS files at runtime (OR logic: TRUE if ANY OJS has it set)
    dual_emcee = False
    for ojs_file in ojs_filenames:
        ojs_path = os.path.join(script_dir, ojs_file)
        if os.path.exists(ojs_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(ojs_path, data_only=True)
                ws = wb["Team and Program Information"]
                dual_emcee_value = ws["F2"].value
                wb.close()
                
                # Convert to boolean
                if isinstance(dual_emcee_value, bool):
                    if dual_emcee_value:
                        dual_emcee = True
                        logger.debug(f"Dual emcee enabled from {ojs_file}")
                        break  # Found TRUE, no need to check other files
                elif isinstance(dual_emcee_value, str):
                    if dual_emcee_value.upper() in ['TRUE', 'YES', '1']:
                        dual_emcee = True
                        logger.debug(f"Dual emcee enabled from {ojs_file}")
                        break
                elif isinstance(dual_emcee_value, (int, float)):
                    if dual_emcee_value:
                        dual_emcee = True
                        logger.debug(f"Dual emcee enabled from {ojs_file}")
                        break
            except Exception as e:
                logger.debug(f"Could not read dual_emcee from {ojs_file}: {e}")
    
    print(f"{Fore.CYAN}Tournament:{Style.RESET_ALL} {info['tournament_long_name']}")
    print(f"{Fore.CYAN}Using divisions:{Style.RESET_ALL} {using_divisions}")
    print(f"{Fore.CYAN}OJS files:{Style.RESET_ALL} {len(ojs_filenames)}")
    print(f"{Fore.CYAN}Dual emcee:{Style.RESET_ALL} {dual_emcee}")
    
    # Validate OJS files exist
    print_header("VALIDATING OJS FILES")
    for ojs_file in ojs_filenames:
        ojs_path = os.path.join(script_dir, ojs_file)
        if os.path.exists(ojs_path):
            print_success(f"Found: {ojs_file}")
        else:
            print_error(logger, f"OJS file not found: {ojs_file}")
    
    # Validate OJS data
    print_header("VALIDATING OJS DATA")
    validator = OJSValidator()
    
    for idx, ojs_file in enumerate(ojs_filenames):
        ojs_path = os.path.join(script_dir, ojs_file)
        division = f"Division {idx + 1}" if using_divisions else ""
        
        print(f"\n{Fore.YELLOW}Validating {ojs_file}...{Style.RESET_ALL}")
        validator.validate_all_sheets(ojs_path, division)
    
    # Display validation results
    if validator.has_errors():
        print(f"\n{Fore.RED}{'═' * 70}{Style.RESET_ALL}")
        print(f"{Fore.RED}VALIDATION FAILED{Style.RESET_ALL}".center(78))
        print(f"{Fore.RED}{'═' * 70}{Style.RESET_ALL}\n")
        
        print(f"{Fore.RED}Errors found:{Style.RESET_ALL}")
        for error in validator.errors:
            print(f"  {error}")
        
        if validator.warnings:
            print(f"\n{Fore.YELLOW}Warnings:{Style.RESET_ALL}")
            for warning in validator.warnings:
                print(f"  {warning}")
        
        print(f"\n{Fore.RED}Please fix the errors above and run the script again.{Style.RESET_ALL}")
        input("\nPress ENTER to exit...")
        sys.exit(1)
    
    if validator.warnings:
        print(f"\n{Fore.YELLOW}Warnings found:{Style.RESET_ALL}")
        for warning in validator.warnings:
            print(f"  {warning}")
        
        response = input(f"\n{Fore.YELLOW}Continue despite warnings? [Y/n]: {Style.RESET_ALL}").strip().lower()
        if response and response not in ['y', 'yes']:
            print("Operation cancelled by user")
            sys.exit(0)
    
    print_success("All validations passed!")
    
    # Collect data
    print_header("COLLECTING AWARD DATA")
    collector = CeremonyDataCollector(config, dual_emcee=dual_emcee)
    template_data = {}
    
    # Basic info
    template_data['tournament_name'] = info['tournament_long_name']
    template_data['using_divisions'] = 1 if using_divisions else 0
    template_data['dual_emcee'] = dual_emcee  # Pass to renderer
    
    # Collect team lists
    print("Collecting team lists...")
    if using_divisions:
        div1_teams = collector.collect_team_list(os.path.join(script_dir, ojs_filenames[0]), "Division 1")
        template_data['div1_list'] = collector.format_team_list_as_html(div1_teams)
        
        if len(ojs_filenames) > 1:
            div2_teams = collector.collect_team_list(os.path.join(script_dir, ojs_filenames[1]), "Division 2")
            template_data['div2_list'] = collector.format_team_list_as_html(div2_teams)
    else:
        all_teams = collector.collect_team_list(os.path.join(script_dir, ojs_filenames[0]))
        template_data['team_list'] = collector.format_team_list_as_html(all_teams)
    
    # Collect advancing teams
    print("Collecting advancing teams...")
    if using_divisions:
        adv_d1 = collector.collect_advancing_teams(os.path.join(script_dir, ojs_filenames[0]), "Division 1")
        template_data['ADV_D1'] = collector.format_team_list_as_html(adv_d1)
        
        if len(ojs_filenames) > 1:
            adv_d2 = collector.collect_advancing_teams(os.path.join(script_dir, ojs_filenames[1]), "Division 2")
            template_data['ADV_D2'] = collector.format_team_list_as_html(adv_d2)
    
    # Collect awards
    print("Collecting award winners...")
    for award in config['AWARDS']:
        award_id = award['ID']
        award_name = award['Name']
        is_div_award = award['DivAwd']
        
        print(f"  Processing {award_name}...")
        
        # Handle Robot Game awards separately
        if award_id == 'P_AWD_RG':
            if using_divisions and is_div_award:
                # Division Robot Game awards
                d1_count = int(award.get('D1_count', 0))
                if d1_count > 0:
                    rg_d1 = collector.collect_robot_game_awards(
                        os.path.join(script_dir, ojs_filenames[0]), d1_count, "Division 1"
                    )
                    tag = award.get('ScriptTagD1', '')
                    if tag:
                        template_data[tag] = collector.format_winners_as_html(rg_d1, include_score=True)
                
                if len(ojs_filenames) > 1:
                    d2_count = int(award.get('D2_count', 0))
                    if d2_count > 0:
                        rg_d2 = collector.collect_robot_game_awards(
                            os.path.join(script_dir, ojs_filenames[1]), d2_count, "Division 2"
                        )
                        tag = award.get('ScriptTagD2', '')
                        if tag:
                            template_data[tag] = collector.format_winners_as_html(rg_d2, include_score=True)
        else:
            # Judged awards
            if using_divisions and is_div_award:
                # Division awards - get labels from config
                labels = award.get('Labels', [])
                
                d1_count = int(award.get('D1_count', 0))
                if d1_count > 0:
                    # Use only the number of labels allocated
                    d1_labels = labels[:d1_count]
                    winners_d1 = collector.collect_judged_awards(
                        os.path.join(script_dir, ojs_filenames[0]), award, d1_labels, "Division 1",
                        ojs_filenames[0]
                    )
                    tag = award.get('ScriptTagD1', '')
                    if tag:
                        template_data[tag] = collector.format_winners_as_html(winners_d1)
                    
                    # Calculate grammar variables
                    if award_id == 'J_AWD_IP':
                        template_data['ip_this_these'] = "this team" if len(winners_d1) == 1 else "these teams"
                    elif award_id == 'J_AWD_RD':
                        template_data['rd_this_these'] = "this team" if len(winners_d1) == 1 else "these teams"
                
                if len(ojs_filenames) > 1:
                    d2_count = int(award.get('D2_count', 0))
                    if d2_count > 0:
                        d2_labels = labels[:d2_count]
                        winners_d2 = collector.collect_judged_awards(
                            os.path.join(script_dir, ojs_filenames[1]), award, d2_labels, "Division 2",
                            ojs_filenames[1]
                        )
                        tag = award.get('ScriptTagD2', '')
                        if tag:
                            template_data[tag] = collector.format_winners_as_html(winners_d2)
            else:
                # Tournament-level awards (like Judges Award)
                tourn_count = int(award.get('TournCount', 0))
                labels = award.get('Labels', [])
                
                if tourn_count > 0:
                    # Collect from all divisions WITHOUT warnings
                    all_winners = []
                    
                    for idx, ojs_file in enumerate(ojs_filenames):
                        # Use the labels from config (not from OJS)
                        winners = collector.collect_judged_awards(
                            os.path.join(script_dir, ojs_file), award, labels, "",
                            ojs_file
                        )
                        all_winners.extend(winners)
                    
                    # NOW check if we got the right total count
                    if len(all_winners) < tourn_count:
                        missing_count = tourn_count - len(all_winners)
                        collector.warnings.append(
                            f"{award_name} tournament award: {len(all_winners)} selected, {tourn_count} allocated ({missing_count} under-allocated)"
                        )
                    elif len(all_winners) > tourn_count:
                        extra_count = len(all_winners) - tourn_count
                        collector.warnings.append(
                            f"{award_name} tournament award: {len(all_winners)} selected, {tourn_count} allocated ({extra_count} OVER-allocated)"
                        )
                    
                    tag = award.get('ScriptTagNoDiv', '')
                    if tag:
                        template_data[tag] = collector.format_winners_as_html(all_winners)
                    
                    # Special handling for Judges Award
                    if award_id == 'J_AWD_Judges':
                        template_data['ja_count'] = len(all_winners)
                        template_data['ja_go_goes'] = "The judges award goes to:" if len(all_winners) == 1 else "The judges awards go to:"
    
    # Set empty strings for any missing variables
    expected_vars = ['div1_list', 'div2_list', 'team_list', 'ADV_D1', 'ADV_D2',
                     'ip_this_these', 'rd_this_these', 'ja_count', 'ja_go_goes']
    for var in expected_vars:
        if var not in template_data:
            template_data[var] = ""
    
    print_success(f"Collected data for {len(template_data)} template variables")
    
    # Display collector warnings
    if collector.warnings:
        print(f"\n{Fore.YELLOW}Data collection warnings:{Style.RESET_ALL}")
        for warning in collector.warnings:
            print(f"  {warning}")
    
    # Render template
    print_header("RENDERING CEREMONY SCRIPT")
    
    renderer = CeremonyRenderer(script_dir)
    template_file = 'script_template.html.jinja'
    
    # Validate template variables
    critical_vars = {'J_AWD_CHAMP_D1', 'J_AWD_CHAMP_D2', 'ADV_D1', 'ADV_D2'}
    errors, warnings = renderer.validate_template_variables(template_file, template_data, critical_vars)
    
    if errors:
        print(f"{Fore.RED}Missing critical template variables:{Style.RESET_ALL}")
        for err in errors:
            print(f"  {err}")
        print(f"\n{Fore.RED}Cannot generate ceremony script with missing critical variables.{Style.RESET_ALL}")
        input("\nPress ENTER to exit...")
        sys.exit(1)
    
    if warnings:
        print(f"{Fore.YELLOW}Missing template variables (will be empty):{Style.RESET_ALL}")
        for warn in warnings:
            print(f"  {warn}")
    
    # Generate output filename
    output_filename = generate_output_filename(ojs_filenames)
    output_path = os.path.join(script_dir, output_filename)
    
    # Render
    success = renderer.render(template_file, template_data, output_path)
    
    if success:
        print(f"\n{Fore.GREEN}{'═' * 70}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}SUCCESS!{Style.RESET_ALL}".center(78))
        print(f"{Fore.GREEN}{'═' * 70}{Style.RESET_ALL}\n")
        print(f"{Fore.GREEN}Closing ceremony script generated:{Style.RESET_ALL}")
        print(f"  {output_path}\n")
        
        # Check if there were any warnings during the process
        has_warnings = (
            (validator.warnings and len(validator.warnings) > 0) or 
            (collector.warnings and len(collector.warnings) > 0) or
            (warnings and len(warnings) > 0)
        )
        
        if has_warnings:
            print(f"{Fore.YELLOW}{'─' * 70}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}⚠ WARNINGS DETECTED{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}{'─' * 70}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Script generated but there are warnings you should review.{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Scroll up to review the warnings and carefully review the script.{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Make changes to the OJS if needed and re-run the script generator.{Style.RESET_ALL}\n")
            input(f"{Fore.YELLOW}Press ENTER to exit...{Style.RESET_ALL}")
        else:
            input("Press ENTER to exit...")
    else:
        print_error(logger, "Failed to render ceremony script")


if __name__ == "__main__":
    main()
