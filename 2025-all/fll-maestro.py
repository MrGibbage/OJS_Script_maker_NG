"""Utility to prepare per-tournament folders and populate OJS spreadsheets.

This script reads a season manifest (`season.json`) and a master
`TournamentList` (or `DivTournamentList`) workbook to create a folder for
each tournament, copy template files, and populate the OJS (Online Judge
System) spreadsheet tables with team/award/meta information.

Usage:
    python build-tournament-folders.py              # Interactive mode (default)
    python build-tournament-folders.py --quiet       # Minimal output, no confirmations
    python build-tournament-folders.py --verbose     # Maximum output with debug logging
"""
import os
import sys
import warnings
import argparse
from openpyxl import load_workbook
from colorama import init, Fore, Style
import pandas as pd
import json

# Import from modules
from modules.logger import setup_logger, print_error
from modules.constants import *
from modules.file_operations import (
    load_json_without_notes, 
    create_folder, 
    copy_files, 
    generate_tournament_config
)
from modules.excel_operations import read_table_as_df, read_table_as_dict, verify_workbooks_closed
from modules.worksheet_setup import (
    set_up_tapi_worksheet,
    set_up_award_worksheet,
    set_up_meta_worksheet,
    resize_worksheets,
    add_essential_conditional_formats,  # Changed from add_conditional_formats
    copy_award_def,
    protect_worksheets,
    hide_worksheets,
    remove_external_links,
    fix_named_ranges,
)
from modules.user_feedback import (
    ValidationSummary,
    ProgressTracker,
    print_section_header,
    print_success,
    print_warning,
    print_info,
    confirm_action,
)

# Suppress openpyxl warnings
warnings.simplefilter(action="ignore", category=UserWarning)

# Initialize colorama
init()


def print_splash():
    """Print MAESTRO splash screen."""
    print(f"\n{Fore.CYAN}{'█' * 70}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}{'  ' * 34}{Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}                       {Fore.YELLOW}╔╦╗╔═╗╔═╗╔═╗╔╦╗╦═╗╔═╗{Style.RESET_ALL}                        {Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}                       {Fore.YELLOW}║║║╠═╣║╣ ╚═╗ ║ ╠╦╝║ ║{Style.RESET_ALL}                        {Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}                       {Fore.YELLOW}╩ ╩╩ ╩╚═╝╚═╝ ╩ ╩╚═╚═╝{Style.RESET_ALL}                        {Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}{'  ' * 34}{Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}           {Fore.WHITE}Managing All Event Seasons, Tournaments,{Style.RESET_ALL}                 {Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}           {Fore.WHITE}Rosters, and OJSs for FIRST LEGO League{Style.RESET_ALL}                  {Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}█{Style.RESET_ALL}{'  ' * 34}{Fore.CYAN}█{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'█' * 70}{Style.RESET_ALL}\n")


def parse_arguments():
    """Parse command-line arguments.
    
    Returns:
        Namespace with parsed arguments
    """
    parser = argparse.ArgumentParser(
        description="Build tournament folders and populate OJS spreadsheets",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=""":
Examples:
  %(prog)s                    Run in quiet mode (default)
  %(prog)s --interactive      Run with prompts, validation summary, and confirmations
  %(prog)s --verbose          Run with INFO-level logging
  %(prog)s --debug            Run with DEBUG-level logging (most detailed)
  %(prog)s --tournament ABC   Build only tournament with short name 'ABC'
        """
    )
    
    parser.add_argument(
        '--interactive', '-i',
        action='store_true',
        help='Interactive mode: show prompts, validation summary, and confirmations'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Verbose mode: enable INFO-level logging and detailed output'
    )
    
    parser.add_argument(
        '--debug', '-d',
        action='store_true',
        help='Debug mode: enable DEBUG-level logging (implies --verbose)'
    )
    
    parser.add_argument(
        '--tournament', '-t',
        type=str,
        metavar='NAME',
        help='Process only the specified tournament (by short name)'
    )
    
    parser.add_argument(
        '--skip-validation',
        action='store_true',
        help='Skip pre-flight validation checks (not recommended)'
    )
    
    parser.add_argument(
        '--no-cleanup',
        action='store_true',
        help='Skip cleanup of existing OJS and config files (not recommended)'
    )
    
    return parser.parse_args()

def validate_environment(dir_path: str, config: dict, tournament_file: str, template_file: str, extrafilelist: list, quiet: bool = False) -> ValidationSummary:
    """Validate the environment before processing.
    
    Args:
        dir_path: Working directory
        config: Configuration dictionary
        tournament_file: Path to tournament file
        template_file: Path to template file
        extrafilelist: List of extra files to copy
        quiet: If True, suppress info messages
        
    Returns:
        ValidationSummary with results
    """
    summary = ValidationSummary()
    
    # Check configuration
    required_config_keys = ["filename", "tournament_template", "copy_file_list", "season_yr", "season_name", "tournament_folder"]
    missing = [k for k in required_config_keys if k not in config]
    if missing:
        summary.add_error(f"Missing config keys: {', '.join(missing)}")
    elif not quiet:
        summary.add_info(f"Configuration valid: {config.get('season_name')} {config.get('season_yr')}")
    
    # Check that workbooks are closed
    try:
        verify_workbooks_closed(tournament_file, template_file)
        if not quiet:
            summary.add_info("Tournament and template files are closed")
    except RuntimeError as e:
        summary.add_error(str(e))
    
    # Check tournament file
    if not os.path.exists(tournament_file):
        summary.add_error(f"Tournament file not found: {tournament_file}")
    elif not quiet:
        summary.add_info(f"Tournament file found: {os.path.basename(tournament_file)}")
    
    # Check template file
    if not os.path.exists(template_file):
        summary.add_error(f"Template file not found: {template_file}")
    elif not quiet:
        summary.add_info(f"Template file found: {os.path.basename(template_file)}")
    
    # Check extra files
    missing_files = []
    for filename in extrafilelist:
        if not os.path.exists(os.path.join(dir_path, filename)):
            missing_files.append(filename)
    
    if missing_files:
        summary.add_warning(f"Missing optional files: {', '.join(missing_files)}")
    elif not quiet:
        summary.add_info(f"All {len(extrafilelist)} extra files found")
    
    return summary

def cleanup_tournament_folders(tournament_folder: str, tournaments_to_process: list[str], quiet: bool = False) -> dict:
    """Remove existing OJS and config files from tournament folders.
    
    Args:
        tournament_folder: Base tournament folder path
        tournaments_to_process: List of tournament short names to clean
        quiet: If True, suppress progress messages
        
    Returns:
        Dictionary with cleanup statistics
    """
    logger.info("Starting cleanup of existing tournament files")
    
    stats = {
        'ojs_deleted': 0,
        'config_deleted': 0,
        'folders_processed': 0,
        'deletion_failures': []
    }
    
    for tourn_name in tournaments_to_process:
        folder_path = os.path.join(tournament_folder, tourn_name)
        
        if not os.path.exists(folder_path):
            continue
            
        stats['folders_processed'] += 1
        
        # Delete OJS files (*.xlsm)
        for file in os.listdir(folder_path):
            if file.endswith('.xlsm'):
                file_path = os.path.join(folder_path, file)
                try:
                    os.remove(file_path)
                    stats['ojs_deleted'] += 1
                    logger.debug(f"Deleted OJS file: {file}")
                except PermissionError as e:
                    error_msg = f"{tourn_name}/{file} (file may be open in Excel)"
                    stats['deletion_failures'].append(error_msg)
                    logger.warning(f"Permission denied deleting {file}: {e}")
                except Exception as e:
                    error_msg = f"{tourn_name}/{file} ({str(e)})"
                    stats['deletion_failures'].append(error_msg)
                    logger.warning(f"Could not delete {file}: {e}")
        
        # Delete tournament_config.json
        config_file = os.path.join(folder_path, 'tournament_config.json')
        if os.path.exists(config_file):
            try:
                os.remove(config_file)
                stats['config_deleted'] += 1
                logger.debug(f"Deleted config: {tourn_name}/tournament_config.json")
            except PermissionError as e:
                error_msg = f"{tourn_name}/tournament_config.json (file may be open)"
                stats['deletion_failures'].append(error_msg)
                logger.warning(f"Permission denied deleting tournament_config.json: {e}")
            except Exception as e:
                error_msg = f"{tourn_name}/tournament_config.json ({str(e)})"
                stats['deletion_failures'].append(error_msg)
                logger.warning(f"Could not delete tournament_config.json: {e}")
    
    if not quiet:
        print_info(f"Cleanup: {stats['ojs_deleted']} OJS files, {stats['config_deleted']} config files removed")
        if stats['deletion_failures']:
            print_warning(f"Failed to delete {len(stats['deletion_failures'])} file(s) - they may be open")
    
    logger.info(f"Cleanup complete: {stats}")
    return stats

def main():
    """Main execution function."""
    args = parse_arguments()
    
    # Quiet mode is default; interactive is opt-in
    quiet = not args.interactive
    
    # Print splash screen (always shown)
    print_splash()
    
    # Determine script directory FIRST
    if getattr(sys, "frozen", False):
        dir_path = os.path.dirname(sys.executable)
    elif __file__:
        dir_path = os.path.dirname(__file__)
    
    # Determine logging level (debug implies verbose)
    log_debug = args.debug or args.verbose
    
    # Set up logger with script directory
    global logger
    logger = setup_logger("ojs_builder", debug=log_debug, log_dir=dir_path)
    
    if args.debug:
        logger.info("Debug logging enabled")
    elif args.verbose:
        logger.info("Verbose logging enabled")
    
    if not quiet:
        print_info(f"Working directory: {dir_path}")
    else:
        logger.info("Starting tournament folder builder")
    
    # Load configuration
    try:
        config = load_json_without_notes(os.path.join(dir_path, CONFIG_FILENAME))
        if not quiet:
            print_success("Configuration loaded")
    except FileNotFoundError:
        print_error(
            logger,
            f"Configuration file '{CONFIG_FILENAME}' not found in {dir_path}",
            error_type='missing_config',
            context={'filename': CONFIG_FILENAME, 'directory': dir_path}
        )
    except json.JSONDecodeError as e:
        print_error(
            logger,
            f"Invalid JSON syntax in '{CONFIG_FILENAME}'",
            e,
            error_type='invalid_json',
            context={'filename': CONFIG_FILENAME}
        )
    except Exception as e:
        print_error(logger, f"Error loading configuration file '{CONFIG_FILENAME}'", e)

    tournament_file = os.path.join(dir_path, config["filename"])
    template_file = os.path.join(dir_path, config["tournament_template"])
    extrafilelist = config["copy_file_list"]
    tournament_folder = config["tournament_folder"]
    
    # Ensure tournament folder exists
    if not os.path.exists(tournament_folder):
        try:
            os.makedirs(tournament_folder)
            logger.info(f"Created tournament folder: {tournament_folder}")
            if not quiet:
                print_success(f"Created tournament folder: {tournament_folder}")
        except Exception as e:
            print_error(
                logger,
                f"Could not create tournament folder: {tournament_folder}",
                e,
                error_type='permission_denied',
                context={'filename': tournament_folder}
            )
    else:
        logger.debug(f"Tournament folder exists: {tournament_folder}")

    # Run validation (unless skipped)
    if not args.skip_validation:
        if not quiet:
            print_section_header("PRE-FLIGHT VALIDATION")
        
        validation = validate_environment(dir_path, config, tournament_file, template_file, extrafilelist, quiet=quiet)
        
        # Always display validation results if there are errors or warnings
        if validation.has_errors() or validation.warnings:
            validation.display()
        elif not quiet:
            validation.display()
        
        if validation.has_errors():
            print_error(
                logger,
                "Validation failed. Please fix the errors above and try again.",
                error_type='invalid_data',
                context={'location': 'pre-flight validation'}
            )
        
        if not quiet and validation.warnings:
            if not confirm_action("⚠ There are warnings. Continue anyway?", default=True):
                logger.info("User cancelled due to warnings")
                sys.exit(0)

    # Load tournament data
    if not quiet:
        print_section_header("LOADING TOURNAMENT DATA")
    
    try:
        logger.info(f"Opening {os.path.basename(tournament_file)}...")
        _ = load_workbook(tournament_file, data_only=True)
        if not quiet:
            print_success("Tournament file opened successfully")
    except Exception as e:
        print_error(
            logger,
            f"Could not open tournament file: {tournament_file}",
            e,
            error_type='file_open',
            context={'filename': os.path.basename(tournament_file)}
        )

    # Read season info
    try:
        dictSeasonInfo = read_table_as_dict(tournament_file, "SeasonInfo", "SeasonInfo")
        divisions_value = dictSeasonInfo["Divisions"]
        
        # Convert to boolean (handle string, bool, int)
        if isinstance(divisions_value, bool):
            using_divisions = divisions_value
        elif isinstance(divisions_value, str):
            using_divisions = divisions_value.upper() in ['TRUE', 'YES', '1']
        elif isinstance(divisions_value, (int, float)):
            using_divisions = bool(divisions_value)
        else:
            using_divisions = False
        
        logger.info(f"Using divisions: {using_divisions} (from value: {divisions_value})")
    except Exception as e:
        print_error(logger, "Could not read the SeasonInfo table", e)

    # Read tournaments
    try:
        if using_divisions:
            dfTournaments = read_table_as_df(
                tournament_file, "DivTournaments", "DivTournamentList"
            ).fillna(0)
        else:
            dfTournaments = read_table_as_df(
                tournament_file, "Tournaments", "TournamentList"
            ).fillna(0)
        logger.info(f"Loaded {len(dfTournaments)} tournament(s)")
    except Exception as e:
        print_error(logger, "Could not read the tournament worksheet", e)

    # Read award definitions
    try:
        dfAwardDef = read_table_as_df(tournament_file, "AwardDef", "AwardDef").fillna(0)
        logger.info(f"Loaded {len(dfAwardDef)} award definitions")
    except Exception as e:
        print_error(logger, "Could not read the AwardDef worksheet", e)

    # Read assignments
    try:
        dfAssignments = read_table_as_df(tournament_file, "Assignments", "Assignments").fillna(0)
        tourn_array = dfTournaments[COL_SHORT_NAME].tolist()
        if not quiet:
            print_success(f"Loaded {len(dfAssignments)} team assignments")
    except Exception as e:
        print_error(
            logger,
            "Could not read the assignments worksheet",
            e,
            error_type='missing_sheet',
            context={
                'workbook': os.path.basename(tournament_file),
                'sheet_name': 'Assignments',
                'available_sheets': []
            }
        )

    # Tournament selection
    if args.tournament:
        # Use tournament from command line
        tourn = args.tournament
        if tourn in tourn_array:
            dfTournaments = dfTournaments.loc[dfTournaments[COL_SHORT_NAME] == tourn]
            logger.info(f"Building single tournament from --tournament arg: {tourn}")
            if not quiet:
                print_success(f"Building single tournament: {tourn}")
        else:
            print_error(
                logger,
                f"Tournament '{tourn}' not found",
                error_type='invalid_data',
                context={
                    'location': 'tournament selection',
                    'expected': f"One of: {', '.join(tourn_array)}",
                    'found': tourn
                }
            )
    else:
        # ALWAYS prompt for tournament selection (even in quiet mode)
        if not quiet:
            print_section_header("TOURNAMENT SELECTION")
        
        if len(tourn_array) > 1:
            if not quiet:
                print_info(f"Available tournaments: {', '.join(tourn_array)}")
            else:
                print(f"\n{Fore.CYAN}Available tournaments: {', '.join(tourn_array)}{Style.RESET_ALL}")
            
            tourn = input(f"{Fore.CYAN}Enter tournament short name, or press ENTER for all: {Style.RESET_ALL}").strip()
        else:
            if not quiet:
                print_info(f"Single tournament detected: {tourn_array[0]}")
            else:
                print(f"{Fore.CYAN}Single tournament detected: {tourn_array[0]}{Style.RESET_ALL}")
            tourn = ""
        
        if tourn != "":
            if tourn in tourn_array:
                dfTournaments = dfTournaments.loc[dfTournaments[COL_SHORT_NAME] == tourn]
                if not quiet:
                    print_success(f"Building single tournament: {tourn}")
                else:
                    logger.info(f"Building single tournament: {tourn}")
            else:
                print_error(
                    logger,
                    f"Tournament '{tourn}' not found",
                    error_type='invalid_data',
                    context={
                        'location': 'tournament selection',
                        'expected': f"One of: {', '.join(tourn_array)}",
                        'found': tourn
                    }
                )

    # Cleanup existing files (unless skipped)
    if not args.no_cleanup:
        if not quiet:
            print_section_header("CLEANUP")
            print_info("Removing existing OJS and config files...")
        
        # Get list of tournaments to clean
        tournaments_to_clean = dfTournaments[COL_SHORT_NAME].unique().tolist()
        
        # In interactive mode, ask for confirmation
        if not quiet:
            cleanup_message = (
                f"This will delete existing OJS files and tournament_config.json "
                f"from {len(tournaments_to_clean)} tournament folder(s).\n"
                f"Proceed with cleanup?"
            )
            if not confirm_action(cleanup_message, default=True):
                logger.info("User skipped cleanup")
                print_warning("Cleanup skipped by user")
            else:
                cleanup_stats = cleanup_tournament_folders(
                    tournament_folder,
                    tournaments_to_clean,
                    quiet=quiet
                )
                
                if not quiet:
                    print_success(f"Cleanup complete: {cleanup_stats['folders_processed']} folders processed")
                
                # Display deletion failures if any
                if cleanup_stats['deletion_failures']:
                    print(f"\n{Fore.YELLOW}⚠ Could not delete {len(cleanup_stats['deletion_failures'])} file(s):{Style.RESET_ALL}")
                    for failure in cleanup_stats['deletion_failures']:
                        print(f"  • {failure}")
                    print(f"\n{Fore.RED}ACTION REQUIRED: Close any open Excel files and try again.{Style.RESET_ALL}\n")
                    
                    if not confirm_action("Continue despite deletion failures?", default=False):
                        logger.info("User cancelled due to deletion failures")
                        print_warning("Operation cancelled by user")
                        sys.exit(0)
        else:
            # Quiet mode - automatic cleanup
            cleanup_stats = cleanup_tournament_folders(
                tournament_folder,
                tournaments_to_clean,
                quiet=quiet
            )
            
            # Still exit if there are deletion failures in quiet mode
            if cleanup_stats['deletion_failures']:
                logger.error(f"Cleanup failed: {len(cleanup_stats['deletion_failures'])} file(s) could not be deleted")
                print_error(
                    logger,
                    f"Could not delete {len(cleanup_stats['deletion_failures'])} file(s). Close any open Excel files and try again.",
                    error_type='file_open',
                    context={'files': cleanup_stats['deletion_failures']}
                )
    else:
        logger.info("Cleanup skipped (--no-cleanup flag)")

    # Confirm before processing (unless quiet)
    if not quiet:
        print_section_header("READY TO PROCESS")
        print_info(f"Tournaments to process: {len(dfTournaments)}")
        if not confirm_action("Proceed with tournament folder creation?", default=True):
            logger.info("User cancelled operation")
            print_warning("Operation cancelled by user")
            sys.exit(0)

    # Process tournaments
    if not quiet:
        print_section_header("PROCESSING TOURNAMENTS")
    else:
        logger.info(f"Processing {len(dfTournaments)} tournament(s)...")
    
    # Track division mismatches and award count mismatches for final summary
    division_mismatches = []
    award_count_issues = {}  # tournament_name -> list of mismatch messages
    
    for index, row in dfTournaments.iterrows():
        tournament_name = f"{row[COL_SHORT_NAME]} {row.get(COL_DIVISION, '')}".strip()
        
        if not quiet:
            print(f"\n{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}  {tournament_name}  {Style.RESET_ALL}".center(70))
            print(f"{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}\n")
            progress = ProgressTracker(8, f"Setting up {row[COL_SHORT_NAME]}")
        else:
            logger.info(f"Processing {tournament_name}")
        
        # Create folder
        newpath = os.path.join(tournament_folder, row[COL_SHORT_NAME])
        create_folder(newpath)
        if not quiet:
            progress.update("Folder created")
        
        # Copy files
        copy_files(row, dir_path, template_file, extrafilelist, tournament_folder, using_divisions)
        if not quiet:
            progress.update("Files copied")

        # Process OJS file
        ojs_name = row.get(COL_OJS_FILENAME)
        if ojs_name is None or (isinstance(ojs_name, float) and pd.isna(ojs_name)):
            if not quiet:
                print_warning(f"No OJS filename for {row[COL_SHORT_NAME]}, skipping")
            logger.warning(f"No OJS filename for {row[COL_SHORT_NAME]}, skipping")
            continue
            
        ojs_path = os.path.join(tournament_folder, row[COL_SHORT_NAME], ojs_name)
        
        ojs_book = load_workbook(ojs_path, read_only=False, keep_vba=True)
        
        # Check if there are teams assigned; skip if not
        has_teams = set_up_tapi_worksheet(row, ojs_book, dfAssignments, using_divisions)
        
        if not has_teams:
            ojs_book.close()
            # Delete the OJS file we just created since there are no teams
            if os.path.exists(ojs_path):
                os.remove(ojs_path)
            if not quiet:
                print_warning(f"No teams assigned to {tournament_name}, OJS file removed")
            logger.warning(f"Skipped {tournament_name} - no teams assigned")
            continue
        
        # Process the tournament (only reached if has_teams is True)
        try:
            if not quiet:
                progress.update("Team info added")
            
            set_up_award_worksheet(row, ojs_book, dfAwardDef, using_divisions)
            if not quiet:
                progress.update("Awards configured")
            
            set_up_meta_worksheet(row, ojs_book, config, tournament_folder, using_divisions)
            if not quiet:
                progress.update("Metadata added")
            
            copy_award_def(row, ojs_book, dfAwardDef)
            if not quiet:
                progress.update("Formatting applied")
            
            hide_worksheets(row, ojs_book)
            if not quiet:
                progress.update("Worksheets hidden")
            
            resize_worksheets(row, ojs_book, dfAssignments, using_divisions)
            if not quiet:
                progress.update("Tables resized")
            
            # Add essential conditional formatting AFTER resize
            add_essential_conditional_formats(ojs_book, len(dfAssignments[dfAssignments[COL_SHORT_NAME] == row[COL_SHORT_NAME]]))
            
            protect_worksheets(row, ojs_book)
            if not quiet:
                progress.update("Protection applied")
            
            # Fix named ranges (especially "Awards" range)
            fix_named_ranges(ojs_book)
            
            # Remove any external workbook links before saving
            remove_external_links(ojs_book)
            if not quiet:
                progress.update("Links removed")
            
        finally:
            ojs_book.save(ojs_path)
            ojs_book.close()
            
        # Generate tournament config file
        mismatch_detected, tourn_name, award_mismatches = generate_tournament_config(
            row, config, dfAwardDef, using_divisions, tournament_folder, quiet=quiet
        )
        
        if mismatch_detected:
            division_mismatches.append(tourn_name)
        
        if award_mismatches:
            award_count_issues[tourn_name] = award_mismatches
            
        if not quiet:
            progress.complete(f"✓ {tournament_name} complete!")
        else:
            logger.info(f"✓ Completed: {tournament_name}")

    if not quiet:
        print(f"\n{Fore.GREEN}{'═' * 60}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}  ALL TOURNAMENTS PROCESSED SUCCESSFULLY!  {Style.RESET_ALL}".center(70))
        print(f"{Fore.GREEN}{'═' * 60}{Style.RESET_ALL}\n")
        
        # Important note about VBA references
        print(f"{Fore.CYAN}{'─' * 60}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}IMPORTANT:{Style.RESET_ALL}")
        print(f"If Excel shows 'Update Links' warnings when opening OJS files:")
        print(f"  1. Click 'Don't Update' or 'Break Links'")
        print(f"  2. Press Alt+F11, go to Tools → References")
        print(f"  3. Uncheck any MISSING references")
        print(f"  4. Save the file\n")
        print(f"To prevent this: Clean the template file's VBA references first.")
        print(f"{Fore.CYAN}{'─' * 60}{Style.RESET_ALL}\n")
    
    logger.info("All tournaments processed successfully")
    
    # Track if there were any warnings or issues
    has_warnings = bool(division_mismatches or award_count_issues)
    
    # Display division mismatch summary if any occurred
    if division_mismatches:
        logger.warning(f"Division mismatches detected in {len(division_mismatches)} tournament(s)")
        print(f"\n{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}  ⚠ DIVISION MISMATCH SUMMARY  {Style.RESET_ALL}".center(70))
        print(f"{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}\n")
        print(f"{Fore.YELLOW}The following tournament(s) had division mismatch issues:{Style.RESET_ALL}")
        for tourn in division_mismatches:
            print(f"  • {tourn}")
        print(f"\n{Fore.YELLOW}These tournaments had tournament_config.json with using_divisions=false,")
        print(f"but appeared to use divisions. The setting has been changed to true.{Style.RESET_ALL}")
        print(f"\n{Fore.RED}⚠ ACTION REQUIRED:{Style.RESET_ALL}")
        print(f"  Check season workbook settings before proceeding.")
        print(f"  OJS files may not work as expected if division settings are incorrect.\n")
        print(f"{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}\n")
    
    # Display award count mismatch summary if any occurred
    if award_count_issues:
        logger.warning(f"Award count mismatches detected in {len(award_count_issues)} tournament(s)")
        print(f"\n{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}  ⚠ AWARD COUNT MISMATCH SUMMARY  {Style.RESET_ALL}".center(70))
        print(f"{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}\n")
        print(f"{Fore.YELLOW}Tournament-level awards should have the same count in both divisions.{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}The following mismatches were found:{Style.RESET_ALL}\n")
        for tourn, mismatches in award_count_issues.items():
            print(f"{Fore.CYAN}{tourn}:{Style.RESET_ALL}")
            for mismatch in mismatches:
                print(f"  • {mismatch}")
            print()
        print(f"{Fore.RED}⚠ ACTION REQUIRED:{Style.RESET_ALL}")
        print(f"  Check the season workbook tournament tables (TournamentList/DivTournamentList).")
        print(f"  Verify tournament-level award counts match between divisions.")
        print(f"  OJS files may not work as expected if award counts are incorrect.\n")
        print(f"{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}\n")
    
    # Final exit message with warning reminder if needed
    if not quiet and has_warnings:
        print(f"\n{Fore.YELLOW}{'─' * 60}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}⚠ WARNINGS DETECTED{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}{'─' * 60}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Tournament files generated but there are warnings above.{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Please review the warnings carefully before distributing OJS files.{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Check the season workbook and re-run if corrections are needed.{Style.RESET_ALL}\n")
        input(f"{Fore.YELLOW}Press ENTER to exit...{Style.RESET_ALL}")
    elif not quiet:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()