"""Utility to prepare per-tournament folders and populate OJS spreadsheets.

This script reads a season manifest (`season.json`) and a master
`TournamentList` (or `DivTournamentList`) workbook to create a folder for
each tournament, copy template files, and populate the OJS (Online Judge
System) spreadsheet tables with team/award/meta information.

Usage:
    python build-tournament-folders.py              # Interactive mode (default)
    python build-tournament-folders.py --quiet       # Minimal output, no confirmations
    python build-tournament-folders.py --verbose     # Maximum output with debug logging
"""

import os
import sys
import warnings
import argparse
from openpyxl import load_workbook
from colorama import init, Fore, Style
import pandas as pd
import json

# Import from modules
from modules.logger import setup_logger, print_error
from modules.constants import *
from modules.file_operations import load_json_without_notes, create_folder, copy_files
from modules.excel_operations import read_table_as_df, read_table_as_dict
from modules.worksheet_setup import (
    set_up_tapi_worksheet,
    set_up_award_worksheet,
    set_up_meta_worksheet,
    resize_worksheets,
    add_conditional_formats,
    copy_award_def,
    protect_worksheets,
    hide_worksheets,
)
from modules.user_feedback import (
    ValidationSummary,
    ProgressTracker,
    print_section_header,
    print_success,
    print_warning,
    print_info,
    confirm_action,
)

# Suppress openpyxl warnings
warnings.simplefilter(action="ignore", category=UserWarning)

# Initialize colorama
init()

def parse_arguments():
    """Parse command-line arguments.
    
    Returns:
        Namespace with parsed arguments
    """
    parser = argparse.ArgumentParser(
        description="Build tournament folders and populate OJS spreadsheets",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=""":
Examples:
  %(prog)s                    Run in quiet mode (default)
  %(prog)s --interactive      Run with prompts, validation summary, and confirmations
  %(prog)s --verbose          Run with debug logging enabled
  %(prog)s --tournament ABC   Build only tournament with short name 'ABC'
        """
    )
    
    parser.add_argument(
        '--interactive', '-i',
        action='store_true',
        help='Interactive mode: show prompts, validation summary, and confirmations'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Verbose mode: enable debug logging and detailed output'
    )
    
    parser.add_argument(
        '--tournament', '-t',
        type=str,
        metavar='NAME',
        help='Process only the specified tournament (by short name)'
    )
    
    parser.add_argument(
        '--skip-validation',
        action='store_true',
        help='Skip pre-flight validation checks (not recommended)'
    )
    
    return parser.parse_args()

def validate_environment(dir_path: str, config: dict, tournament_file: str, template_file: str, extrafilelist: list, quiet: bool = False) -> ValidationSummary:
    """Validate the environment before processing.
    
    Args:
        dir_path: Working directory
        config: Configuration dictionary
        tournament_file: Path to tournament file
        template_file: Path to template file
        extrafilelist: List of extra files to copy
        quiet: If True, suppress info messages
        
    Returns:
        ValidationSummary with results
    """
    summary = ValidationSummary()
    
    # Check configuration
    required_config_keys = ["filename", "tournament_template", "copy_file_list", "season_yr", "season_name"]
    missing = [k for k in required_config_keys if k not in config]
    if missing:
        summary.add_error(f"Missing config keys: {', '.join(missing)}")
    elif not quiet:
        summary.add_info(f"Configuration valid: {config.get('season_name')} {config.get('season_yr')}")
    
    # Check tournament file
    if not os.path.exists(tournament_file):
        summary.add_error(f"Tournament file not found: {tournament_file}")
    elif not quiet:
        summary.add_info(f"Tournament file found: {os.path.basename(tournament_file)}")
    
    # Check template file
    if not os.path.exists(template_file):
        summary.add_error(f"Template file not found: {template_file}")
    elif not quiet:
        summary.add_info(f"Template file found: {os.path.basename(template_file)}")
    
    # Check extra files
    missing_files = []
    for filename in extrafilelist:
        if not os.path.exists(os.path.join(dir_path, filename)):
            missing_files.append(filename)
    
    if missing_files:
        summary.add_warning(f"Missing optional files: {', '.join(missing_files)}")
    elif not quiet:
        summary.add_info(f"All {len(extrafilelist)} extra files found")
    
    return summary


def main():
    """Main execution function."""
    args = parse_arguments()
    
    # Quiet mode is default; interactive is opt-in
    quiet = not args.interactive
    
    # Set up logger with appropriate verbosity
    global logger
    logger = setup_logger("ojs_builder", debug=args.verbose)
    
    # Determine script directory
    if getattr(sys, "frozen", False):
        dir_path = os.path.dirname(sys.executable)
    elif __file__:
        dir_path = os.path.dirname(__file__)
    
    if not quiet:
        print_section_header("OJS TOURNAMENT FOLDER BUILDER")
        print_info(f"Working directory: {dir_path}")
    else:
        logger.info("Starting tournament folder builder")
    
    # Load configuration
    try:
        config = load_json_without_notes(os.path.join(dir_path, CONFIG_FILENAME))
        if not quiet:
            print_success("Configuration loaded")
    except FileNotFoundError:
        print_error(
            logger,
            f"Configuration file '{CONFIG_FILENAME}' not found in {dir_path}",
            error_type='missing_config',
            context={'filename': CONFIG_FILENAME, 'directory': dir_path}
        )
    except json.JSONDecodeError as e:
        print_error(
            logger,
            f"Invalid JSON syntax in '{CONFIG_FILENAME}'",
            e,
            error_type='invalid_json',
            context={'filename': CONFIG_FILENAME}
        )
    except Exception as e:
        print_error(logger, f"Error loading configuration file '{CONFIG_FILENAME}'", e)

    tournament_file = os.path.join(dir_path, config["filename"])
    template_file = os.path.join(dir_path, config["tournament_template"])
    extrafilelist = config["copy_file_list"]

    # Run validation (unless skipped)
    if not args.skip_validation:
        if not quiet:
            print_section_header("PRE-FLIGHT VALIDATION")
        
        validation = validate_environment(dir_path, config, tournament_file, template_file, extrafilelist, quiet=quiet)
        
        if not quiet:
            validation.display()
        
        if validation.has_errors():
            print_error(
                logger,
                "Validation failed. Please fix the errors above and try again.",
                error_type='invalid_data',
                context={'location': 'pre-flight validation'}
            )
        
        if not quiet and validation.warnings:
            if not confirm_action("⚠ There are warnings. Continue anyway?", default=True):
                logger.info("User cancelled due to warnings")
                sys.exit(0)

    # Load tournament data
    if not quiet:
        print_section_header("LOADING TOURNAMENT DATA")
    
    try:
        logger.info(f"Opening {os.path.basename(tournament_file)}...")
        _ = load_workbook(tournament_file, data_only=True)
        if not quiet:
            print_success("Tournament file opened successfully")
    except Exception as e:
        print_error(
            logger,
            f"Could not open tournament file: {tournament_file}",
            e,
            error_type='file_open',
            context={'filename': os.path.basename(tournament_file)}
        )

    # Read season info
    try:
        dictSeasonInfo = read_table_as_dict(tournament_file, "SeasonInfo", "SeasonInfo")
        using_divisions = dictSeasonInfo["Divisions"]
        logger.info(f"Using divisions: {using_divisions}")
    except Exception as e:
        print_error(logger, "Could not read the SeasonInfo table", e)

    # Read tournaments
    try:
        if using_divisions:
            dfTournaments = read_table_as_df(
                tournament_file, "DivTournaments", "DivTournamentList"
            ).fillna(0)
        else:
            dfTournaments = read_table_as_df(
                tournament_file, "Tournaments", "TournamentList"
            ).fillna(0)
        logger.info(f"Loaded {len(dfTournaments)} tournament(s)")
    except Exception as e:
        print_error(logger, "Could not read the tournament worksheet", e)

    # Read award definitions
    try:
        dfAwardDef = read_table_as_df(tournament_file, "AwardDef", "AwardDef").fillna(0)
        logger.info(f"Loaded {len(dfAwardDef)} award definitions")
    except Exception as e:
        print_error(logger, "Could not read the AwardDef worksheet", e)

    # Read assignments
    try:
        dfAssignments = read_table_as_df(tournament_file, "Assignments", "Assignments").fillna(0)
        tourn_array = dfTournaments[COL_SHORT_NAME].tolist()
        if not quiet:
            print_success(f"Loaded {len(dfAssignments)} team assignments")
    except Exception as e:
        print_error(
            logger,
            "Could not read the assignments worksheet",
            e,
            error_type='missing_sheet',
            context={
                'workbook': os.path.basename(tournament_file),
                'sheet_name': 'Assignments',
                'available_sheets': []
            }
        )

    # Tournament selection
    if args.tournament:
        # Use tournament from command line
        tourn = args.tournament
        if tourn in tourn_array:
            dfTournaments = dfTournaments.loc[dfTournaments[COL_SHORT_NAME] == tourn]
            logger.info(f"Building single tournament from --tournament arg: {tourn}")
            if not quiet:
                print_success(f"Building single tournament: {tourn}")
        else:
            print_error(
                logger,
                f"Tournament '{tourn}' not found",
                error_type='invalid_data',
                context={
                    'location': 'tournament selection',
                    'expected': f"One of: {', '.join(tourn_array)}",
                    'found': tourn
                }
            )
    else:
        # ALWAYS prompt for tournament selection (even in quiet mode)
        if not quiet:
            print_section_header("TOURNAMENT SELECTION")
        
        if len(tourn_array) > 1:
            if not quiet:
                print_info(f"Available tournaments: {', '.join(tourn_array)}")
            else:
                print(f"\n{Fore.CYAN}Available tournaments: {', '.join(tourn_array)}{Style.RESET_ALL}")
            
            tourn = input(f"{Fore.CYAN}Enter tournament short name, or press ENTER for all: {Style.RESET_ALL}").strip()
        else:
            if not quiet:
                print_info(f"Single tournament detected: {tourn_array[0]}")
            else:
                print(f"{Fore.CYAN}Single tournament detected: {tourn_array[0]}{Style.RESET_ALL}")
            tourn = ""
        
        if tourn != "":
            if tourn in tourn_array:
                dfTournaments = dfTournaments.loc[dfTournaments[COL_SHORT_NAME] == tourn]
                if not quiet:
                    print_success(f"Building single tournament: {tourn}")
                else:
                    logger.info(f"Building single tournament: {tourn}")
            else:
                print_error(
                    logger,
                    f"Tournament '{tourn}' not found",
                    error_type='invalid_data',
                    context={
                        'location': 'tournament selection',
                        'expected': f"One of: {', '.join(tourn_array)}",
                        'found': tourn
                    }
                )

    # Confirm before processing (unless quiet)
    if not quiet:
        print_section_header("READY TO PROCESS")
        print_info(f"Tournaments to process: {len(dfTournaments)}")
        if not confirm_action("Proceed with tournament folder creation?", default=True):
            logger.info("User cancelled operation")
            print_warning("Operation cancelled by user")
            sys.exit(0)

    # Process tournaments
    if not quiet:
        print_section_header("PROCESSING TOURNAMENTS")
    else:
        logger.info(f"Processing {len(dfTournaments)} tournament(s)...")
    
    for index, row in dfTournaments.iterrows():
        tournament_name = f"{row[COL_SHORT_NAME]} {row.get(COL_DIVISION, '')}".strip()
        
        if not quiet:
            print(f"\n{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}  {tournament_name}  {Style.RESET_ALL}".center(70))
            print(f"{Fore.YELLOW}{'═' * 60}{Style.RESET_ALL}\n")
            progress = ProgressTracker(8, f"Setting up {row[COL_SHORT_NAME]}")
        else:
            logger.info(f"Processing {tournament_name}")
        
        # Create folder
        newpath = os.path.join(dir_path, FOLDER_TOURNAMENTS, row[COL_SHORT_NAME])
        create_folder(newpath)
        if not quiet:
            progress.update("Folder created")
        
        # Copy files
        copy_files(row, dir_path, template_file, extrafilelist)
        if not quiet:
            progress.update("Files copied")

        # Process OJS file
        ojs_name = row.get(COL_OJS_FILENAME)
        if ojs_name is None or (isinstance(ojs_name, float) and pd.isna(ojs_name)):
            if not quiet:
                print_warning(f"No OJS filename for {row[COL_SHORT_NAME]}, skipping")
            logger.warning(f"No OJS filename for {row[COL_SHORT_NAME]}, skipping")
            continue
            
        ojs_path = os.path.join(dir_path, FOLDER_TOURNAMENTS, row[COL_SHORT_NAME], ojs_name)
        
        ojs_book = load_workbook(ojs_path, read_only=False, keep_vba=True)
        try:
            set_up_tapi_worksheet(row, ojs_book, dfAssignments, using_divisions)
            if not quiet:
                progress.update("Team info added")
            
            set_up_award_worksheet(row, ojs_book, dfAwardDef, using_divisions)
            if not quiet:
                progress.update("Awards configured")
            
            set_up_meta_worksheet(row, ojs_book, config, dir_path, using_divisions)
            if not quiet:
                progress.update("Metadata added")
            
            add_conditional_formats(row, ojs_book)
            copy_award_def(row, ojs_book, dfAwardDef)
            if not quiet:
                progress.update("Formatting applied")
            
            hide_worksheets(row, ojs_book)
            if not quiet:
                progress.update("Worksheets hidden")
            
            resize_worksheets(row, ojs_book, dfAssignments, using_divisions)
            if not quiet:
                progress.update("Tables resized")
            
            protect_worksheets(row, ojs_book)
            if not quiet:
                progress.update("Protection applied")
            
        finally:
            ojs_book.save(ojs_path)
            ojs_book.close()
            if not quiet:
                progress.complete(f"✓ {tournament_name} complete!")
            else:
                logger.info(f"✓ Completed: {tournament_name}")

    if not quiet:
        print(f"\n{Fore.GREEN}{'═' * 60}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}  ALL TOURNAMENTS PROCESSED SUCCESSFULLY!  {Style.RESET_ALL}".center(70))
        print(f"{Fore.GREEN}{'═' * 60}{Style.RESET_ALL}\n")
    
    logger.info("All tournaments processed successfully")


if __name__ == "__main__":
    main()