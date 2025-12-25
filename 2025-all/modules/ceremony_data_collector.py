"""Collects data from OJS files for ceremony script generation."""

import logging
from typing import List, Tuple, Dict
from openpyxl import load_workbook
from dataclasses import dataclass

from .constants import (
    SHEET_RESULTS, TABLE_TOURNAMENT_DATA,
    SHEET_TEAM_INFO, TABLE_TEAM_LIST,
    COL_TEAM_NUMBER, COL_TEAM_NAME
)

logger = logging.getLogger("ceremony_generator")


@dataclass
class AwardWinner:
    """Represents an award winner."""
    team_number: int
    team_name: str
    label: str = ""  # Optional label like "1st Place", "Winner", etc.
    score: int = None  # For robot game awards


class HighlightTracker:
    """Tracks alternating highlight classes for dual emcee mode."""
    
    def __init__(self, enabled: bool = True):
        """Initialize highlight tracker.
        
        Args:
            enabled: If True, wrap text in highlight spans. If False, return plain text.
        """
        self.enabled = enabled
        self.current = 0  # Start with highlight0
    
    def wrap(self, text: str) -> str:
        """Wrap text in highlight span and toggle state.
        
        Args:
            text: Text to wrap
            
        Returns:
            Text wrapped in <span class="highlightN"> if enabled, otherwise plain text
        """
        if not self.enabled:
            return text
        
        result = f'<span class="highlight{self.current}">{text}</span>'
        self.current = 1 - self.current  # Toggle 0<->1
        return result
    
    def wrap_paragraph(self, content: str) -> str:
        """Wrap content in <p> tag with highlight span.
        
        Args:
            content: Content to wrap
            
        Returns:
            Content wrapped in <p> and <span> tags
        """
        return f'<p>{self.wrap(content)}</p>'


class CeremonyDataCollector:
    """Collects award and team data from OJS files for ceremony script."""
    
    def __init__(self, config: dict, dual_emcee: bool = False):
        """Initialize data collector.
        
        Args:
            config: Tournament configuration dictionary
            dual_emcee: Whether to enable dual emcee highlighting
        """
        self.config = config
        self.warnings = []
        
        # Initialize highlight tracker
        self.highlight_tracker = HighlightTracker(enabled=dual_emcee)
        
        logger.debug(f"Highlight tracker initialized (enabled={dual_emcee})")
    
    def collect_team_list(self, ojs_path: str, division: str = "") -> List[Tuple[int, str]]:
        """Collect list of teams from OJS file.
        
        Args:
            ojs_path: Path to OJS file
            division: Division name (for logging)
            
        Returns:
            List of (team_number, team_name) tuples
        """
        logger.info(f"Collecting team list from {ojs_path}")
        
        teams = []
        
        try:
            wb = load_workbook(ojs_path, data_only=True)
            ws = wb[SHEET_TEAM_INFO]
            
            # Find the table
            table = ws.tables.get(TABLE_TEAM_LIST)
            if not table:
                logger.error(f"Table {TABLE_TEAM_LIST} not found in {ojs_path}")
                return teams
            
            # Get table range
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            # Read header row to find column indices
            header_row = min_row
            team_num_col = None
            team_name_col = None
            
            for col_idx in range(min_col, max_col + 1):
                header_val = ws.cell(row=header_row, column=col_idx).value
                if header_val == COL_TEAM_NUMBER:
                    team_num_col = col_idx
                elif header_val == COL_TEAM_NAME:
                    team_name_col = col_idx
            
            if team_num_col is None or team_name_col is None:
                logger.error(f"Required columns not found in {TABLE_TEAM_LIST}")
                return teams
            
            # Read data rows
            for row_idx in range(min_row + 1, max_row + 1):
                team_num = ws.cell(row=row_idx, column=team_num_col).value
                team_name = ws.cell(row=row_idx, column=team_name_col).value
                
                if team_num and team_name:
                    teams.append((int(team_num), str(team_name)))
            
            wb.close()
            logger.info(f"Collected {len(teams)} teams from {division if division else 'tournament'}")
            
        except Exception as e:
            logger.error(f"Error collecting team list: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        
        return teams
    
    def collect_advancing_teams(self, ojs_path: str, division: str = "") -> List[Tuple[int, str]]:
        """Collect list of advancing teams from OJS file.
        
        Args:
            ojs_path: Path to OJS file
            division: Division name (for logging)
            
        Returns:
            List of (team_number, team_name) tuples for advancing teams
        """
        logger.info(f"Collecting advancing teams from {ojs_path}")
        
        advancing = []
        
        try:
            wb = load_workbook(ojs_path, data_only=True)
            ws = wb[SHEET_RESULTS]
            
            # Find the table
            table = ws.tables.get(TABLE_TOURNAMENT_DATA)
            if not table:
                logger.error(f"Table {TABLE_TOURNAMENT_DATA} not found in {ojs_path}")
                return advancing
            
            # Get table range
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            # Read header row to find column indices
            header_row = min_row
            team_num_col = None
            team_name_col = None
            advance_col = None
            
            for col_idx in range(min_col, max_col + 1):
                header_val = ws.cell(row=header_row, column=col_idx).value
                if header_val == COL_TEAM_NUMBER or header_val == "Team Number":
                    team_num_col = col_idx
                elif header_val == COL_TEAM_NAME:
                    team_name_col = col_idx
                elif header_val == "Advance?":
                    advance_col = col_idx
            
            if team_num_col is None or team_name_col is None or advance_col is None:
                logger.error(f"Required columns not found in {TABLE_TOURNAMENT_DATA}")
                return advancing
            
            # Read data rows
            for row_idx in range(min_row + 1, max_row + 1):
                advance_val = ws.cell(row=row_idx, column=advance_col).value
                
                if advance_val == "Yes":
                    team_num = ws.cell(row=row_idx, column=team_num_col).value
                    team_name = ws.cell(row=row_idx, column=team_name_col).value
                    
                    if team_num and team_name:
                        advancing.append((int(team_num), str(team_name)))
            
            wb.close()
            logger.info(f"Collected {len(advancing)} advancing teams from {division if division else 'tournament'}")
            
        except Exception as e:
            logger.error(f"Error collecting advancing teams: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        
        return advancing
    
    def collect_robot_game_awards(self, ojs_path: str, count: int, division: str = "") -> List[AwardWinner]:
        """Collect robot game award winners.
        
        Args:
            ojs_path: Path to OJS file
            count: Number of awards to collect
            division: Division name (for logging)
            
        Returns:
            List of AwardWinner objects
        """
        logger.info(f"Collecting top {count} robot game awards from {ojs_path}")
        
        winners = []
        
        try:
            wb = load_workbook(ojs_path, data_only=True)
            ws = wb[SHEET_RESULTS]
            
            # Find the table
            table = ws.tables.get(TABLE_TOURNAMENT_DATA)
            if not table:
                logger.error(f"Table {TABLE_TOURNAMENT_DATA} not found in {ojs_path}")
                return winners
            
            # Get table range
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            # Read header row
            header_row = min_row
            team_num_col = None
            team_name_col = None
            rg_rank_col = None
            rg_score_col = None
            
            # Read all headers for debugging
            all_headers = []
            for col_idx in range(min_col, max_col + 1):
                header_val = ws.cell(row=header_row, column=col_idx).value
                all_headers.append(header_val)
                if header_val == COL_TEAM_NUMBER or header_val == "Team Number":
                    team_num_col = col_idx
                elif header_val == COL_TEAM_NAME:
                    team_name_col = col_idx
                elif header_val == "Robot Game Rank":
                    rg_rank_col = col_idx
                elif header_val == "Max Robot Game Score":
                    rg_score_col = col_idx
            
            logger.debug(f"Table columns: {all_headers}")
            logger.debug(f"Found columns - Team#: {team_num_col}, Name: {team_name_col}, RG Rank: {rg_rank_col}, RG Score: {rg_score_col}")
            
            # Validate that all required columns were found
            if team_num_col is None:
                logger.error(f"Column '{COL_TEAM_NUMBER}' or 'Team Number' not found in {TABLE_TOURNAMENT_DATA}")
                wb.close()
                return winners
            if team_name_col is None:
                logger.error(f"Column '{COL_TEAM_NAME}' not found in {TABLE_TOURNAMENT_DATA}")
                wb.close()
                return winners
            if rg_rank_col is None:
                logger.error(f"Column 'Robot Game Rank' not found in {TABLE_TOURNAMENT_DATA}")
                wb.close()
                return winners
            if rg_score_col is None:
                logger.error(f"Column 'Max Robot Game Score' not found in {TABLE_TOURNAMENT_DATA}")
                wb.close()
                return winners
            
            # Collect teams with ranks 1 through count
            for row_idx in range(min_row + 1, max_row + 1):
                rank = ws.cell(row=row_idx, column=rg_rank_col).value
                
                if rank and 1 <= int(rank) <= count:
                    team_num = ws.cell(row=row_idx, column=team_num_col).value
                    team_name = ws.cell(row=row_idx, column=team_name_col).value
                    score = ws.cell(row=row_idx, column=rg_score_col).value
                    
                    # Determine label based on rank
                    rank_labels = {1: "1st Place", 2: "2nd Place", 3: "3rd Place"}
                    label = rank_labels.get(int(rank), f"{rank}th Place")
                    
                    if team_num and team_name:
                        winners.append(AwardWinner(
                            team_number=int(team_num),
                            team_name=str(team_name),
                            label=label,
                            score=int(score) if score else None
                        ))
            
            wb.close()
            logger.info(f"Collected {len(winners)} robot game winners")
            
        except Exception as e:
            logger.error(f"Error collecting robot game awards: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        
        return winners
    
    def collect_judged_awards(self, ojs_path: str, award: dict, labels: List[str], 
                              division: str, ojs_filename: str) -> List[AwardWinner]:
        """Collect judged award winners.
        
        Args:
            ojs_path: Path to OJS file
            award: Award configuration dict
            labels: List of award labels
            division: Division name
            ojs_filename: OJS filename (for warnings)
            
        Returns:
            List of AwardWinner objects
        """
        award_id = award['ID']
        award_name = award['Name']
        
        logger.info(f"Collecting {award_name} from {ojs_path}")
        
        winners = []
        
        try:
            wb = load_workbook(ojs_path, data_only=True)
            ws = wb[SHEET_RESULTS]
            
            # Find the table
            table = ws.tables.get(TABLE_TOURNAMENT_DATA)
            if not table:
                logger.error(f"Table {TABLE_TOURNAMENT_DATA} not found in {ojs_path}")
                return winners
            
            # Get table range
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            # Read header row
            header_row = min_row
            team_num_col = None
            team_name_col = None
            award_col = None
            
            for col_idx in range(min_col, max_col + 1):
                header_val = ws.cell(row=header_row, column=col_idx).value
                if header_val == COL_TEAM_NUMBER or header_val == "Team Number":
                    team_num_col = col_idx
                elif header_val == COL_TEAM_NAME:
                    team_name_col = col_idx
                elif header_val == "Award":
                    award_col = col_idx
            
            # Collect teams with matching awards
            for label in labels:
                found = False
                for row_idx in range(min_row + 1, max_row + 1):
                    award_val = ws.cell(row=row_idx, column=award_col).value
                    
                    if award_val == label:
                        team_num = ws.cell(row=row_idx, column=team_num_col).value
                        team_name = ws.cell(row=row_idx, column=team_name_col).value
                        
                        if team_num and team_name:
                            winners.append(AwardWinner(
                                team_number=int(team_num),
                                team_name=str(team_name),
                                label=label
                            ))
                            found = True
                            break
                
                if not found and division:  # Only warn for division awards
                    self.warnings.append(
                        f"{ojs_filename}: {award_name} '{label}' not assigned"
                    )
            
            wb.close()
            logger.info(f"Collected {len(winners)} winners for {award_name}")
            
        except Exception as e:
            logger.error(f"Error collecting judged awards: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        
        return winners
    
    def format_team_list_as_html(self, teams: List[Tuple[int, str]]) -> str:
        """Format team list as HTML paragraphs with highlighting.
        
        Args:
            teams: List of (team_number, team_name) tuples
            
        Returns:
            HTML string with team list
        """
        if not teams:
            return ""
        
        html_lines = []
        for team_num, team_name in teams:
            line = f"Team {team_num}, {team_name}"
            html_line = self.highlight_tracker.wrap_paragraph(line)
            html_lines.append(html_line)
        
        return "\n".join(html_lines)
    
    def format_winners_as_html(self, winners: List[AwardWinner], include_score: bool = False) -> str:
        """Format award winners as HTML paragraphs with highlighting.
        
        Args:
            winners: List of AwardWinner objects
            include_score: If True, include robot game score in output
            
        Returns:
            HTML string with formatted winners
        """
        if not winners:
            return ""
        
        html_lines = []
        for winner in winners:
            parts = []
            
            # Add label if present
            if winner.label:
                parts.append(f"{winner.label}:")
            
            # Add team info
            parts.append(f"Team {winner.team_number}, {winner.team_name}")
            
            # Add score if requested
            if include_score and winner.score is not None:
                parts.append(f"with a score of {winner.score}")
            
            line = " ".join(parts)
            html_line = self.highlight_tracker.wrap_paragraph(line)
            html_lines.append(html_line)
        
        return "\n".join(html_lines)
