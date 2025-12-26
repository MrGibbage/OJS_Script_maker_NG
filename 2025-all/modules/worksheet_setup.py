"""Worksheet setup functions for tournament OJS files."""

import os
import logging
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import Rule
import openpyxl.styles.differential
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange
from copy import copy as _copy
from openpyxl.workbook.external_link import ExternalLink
from openpyxl.workbook.defined_name import DefinedName

from .constants import (
    SHEET_PASSWORD, REQUIRED_COLUMNS,
    COL_TEAM_NUMBER, COL_TEAM_NAME, COL_COACH_NAME, COL_POD_NUMBER,
    COL_SHORT_NAME, COL_LONG_NAME, COL_OJS_FILENAME, COL_DIVISION, COL_ADVANCING,
    SHEET_TEAM_INFO, SHEET_AWARD_DROPDOWNS, SHEET_META, SHEET_AWARD_DEF,
    SHEET_ROBOT_GAME, SHEET_INNOVATION, SHEET_ROBOT_DESIGN, SHEET_CORE_VALUES, SHEET_RESULTS,
    TABLE_TEAM_LIST, TABLE_ROBOT_GAME_AWARDS, TABLE_AWARD_DROPDOWNS, TABLE_META, TABLE_AWARD_DEF,
    TABLE_ROBOT_GAME, TABLE_INNOVATION, TABLE_ROBOT_DESIGN, TABLE_CORE_VALUES, TABLE_TOURNAMENT_DATA,
    FILE_CLOSING_CEREMONY, AWARD_COLUMN_PREFIX_JUDGED, AWARD_COLUMN_ROBOT_GAME, AWARD_LABEL_PREFIX
)
from .excel_operations import add_table_dataframe, _to_int
from .logger import print_error


logger = logging.getLogger("ojs_builder")


def set_up_tapi_worksheet(
    tournament: pd.Series,
    book: Workbook,
    dfAssignments: pd.DataFrame,
    using_divisions: bool
) -> bool:
    """Populate the 'Team and Program Information' table.

    Args:
        tournament: A pandas Series representing the tournament row
        book: An open openpyxl Workbook object
        dfAssignments: DataFrame containing team assignments
        using_divisions: Boolean indicating if divisions are used
        
    Returns:
        True if successful, False if no teams assigned (should skip this tournament)
    """
    d = ""
    logger.info(f"Setting up Team and Program Information for {tournament[COL_SHORT_NAME]}")
    
    if using_divisions:
        d = tournament[COL_DIVISION]
        if isinstance(tournament[COL_OJS_FILENAME], float):
            print_error(logger, f"Invalid OJS filename for tournament {tournament[COL_SHORT_NAME]}: "
                       "expected string, got float (possibly missing value)")
        
        assignees: pd.DataFrame = dfAssignments[
            (dfAssignments[COL_SHORT_NAME] == tournament[COL_SHORT_NAME])
            & (dfAssignments[COL_DIVISION] == d)
        ]
        logger.info(f"Found {len(assignees)} teams in {tournament[COL_SHORT_NAME]} {d}")
    else:
        if isinstance(tournament[COL_OJS_FILENAME], float):
            print_error(logger, f"Invalid OJS filename for tournament {tournament[COL_SHORT_NAME]}: "
                       "expected string, got float (possibly missing value)")
        
        assignees: pd.DataFrame = dfAssignments[
            (dfAssignments[COL_SHORT_NAME] == tournament[COL_SHORT_NAME])
        ]
        logger.info(f"Found {len(assignees)} teams in {tournament[COL_SHORT_NAME]}")

    # Check if there are any teams assigned
    if len(assignees) == 0:
        logger.warning(f"No teams assigned to {tournament[COL_SHORT_NAME]} {d} - skipping")
        return False

    # Validate required columns exist
    keep = [COL_TEAM_NUMBER, COL_TEAM_NAME, COL_COACH_NAME]
    keep_safe = [c for c in keep if c in assignees.columns]
    
    if len(keep_safe) != len(keep):
        missing = set(keep) - set(keep_safe)
        print_error(logger, f"Missing required columns in assignments: {', '.join(missing)}")
    
    assignees = assignees[keep_safe]
    assignees[COL_POD_NUMBER] = 0
    sorted_assignees = assignees.sort_values(by=COL_TEAM_NUMBER, ascending=True)
    
    add_table_dataframe(book, SHEET_TEAM_INFO, TABLE_TEAM_LIST, sorted_assignees)
    
    # Now copy formatting from the template row to all data rows
    from openpyxl.styles.protection import Protection
    from openpyxl.utils import range_boundaries
    
    ws = book[SHEET_TEAM_INFO]
    table = ws.tables[TABLE_TEAM_LIST]
    
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    
    # First data row is the template
    first_data_row = min_row + 1
    
    # Copy font and row height from first data row to all other rows
    for col_idx in range(min_col, max_col + 1):
        template_cell = ws.cell(row=first_data_row, column=col_idx)
        
        # Copy font to all data rows
        for row_idx in range(first_data_row + 1, max_row + 1):
            target_cell = ws.cell(row=row_idx, column=col_idx)
            try:
                target_cell.font = _copy(template_cell.font)
            except Exception:
                pass
    
    # Copy row height from template row to all data rows
    try:
        template_height = ws.row_dimensions[first_data_row].height
        if template_height is not None:
            for row_idx in range(first_data_row + 1, max_row + 1):
                ws.row_dimensions[row_idx].height = template_height
            logger.debug(f"Copied row height ({template_height}) to all data rows")
    except Exception:
        pass
    
    # Unlock Pod Number column (column 4/D) so users can edit it
    pod_col_idx = 4  # D is the 4th column
    
    for row_idx in range(min_row + 1, max_row + 1):  # Skip header row
        cell = ws.cell(row=row_idx, column=pod_col_idx)
        cell.protection = Protection(locked=False)
    
    logger.debug(f"Unlocked Pod Number column (D{min_row + 1}:D{max_row})")
    
    return True


def set_up_award_worksheet(
    tournament: pd.Series,
    book: Workbook,
    dfAwardDef: pd.DataFrame,
    using_divisions: bool
) -> None:
    """Prepare award tables used by OJS dropdowns and closing scripts.

    Args:
        tournament: A pandas Series representing the tournament row
        book: An open openpyxl Workbook object
        dfAwardDef: DataFrame containing award definitions
        using_divisions: Boolean indicating if divisions are used
    """
    thisDiv = tournament[COL_DIVISION] if using_divisions else ""
    logger.info(f"Setting up awards for {tournament[COL_SHORT_NAME]} {thisDiv}")
    
    # Get available Label columns dynamically
    label_columns = [col for col in dfAwardDef.columns if col.startswith(AWARD_LABEL_PREFIX)]
    logger.debug(f"Found {len(label_columns)} label columns: {label_columns}")

    # Robot Game awards
    rg_raw = (
        tournament.get(AWARD_COLUMN_ROBOT_GAME)
        if hasattr(tournament, "get")
        else tournament[AWARD_COLUMN_ROBOT_GAME]
    )
    try:
        rg_awards = int(rg_raw) if not pd.isna(rg_raw) else 0
    except (ValueError, TypeError) as e:
        print_error(logger, f"Invalid robot game award count for {tournament[COL_SHORT_NAME]}: {rg_raw}", e)
        
    rg_awards_df = pd.DataFrame(columns=["Robot Game Awards"])
    for rg_award_num in range(1, rg_awards + 1):
        thisLabel = AWARD_LABEL_PREFIX + str(rg_award_num)
        
        # Check if this Label column exists
        if thisLabel not in label_columns:
            logger.warning(f"Label column '{thisLabel}' not found in AwardDef table, using blank value")
            thisValue = None
        else:
            sel = dfAwardDef.loc[dfAwardDef["ColumnName"] == AWARD_COLUMN_ROBOT_GAME, thisLabel]
            try:
                thisValue = sel.iat[0]
            except (IndexError, KeyError):
                logger.warning(f"Could not find value for {AWARD_COLUMN_ROBOT_GAME} in column {thisLabel}")
                thisValue = None
        
        rg_awards_df.loc[len(rg_awards_df)] = [thisValue]

    add_table_dataframe(book, SHEET_AWARD_DROPDOWNS, TABLE_ROBOT_GAME_AWARDS, rg_awards_df)

    # Other judged awards - now with ID column
    j_cols = tournament.filter(regex=f"^{AWARD_COLUMN_PREFIX_JUDGED}")
    j_awards_df = pd.DataFrame(columns=["Award", "ID"])
    
    for this_col_name, series in j_cols.items():
        count = _to_int(series)
        for award_num in range(1, count + 1):
            label_col = AWARD_LABEL_PREFIX + str(award_num)
            
            # Check if this Label column exists
            if label_col not in label_columns:
                logger.warning(f"Label column '{label_col}' not found in AwardDef table, using blank value")
                thisValue = None
            else:
                sel = dfAwardDef.loc[dfAwardDef["ColumnName"] == this_col_name, label_col]
                try:
                    thisValue = sel.iat[0]
                except (IndexError, KeyError):
                    logger.warning(f"Could not find value for {this_col_name} in column {label_col}")
                    thisValue = None
            
            j_awards_df.loc[len(j_awards_df)] = [thisValue, this_col_name]
            
    add_table_dataframe(book, SHEET_AWARD_DROPDOWNS, TABLE_AWARD_DROPDOWNS, j_awards_df)
    logger.debug(f"Added {len(rg_awards_df)} robot game awards and {len(j_awards_df)} judged awards")


def set_up_meta_worksheet(
    tournament: pd.Series,
    book: Workbook,
    config: dict,
    tournament_folder: str,
    using_divisions: bool
) -> None:
    """Populate the metadata worksheet with tournament information."""
    d = tournament[COL_DIVISION] if using_divisions else ""
    logger.info(f"Setting up metadata for {tournament[COL_SHORT_NAME]} {d}")
    
    scriptfile = os.path.join(
        tournament_folder,
        tournament[COL_SHORT_NAME],
        FILE_CLOSING_CEREMONY
    )
    
    # Validate required config keys
    required_keys = ["season_yr", "season_name"]
    for key in required_keys:
        if key not in config:
            print_error(logger, f"Missing required configuration key: '{key}' in {CONFIG_FILENAME}")
    
    dfMeta = pd.DataFrame(columns=["Key", "Value"])
    dfMeta.loc[len(dfMeta)] = {"Key": "Tournament Year", "Value": config["season_yr"]}
    dfMeta.loc[len(dfMeta)] = {"Key": "FLL Season Title", "Value": config["season_name"]}
    dfMeta.loc[len(dfMeta)] = {"Key": "Tournament Short Name", "Value": tournament[COL_SHORT_NAME]}
    dfMeta.loc[len(dfMeta)] = {"Key": "Tournament Long Name", "Value": tournament[COL_LONG_NAME]}
    dfMeta.loc[len(dfMeta)] = {"Key": "Completed Script File", "Value": scriptfile}
    dfMeta.loc[len(dfMeta)] = {"Key": "Using Divisions", "Value": using_divisions}
    
    if using_divisions:
        dfMeta.loc[len(dfMeta)] = {"Key": "Division", "Value": d}
        
    dfMeta.loc[len(dfMeta)] = {"Key": "Advancing", "Value": tournament[COL_ADVANCING]}
    
    if tournament[COL_OJS_FILENAME] is not None:
        add_table_dataframe(book, SHEET_META, TABLE_META, dfMeta, debug=False)


def copy_team_numbers(
    source_sheet: Worksheet,
    target_sheet: Worksheet,
    target_start_row: int,
    source_start_row: int = 3,
    debug: bool = False,
) -> int:
    """Copy team numbers from column A of source to target sheet.
    
    Args:
        source_sheet: Source worksheet
        target_sheet: Target worksheet
        target_start_row: Row to start writing in target
        source_start_row: Row to start reading from source
        debug: If True, log diagnostic information
        
    Returns:
        Number of rows copied
    """
    col = 1
    max_row = source_sheet.max_row

    if debug:
        logger.debug(
            f"copy_team_numbers: source={source_sheet.title}, target={target_sheet.title}, "
            f"max_row={max_row}, source_start={source_start_row}, target_start={target_start_row}"
        )

    last_row = None
    for r in range(source_start_row, max_row + 1):
        v = source_sheet.cell(row=r, column=col).value
        if v is not None and v != "":
            last_row = r

    if last_row is None:
        logger.debug("copy_team_numbers: no non-empty rows found")
        return 0

    dest_row = target_start_row
    copied = 0
    for r in range(source_start_row, last_row + 1):
        cell_value = source_sheet.cell(row=r, column=col).value
        
        if pd.isna(cell_value):
            write_val = None
        elif isinstance(cell_value, (np.integer,)):
            write_val = int(cell_value)
        elif isinstance(cell_value, (np.floating,)):
            fv = float(cell_value)
            write_val = int(fv) if fv.is_integer() else fv
        else:
            write_val = cell_value

        target_sheet.cell(row=dest_row, column=col).value = write_val
        dest_row += 1
        copied += 1

    logger.debug(f"copy_team_numbers: copied {copied} rows")
    return copied


def _copy_data_validations_for_range(
    ws: Worksheet,
    start_col_idx: int,
    end_col_idx: int,
    first_data_row: int,
    new_end_row: int,
):
    """Duplicate data validations from template row to new range.
    
    Args:
        ws: Worksheet to modify
        start_col_idx: Starting column index
        end_col_idx: Ending column index
        first_data_row: Template data row number
        new_end_row: New ending row number
    """
    try:
        existing = list(ws.data_validations.dataValidation)
    except Exception:
        existing = []

    for dv in existing:
        try:
            for rng in dv.ranges:
                try:
                    cr = CellRange(str(rng))
                except Exception:
                    continue

                if not (cr.min_row <= first_data_row <= cr.max_row):
                    continue

                orig_min_col = cr.min_col
                orig_max_col = cr.max_col
                new_min_col = max(orig_min_col, start_col_idx)
                new_max_col = min(orig_max_col, end_col_idx)

                if new_min_col > new_max_col:
                    continue

                new_range = f"{get_column_letter(new_min_col)}{first_data_row}:{get_column_letter(new_max_col)}{new_end_row}"

                newdv = DataValidation(
                    type=dv.type,
                    formula1=getattr(dv, "formula1", None),
                    formula2=getattr(dv, "formula2", None),
                    allow_blank=getattr(dv, "allow_blank", None),
                    operator=getattr(dv, "operator", None),
                    showDropDown=getattr(dv, "showDropDown", None),
                    error=getattr(dv, "error", None),
                    errorTitle=getattr(dv, "errorTitle", None),
                    prompt=getattr(dv, "prompt", None),
                    promptTitle=getattr(dv, "promptTitle", None),
                )

                try:
                    newdv.add(new_range)
                    ws.add_data_validation(newdv)
                except Exception:
                    continue
        except Exception:
            continue


def resize_worksheets(
    tournament: pd.Series,
    book: Workbook,
    dfAssignments: pd.DataFrame,
    using_divisions: bool
) -> None:
    """Resize all tables in the workbook based on number of teams.
    
    Args:
        tournament: A pandas Series representing the tournament row
        book: An open openpyxl Workbook object
        dfAssignments: DataFrame containing team assignments
        using_divisions: Whether the tournament uses divisions
    """
    logger.info(f"Resizing worksheets for {book.properties.title}")
    
    worksheetNames = [
        SHEET_ROBOT_GAME,
        SHEET_INNOVATION,
        SHEET_ROBOT_DESIGN,
        SHEET_CORE_VALUES,
        SHEET_RESULTS,
    ]
    worksheetTables = [
        TABLE_ROBOT_GAME,
        TABLE_INNOVATION,
        TABLE_ROBOT_DESIGN,
        TABLE_CORE_VALUES,
        TABLE_TOURNAMENT_DATA,
    ]
    worksheet_start_row = [2, 2, 2, 2, 3]
    
    div = tournament.get(COL_DIVISION, None)
    
    if using_divisions:
        assignees: pd.DataFrame = dfAssignments[
            (dfAssignments[COL_SHORT_NAME] == tournament[COL_SHORT_NAME])
            & (dfAssignments[COL_DIVISION] == div)
        ]
    else:
        assignees: pd.DataFrame = dfAssignments[
            (dfAssignments[COL_SHORT_NAME] == tournament[COL_SHORT_NAME])
        ]

    # Copy team numbers to each worksheet
    sheet_tables = zip(worksheetNames, worksheetTables, worksheet_start_row)
    copied_counts: dict[str, int] = {}
    for s, t, r in sheet_tables:
        if s in book.sheetnames:
            ws = book[s]
            tapi_sheet = book[SHEET_TEAM_INFO]
            copied = copy_team_numbers(
                source_sheet=tapi_sheet,
                target_sheet=ws,
                target_start_row=r,
                debug=False,
            )
            copied_counts[s] = copied

    # Resize the tables
    sheet_tables = zip(worksheetNames, worksheetTables, worksheet_start_row)
    for s, t, r in sheet_tables:
        if s not in book.sheetnames:
            continue
        ws = book[s]
        table: Table = ws.tables[t]
        table_range: str = table.ref

        start_cell, end_cell = table_range.split(":")
        start_col_letter, start_row_num = coordinate_from_string(start_cell)
        end_col_letter, end_row_num = coordinate_from_string(end_cell)

        rows_for_table = copied_counts.get(s, len(assignees))
        new_end_row = start_row_num + rows_for_table

        table.ref = f"{start_col_letter}{start_row_num}:{end_col_letter}{new_end_row}"
        logger.debug(f"Resized table {t}: {table.ref}")

        # Copy formulas
        first_data_row = start_row_num + 1
        start_col_idx = column_index_from_string(start_col_letter)
        end_col_idx = column_index_from_string(end_col_letter)
        
        for col_idx in range(start_col_idx + 1, end_col_idx + 1):
            template = ws.cell(row=first_data_row, column=col_idx).value
            if isinstance(template, str) and template.startswith("="):
                for rr in range(first_data_row, new_end_row + 1):
                    ws.cell(row=rr, column=col_idx).value = template

        # Copy cell protection
        for col_idx in range(start_col_idx, end_col_idx + 1):
            template_cell = ws.cell(row=first_data_row, column=col_idx)
            tpl_prot = getattr(template_cell, "protection", None)
            if tpl_prot is None:
                continue
            for rr in range(first_data_row, new_end_row + 1):
                tgt = ws.cell(row=rr, column=col_idx)
                try:
                    tgt.protection = _copy(tpl_prot)
                except Exception:
                    pass

        # Copy cell styles
        for col_idx in range(start_col_idx, end_col_idx + 1):
            template_cell = ws.cell(row=first_data_row, column=col_idx)
            for rr in range(first_data_row, new_end_row + 1):
                tgt = ws.cell(row=rr, column=col_idx)
                try:
                    if hasattr(template_cell, "_style"):
                        tgt._style = _copy(template_cell._style)
                    tgt.number_format = template_cell.number_format
                    tgt.font = _copy(template_cell.font)
                    tgt.fill = _copy(template_cell.fill)
                    tgt.alignment = _copy(template_cell.alignment)
                    tgt.border = _copy(template_cell.border)
                except Exception:
                    pass

        # Copy row height
        try:
            template_height = ws.row_dimensions[first_data_row].height
            if template_height is not None:
                for rr in range(first_data_row, new_end_row + 1):
                    ws.row_dimensions[rr].height = template_height
        except Exception:
            pass

        # Copy data validations
        try:
            _copy_data_validations_for_range(
                ws, start_col_idx, end_col_idx, first_data_row, new_end_row
            )
        except Exception:
            pass

        # Delete extra rows
        ws.delete_rows(idx=new_end_row + 1, amount=200)

    logger.debug("Worksheet resizing complete")


def copy_award_def(tournament: pd.Series, book: Workbook, dfAwardDef: pd.DataFrame) -> None:
    """Add the award definitions table to the workbook with counts from tournament row.
    
    Args:
        tournament: A pandas Series representing the tournament row
        book: An open openpyxl Workbook object
        dfAwardDef: DataFrame containing award definitions
    """
    logger.debug("Adding award definitions table with tournament counts")
    
    # Make a copy to avoid modifying the original
    dfAwardDef_copy = dfAwardDef.copy()
    
    # Drop old count columns if they exist (they've been replaced by 'Count')
    columns_to_drop = ['D1Count', 'D2Count', 'TournCount']
    for col in columns_to_drop:
        if col in dfAwardDef_copy.columns:
            dfAwardDef_copy = dfAwardDef_copy.drop(columns=[col])
            logger.debug(f"Dropped obsolete column: {col}")
    
    # Replace 0 with empty string for Label* and ScriptTag* columns
    for col in dfAwardDef_copy.columns:
        if col.startswith('Label') or col.startswith('ScriptTag'):
            # Replace 0, 0.0, and '0' with empty string
            dfAwardDef_copy[col] = dfAwardDef_copy[col].replace([0, 0.0, '0', '0.0'], '')
            # Also replace NaN with empty string
            dfAwardDef_copy[col] = dfAwardDef_copy[col].fillna('')
    
    # Convert DivAward column from 1/0 to TRUE/FALSE
    if 'DivAward' in dfAwardDef_copy.columns:
        dfAwardDef_copy['DivAward'] = dfAwardDef_copy['DivAward'].map({
            1: 'TRUE', 0: 'FALSE', 
            True: 'TRUE', False: 'FALSE',
            '1': 'TRUE', '0': 'FALSE',
            'TRUE': 'TRUE', 'FALSE': 'FALSE'
        })
    
    # Add Count column if it doesn't exist
    if 'Count' not in dfAwardDef_copy.columns:
        dfAwardDef_copy['Count'] = None
    
    # Populate Count column from tournament row
    for idx, row in dfAwardDef_copy.iterrows():
        column_name = row['ColumnName']
        
        # Get the count from the tournament row for this award
        if column_name in tournament.index:
            count = tournament.get(column_name, 0)
            
            # Convert to int, handling NaN
            try:
                count = int(count) if not pd.isna(count) else 0
            except (ValueError, TypeError):
                count = 0
            
            dfAwardDef_copy.at[idx, 'Count'] = count
            logger.debug(f"Set Count={count} for award {column_name}")
        else:
            # Award column not in tournament row, default to 0
            dfAwardDef_copy.at[idx, 'Count'] = 0
            logger.debug(f"Award {column_name} not in tournament row, set Count=0")
    
    add_table_dataframe(book, SHEET_AWARD_DEF, TABLE_AWARD_DEF, dfAwardDef_copy)


def add_essential_conditional_formats(book: Workbook, num_teams: int) -> None:
    """Add essential conditional formatting to Results and Rankings sheet.
    
    Args:
        book: An open openpyxl Workbook object
        num_teams: Number of teams (to determine range)
    """
    logger.info("Adding essential conditional formatting")
    
    try:
        ws = book[SHEET_RESULTS]
        
        # Find the TournamentData table
        table = None
        for tbl in ws.tables.values():
            if tbl.name == TABLE_TOURNAMENT_DATA:
                table = tbl
                break
        
        if not table:
            logger.warning(f"Table {TABLE_TOURNAMENT_DATA} not found, skipping CF")
            return
        
        from openpyxl.utils import range_boundaries, get_column_letter
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        
        # Get count of Robot Game awards from RobotGameAwards table
        rg_award_count = 0
        try:
            rg_ws = book[SHEET_AWARD_DROPDOWNS]
            rg_table = None
            for tbl in rg_ws.tables.values():
                if tbl.name == TABLE_ROBOT_GAME_AWARDS:
                    rg_table = tbl
                    break
            
            if rg_table:
                rg_min_col, rg_min_row, rg_max_col, rg_max_row = range_boundaries(rg_table.ref)
                # Count rows excluding header
                rg_award_count = rg_max_row - rg_min_row
                logger.debug(f"Found {rg_award_count} Robot Game awards")
        except Exception as e:
            logger.warning(f"Could not count Robot Game awards: {e}")
        
        # Define ALL column variables at the beginning
        # Column J (Robot Game Rank) is the 10th column
        rg_rank_col = get_column_letter(10)
        # Column O (Champion's Rank) is the 15th column
        champ_rank_col = get_column_letter(15)
        # Column P (Award) is the 16th column
        award_col = get_column_letter(16)
        # Column Q (Advance?) is the 17th column
        advance_col = get_column_letter(17)
        # Column W is the 23rd column
        col_w = get_column_letter(23)
        
        # Add rules in REVERSE order of priority
        # Excel applies CF rules from bottom to top, so add low-priority first
        
        # Rule 1 (LOW PRIORITY): Award column header - bright green when all selected
        award_list_ws = book[SHEET_AWARD_DROPDOWNS]
        award_list_table = None
        for tbl in award_list_ws.tables.values():
            if tbl.name == TABLE_AWARD_DROPDOWNS:
                award_list_table = tbl
                break
        
        if award_list_table:
            list_min_col, list_min_row, list_max_col, list_max_row = range_boundaries(award_list_table.ref)
            
            bright_green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            purple_italic_font = Font(color="800080", italic=True, size=12)
            
            # Build the formula WITHOUT the sheet name wrapper - just the range
            # Excel formula: =COUNTA(AwardListDropdowns!$A$2:$A$7)=COUNTA($P$3:$P$7)
            award_range = f"${award_col}${min_row + 1}:${award_col}${max_row}"
            
            # For the list range, check if sheet name has spaces
            sheet_name = SHEET_AWARD_DROPDOWNS
            if ' ' in sheet_name:
                list_range = f"'{sheet_name}'!$A${list_min_row + 1}:$A${list_max_row}"
            else:
                list_range = f"{sheet_name}!$A${list_min_row + 1}:$A${list_max_row}"
            
            formula = f'COUNTA({list_range})=COUNTA({award_range})'
            
            logger.debug(f"CF Formula: {formula}")
            
            header_rule = Rule(type="expression", formula=[formula], stopIfTrue=False)
            header_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=bright_green_fill, font=purple_italic_font)
            
            header_cell = f"{award_col}{min_row}"
            ws.conditional_formatting.add(header_cell, header_rule)
            
            logger.info(f"Added 'all awards selected' CF to {header_cell}")
        else:
            logger.warning(f"Table {TABLE_AWARD_DROPDOWNS} not found, skipping Award column CF")
        
        # Rule 2 (LOW PRIORITY): Duplicate awards - red
        bright_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        purple_italic_font = Font(color="800080", italic=True, size=11)
        
        award_data_range = f"{award_col}{min_row + 1}:{award_col}{max_row}"
        duplicate_formula = f'COUNTIF(${award_col}${min_row + 1}:${award_col}${max_row},{award_col}{min_row + 1})>1'
        
        duplicate_rule = Rule(type="expression", formula=[duplicate_formula], stopIfTrue=False)
        duplicate_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=bright_red_fill, font=purple_italic_font)
        
        ws.conditional_formatting.add(award_data_range, duplicate_rule)
        
        logger.info(f"Added duplicate detection CF to {award_data_range}")
        logger.debug(f"Duplicate CF Formula: {duplicate_formula}")
        
        # Rule 3 (LOW PRIORITY): Champion's Rank - blue for top N
        medium_blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        white_font = Font(color="FFFFFF")
        
        champ_rank_range = f"{champ_rank_col}{min_row + 1}:{champ_rank_col}{max_row}"
        champ_rank_formula = f'{champ_rank_col}{min_row + 1}<=${advance_col}$1'
        
        champ_rank_rule = Rule(type="expression", formula=[champ_rank_formula], stopIfTrue=False)
        champ_rank_rule.dxf = openpyxl.styles.differential.DifferentialStyle(
            fill=medium_blue_fill, 
            font=white_font
        )
        
        ws.conditional_formatting.add(champ_rank_range, champ_rank_rule)
        logger.debug(f"Champion's Rank CF Formula: {champ_rank_formula}")
        
        # Rule 4 (LOW PRIORITY): Q1 and Q2 header - green when all advancing selected
        bright_green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Apply to Q1 (header) and Q2 (count cell)
        # Formula: COUNTIF($Q$3:$Q$max,"Yes")=$Q$1
        # This counts "Yes" values and compares to the advancing count in Q1
        advance_header_range = f"{advance_col}1:{advance_col}2"
        advance_count_formula = f'COUNTIF(${advance_col}${min_row + 1}:${advance_col}${max_row},"Yes")=${advance_col}$1'
        
        advance_header_rule = Rule(type="expression", formula=[advance_count_formula], stopIfTrue=False)
        advance_header_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=bright_green_fill)
        
        ws.conditional_formatting.add(advance_header_range, advance_header_rule)
        
        logger.info(f"Added Advance? header highlighting CF to {advance_header_range}")
        logger.debug(f"Advance header CF Formula: {advance_count_formula}")
        
        # Rule 5 (LOW PRIORITY): W1 - green when correct advancing + one Alt
        bright_green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Apply to W1 only
        w1_cell = f"{col_w}1"
        
        # Formula: Check two conditions:
        # 1. COUNTIF($Q$3:$Q$max,"Yes")=$Q$1 (correct number of Yes)
        # 2. COUNTIF($Q$3:$Q$max,"Alt")=1 (exactly one Alt)
        w1_formula = f'AND(COUNTIF(${advance_col}${min_row + 1}:${advance_col}${max_row},"Yes")=${advance_col}$1,COUNTIF(${advance_col}${min_row + 1}:${advance_col}${max_row},"Alt")=1)'
        
        w1_rule = Rule(type="expression", formula=[w1_formula], stopIfTrue=False)
        w1_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=bright_green_fill)
        
        ws.conditional_formatting.add(w1_cell, w1_rule)
        
        logger.info(f"Added W1 Alt team highlighting CF to {w1_cell}")
        logger.debug(f"W1 CF Formula: {w1_formula}")
        
        # Rule 6 (MEDIUM PRIORITY): Yellow row highlight when award selected
        # Apply to entire row BUT exclude columns J (RG Rank), P (Award), Q (Advance?)
        # We'll add the rule in multiple segments to skip those columns
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        row_highlight_formula = f'${award_col}{min_row + 1}<>""'
        row_highlight_rule = Rule(type="expression", formula=[row_highlight_formula], stopIfTrue=False)
        row_highlight_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=yellow_fill)
        
        start_col_letter = get_column_letter(min_col)
        end_col_letter = get_column_letter(max_col)
        
        # Apply yellow to columns A-I (before Robot Game Rank in column J)
        range_before_j = f"{start_col_letter}{min_row + 1}:I{max_row}"
        ws.conditional_formatting.add(range_before_j, row_highlight_rule)
        
        # Apply yellow to columns K-O (between RG Rank and Award)
        range_k_to_o = f"K{min_row + 1}:O{max_row}"
        ws.conditional_formatting.add(range_k_to_o, row_highlight_rule)
        
        # Apply yellow to column R onwards (after Advance?)
        range_after_q = f"R{min_row + 1}:{end_col_letter}{max_row}"
        ws.conditional_formatting.add(range_after_q, row_highlight_rule)
        
        logger.debug(f"Row highlight CF Formula: {row_highlight_formula}")
        logger.debug(f"Applied to ranges: {range_before_j}, {range_k_to_o}, {range_after_q}")
        
        # Rules 7-11 (HIGH PRIORITY): Column-specific highlights
        # Now these can be applied without competition
        
        # Rule 7: Robot Game Gold (Column J)
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        rg_rank_range = f"{rg_rank_col}{min_row + 1}:{rg_rank_col}{max_row}"
        
        gold_formula = f'{rg_rank_col}{min_row + 1}=1'
        gold_rule = Rule(type="expression", formula=[gold_formula], stopIfTrue=False)
        gold_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=gold_fill)
        ws.conditional_formatting.add(rg_rank_range, gold_rule)
        logger.debug(f"Added Robot Game Gold CF: {gold_formula}")
        
        # Rule 8: Robot Game Silver (Column J)
        if rg_award_count >= 2:
            silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            silver_formula = f'{rg_rank_col}{min_row + 1}=2'
            silver_rule = Rule(type="expression", formula=[silver_formula], stopIfTrue=False)
            silver_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=silver_fill)
            ws.conditional_formatting.add(rg_rank_range, silver_rule)
            logger.debug(f"Added Robot Game Silver CF: {silver_formula}")
        
        # Rule 9: Robot Game Bronze (Column J)
        if rg_award_count >= 3:
            bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
            # Bronze should highlight ranks from 3 up to the total robot game award count
            # For example: if rg_award_count=3, highlight rank 3; if rg_award_count=5, highlight ranks 3-5
            if rg_award_count == 3:
                bronze_formula = f'{rg_rank_col}{min_row + 1}=3'
            else:
                bronze_formula = f'AND({rg_rank_col}{min_row + 1}>=3,{rg_rank_col}{min_row + 1}<={rg_award_count})'
            bronze_rule = Rule(type="expression", formula=[bronze_formula], stopIfTrue=False)
            bronze_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=bronze_fill)
            ws.conditional_formatting.add(rg_rank_range, bronze_rule)
            logger.debug(f"Added Robot Game Bronze CF: {bronze_formula}")
        
        # Rule 10: Award column duplicates (Column P) - already added above as Rule 2
        
        # Rule 11: Advance "Yes" - medium blue (Column Q)
        medium_blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        white_font = Font(color="FFFFFF")
        
        advance_data_range = f"{advance_col}{min_row + 1}:{advance_col}{max_row}"
        
        yes_formula = f'{advance_col}{min_row + 1}="Yes"'
        yes_rule = Rule(type="expression", formula=[yes_formula], stopIfTrue=False)
        yes_rule.dxf = openpyxl.styles.differential.DifferentialStyle(
            fill=medium_blue_fill,
            font=white_font
        )
        ws.conditional_formatting.add(advance_data_range, yes_rule)
        logger.debug(f"Added 'Yes' highlighting CF: {yes_formula}")
        
        # Rule 12: Advance "Alt" - light blue (Column Q)
        light_blue_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        
        alt_formula = f'{advance_col}{min_row + 1}="Alt"'
        alt_rule = Rule(type="expression", formula=[alt_formula], stopIfTrue=False)
        alt_rule.dxf = openpyxl.styles.differential.DifferentialStyle(fill=light_blue_fill)
        ws.conditional_formatting.add(advance_data_range, alt_rule)
        logger.debug(f"Added 'Alt' highlighting CF: {alt_formula}")
        
        logger.info(f"Added all conditional formatting rules")
        
    except Exception as e:
        logger.warning(f"Could not add conditional formatting: {e}")
        import traceback
        logger.debug(traceback.format_exc())


def protect_worksheets(tournament: pd.Series, book: Workbook) -> None:
    """Apply protection settings to every worksheet.
    
    Args:
        tournament: Tournament row
        book: Workbook to protect
    """
    logger.info(f"Protecting worksheets for {tournament[COL_OJS_FILENAME]}")
    for ws in book.worksheets:
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = False
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.autoFilter = False
        ws.protection.sort = False
        ws.protection.set_password(SHEET_PASSWORD)
        ws.protection.enable()
    logger.debug("Worksheet protection applied")


def hide_worksheets(tournament: pd.Series, book: Workbook) -> None:
    """Hide utility worksheets.
    
    Args:
        tournament: Tournament row
        book: Workbook
    """
    logger.info(f"Hiding worksheets for {tournament[COL_OJS_FILENAME]}")
    worksheetNames = ["Data Validation", SHEET_META, SHEET_AWARD_DROPDOWNS, SHEET_AWARD_DEF]
    
    for sheetname in worksheetNames:
        if sheetname in book.sheetnames:
            ws = book[sheetname]
            ws.sheet_state = "hidden"
            logger.debug(f"Hid worksheet: {sheetname}")


def remove_external_links(book: Workbook) -> None:
    """Remove all external workbook links from the workbook.
    
    Args:
        book: An open openpyxl Workbook object
    """
    logger.debug("Checking for external workbook links")
    
    removed_count = 0
    
    # Method 1: Check _external_links attribute
    if hasattr(book, '_external_links') and book._external_links:
        original_count = len(book._external_links)
        logger.debug(f"Found {original_count} external link(s) in _external_links")
        book._external_links = []
        removed_count += original_count
    
    # Method 2: Check defined names (which can contain external references)
    if hasattr(book, 'defined_names'):
        names_to_remove = []
        
        for name_obj in book.defined_names.values():
            if hasattr(name_obj, 'value') and name_obj.value:
                # Check for external reference pattern: [filename.xlsx]
                if '[' in str(name_obj.value) and ']' in str(name_obj.value):
                    logger.debug(f"Found external reference in defined name: {name_obj.name} = {name_obj.value}")
                    names_to_remove.append(name_obj.name)
                    removed_count += 1
        
        for name in names_to_remove:
            del book.defined_names[name]
    
    # Method 3: Fix conditional formatting formulas that reference old sheet names
    # ONLY replace "AwardList!" if it's NOT already "AwardListDropdowns!"
    try:
        for ws in book.worksheets:
            if hasattr(ws, 'conditional_formatting'):
                for cf_range in list(ws.conditional_formatting._cf_rules.keys()):
                    rules = ws.conditional_formatting._cf_rules[cf_range]
                    
                    for rule in rules:
                        if hasattr(rule, 'formula') and rule.formula:
                            for i, formula in enumerate(rule.formula):
                                if formula:
                                    formula_str = str(formula)
                                    # Only replace if it has the OLD sheet name, not the new one
                                    if 'AwardList!' in formula_str and 'AwardListDropdowns!' not in formula_str:
                                        new_formula = formula_str.replace('AwardList!', 'AwardListDropdowns!')
                                        rule.formula[i] = new_formula
                                        logger.debug(f"Fixed CF formula: {formula} -> {new_formula}")
                
    except Exception as e:
        logger.debug(f"Could not update conditional formatting: {e}")
    
    if removed_count > 0:
        logger.info(f"Removed {removed_count} external reference(s)")
    else:
        logger.debug("No external workbook links found")


def fix_named_ranges(book: Workbook) -> None:
    """Recreate named ranges that reference renamed sheets.
    
    Args:
        book: An open openpyxl Workbook object
    """
    logger.debug("Fixing named ranges")
    
    try:
        # Fix the "Awards" named range to point to AwardListDropdowns sheet
        
        # First, remove the old "Awards" named range if it exists
        if "Awards" in book.defined_names:
            del book.defined_names["Awards"]
            logger.debug("Removed old 'Awards' named range")
        
        # Find the AwardListDropdowns table to get the correct range
        if SHEET_AWARD_DROPDOWNS in book.sheetnames:
            ws = book[SHEET_AWARD_DROPDOWNS]
            
            # Find the table
            table = None
            for tbl in ws.tables.values():
                if tbl.name == TABLE_AWARD_DROPDOWNS:
                    table = tbl
                    break
            
            if table:
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                
                # Assuming "Award" is the first column (column A)
                # Create named range pointing to the Award column (excluding header)
                award_range = f"'{SHEET_AWARD_DROPDOWNS}'!$A${min_row + 1}:$A${max_row}"
                
                # Create the named range - use dictionary assignment, not append
                defn = DefinedName("Awards", attr_text=award_range)
                book.defined_names["Awards"] = defn
                
                logger.info(f"Created 'Awards' named range: {award_range}")
            else:
                logger.warning(f"Table {TABLE_AWARD_DROPDOWNS} not found")
        else:
            logger.warning(f"Sheet {SHEET_AWARD_DROPDOWNS} not found")
        
        # Also fix any other named ranges that reference "AwardList" (old sheet name)
        # Iterate using the correct method for DefinedNameDict
        for name_key, name_obj in book.defined_names.items():
            if hasattr(name_obj, 'value') and name_obj.value and 'AwardList!' in str(name_obj.value):
                old_value = name_obj.value
                name_obj.value = str(name_obj.value).replace('AwardList!', 'AwardListDropdowns!')
                logger.debug(f"Updated named range '{name_obj.name}': {old_value} -> {name_obj.value}")
    
    except Exception as e:
        logger.warning(f"Could not fix named ranges: {e}")
        import traceback
        logger.debug(traceback.format_exc())
