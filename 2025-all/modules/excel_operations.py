"""Excel file reading and table manipulation operations."""

import os
import logging
from typing import Any
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from .logger import print_error
from .constants import REQUIRED_COLUMNS


logger = logging.getLogger("ojs_builder")


def _to_int(val: Any, default: int = 0) -> int:
    """Safely coerce value to an int.
    
    Accepts pandas.Series, numpy arrays, lists/tuples (uses first element), or scalars.
    Returns `default` on missing/NaN/uncoercible values.
    """
    try:
        if isinstance(val, pd.Series):
            if val.empty:
                return default
            v = val.iat[0]
        elif isinstance(val, np.ndarray):
            if val.size == 0:
                return default
            v = val.flat[0]
        elif isinstance(val, (list, tuple)):
            if len(val) == 0:
                return default
            v = val[0]
        else:
            v = val

        if pd.isna(v):
            return default
        return int(v)
    except Exception:
        return default


def check_workbook_is_closed(xlsx_path: str) -> bool:
    """Check if an Excel workbook is closed by looking for temporary lock files.
    
    Excel creates a temporary lock file (starting with ~$) when a workbook is open.
    
    Args:
        xlsx_path: Path to the Excel workbook
        
    Returns:
        True if workbook appears to be closed, False if it's open
    """
    if not os.path.exists(xlsx_path):
        logger.warning(f"Workbook not found: {xlsx_path}")
        return True
    
    directory = os.path.dirname(xlsx_path)
    filename = os.path.basename(xlsx_path)
    lock_filename = f"~${filename}"
    lock_path = os.path.join(directory, lock_filename)
    
    is_closed = not os.path.exists(lock_path)
    
    if not is_closed:
        logger.debug(f"Lock file detected for {filename}: {lock_filename}")
    
    return is_closed


def verify_workbooks_closed(*workbook_paths: str) -> None:
    """Verify that all specified workbooks are closed.
    
    Raises an error if any workbook appears to be open (has a temporary lock file).
    
    Args:
        *workbook_paths: Variable number of paths to Excel workbooks
        
    Raises:
        RuntimeError: If any workbook is currently open
    """
    open_workbooks = []
    
    for path in workbook_paths:
        if not check_workbook_is_closed(path):
            open_workbooks.append(os.path.basename(path))
    
    if open_workbooks:
        files_str = ", ".join(open_workbooks)
        raise RuntimeError(
            f"The following workbook(s) must be closed before proceeding: {files_str}\n"
            "Please close them in Excel and try again."
        )
    
    logger.debug(f"Verified {len(workbook_paths)} workbook(s) are closed")


def read_table_as_df(
    xlsx_path: str,
    sheet_name: str,
    table_name: str,
    require_table: bool = True,
    convert_integer_floats: bool = True,
) -> pd.DataFrame:
    """Read an Excel table (ListObject) by name into a pandas DataFrame.

    Args:
        xlsx_path: Path to the Excel workbook
        sheet_name: Name of the worksheet containing the table
        table_name: Name of the Excel table to read
        require_table: If True, raise an error if table is not found
        convert_integer_floats: If True, convert float columns with whole numbers to Int64
        
    Returns:
        DataFrame containing the table data
        
    Raises:
        RuntimeError: If workbook cannot be opened
        KeyError: If sheet or table is not found (when require_table=True)
        ValueError: If table reference format is invalid
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    
    logger.debug(f"Reading table '{table_name}' from sheet '{sheet_name}' in {xlsx_path}")
    
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        raise RuntimeError(f"Failed to open workbook '{xlsx_path}': {e}") from e

    if sheet_name not in wb.sheetnames:
        if require_table:
            raise KeyError(f"Sheet not found: {sheet_name} in {xlsx_path}")
        return pd.DataFrame()

    ws = wb[sheet_name]

    if table_name not in ws.tables:
        if require_table:
            raise KeyError(
                f"Table {table_name!r} not found on sheet {sheet_name!r} in {xlsx_path}"
            )
        return pd.DataFrame()

    table = ws.tables[table_name]
    ref = table.ref
    if not isinstance(ref, str) or ":" not in ref:
        raise ValueError(
            f"Unexpected table.ref for {table_name!r} on {sheet_name!r}: {ref!r}"
        )

    try:
        start, end = ref.split(":")
        start_col, start_row = coordinate_from_string(start)
        end_col, end_row = coordinate_from_string(end)
    except Exception as e:
        raise ValueError(
            f"Could not parse table.ref '{ref}' for {table_name!r}: {e}"
        ) from e

    header_row_idx = int(start_row) - 1
    usecols = f"{start_col}:{end_col}"
    nrows = int(end_row) - int(start_row)
    
    if nrows <= 0:
        try:
            df = pd.read_excel(
                xlsx_path,
                sheet_name=sheet_name,
                header=header_row_idx,
                usecols=usecols,
                nrows=0,
                engine="openpyxl",
            )
            df.columns = df.columns.str.strip()
            return df
        except Exception:
            return pd.DataFrame()

    try:
        df = pd.read_excel(
            xlsx_path,
            sheet_name=sheet_name,
            header=header_row_idx,
            usecols=usecols,
            nrows=nrows,
            engine="openpyxl",
        )
    except Exception as e:
        raise RuntimeError(
            f"pandas.read_excel failed for table {table_name!r} on sheet {sheet_name!r}: {e}"
        ) from e

    df.columns = df.columns.str.strip()

    def _trim_series(s: pd.Series) -> pd.Series:
        if pd.api.types.is_string_dtype(s):
            return s.str.strip()
        if s.dtype == object:
            return s.map(lambda v: v.strip() if isinstance(v, str) else v)
        return s

    df = df.apply(_trim_series)

    if convert_integer_floats:
        float_cols = df.select_dtypes(include=["float"]).columns
        for col in float_cols:
            ser = df[col]
            non_na = ser.dropna()
            if non_na.empty:
                continue
            try:
                if ((non_na % 1) == 0).all():
                    df[col] = df[col].astype("Int64")
            except Exception:
                continue

    logger.debug(f"Read {len(df)} rows from table '{table_name}'")
    return df


def read_table_as_dict(
    xlsx_path: str,
    sheet_name: str,
    table_name: str,
    key_col: str | None = None,
    value_col: str | None = None,
    require_unique_keys: bool = True,
) -> dict:
    """Read a two-column Excel table and return a dict mapping key->value.

    Args:
        xlsx_path: Path to the Excel workbook
        sheet_name: Name of the worksheet containing the table
        table_name: Name of the Excel table to read
        key_col: Column name to use as key (defaults to first column)
        value_col: Column name to use as value (defaults to second column)
        require_unique_keys: If True, raise error on duplicate keys
        
    Returns:
        Dictionary mapping keys to values
        
    Raises:
        ValueError: If table doesn't have exactly 2 columns or has duplicate keys
        KeyError: If specified key/value columns are not found
    """
    logger.debug(f"Reading table '{table_name}' as dictionary")
    df = read_table_as_df(xlsx_path, sheet_name, table_name, require_table=True)

    if df.shape[1] != 2:
        raise ValueError(
            f"Table {table_name!r} on sheet {sheet_name!r} must have exactly 2 columns (found {df.shape[1]})"
        )

    col_names = list(df.columns)
    key_col_name = key_col if key_col is not None else col_names[0]
    value_col_name = value_col if value_col is not None else col_names[1]

    if key_col_name not in df.columns or value_col_name not in df.columns:
        raise KeyError(
            f"Specified key/value columns not found in table columns: {df.columns.tolist()}"
        )

    mapping: dict = {}
    for idx, row in df.iterrows():
        raw_key = row[key_col_name]
        raw_val = row[value_col_name]

        if pd.isna(raw_key):
            continue

        key = raw_key.strip() if isinstance(raw_key, str) else raw_key
        val = raw_val.strip() if isinstance(raw_val, str) else raw_val

        if require_unique_keys and key in mapping:
            raise ValueError(
                f"Duplicate key found in table {table_name!r} on sheet {sheet_name!r}: {key!r}"
            )
        mapping[key] = val

    logger.debug(f"Loaded {len(mapping)} key-value pairs from table '{table_name}'")
    return mapping


def add_table_dataframe(
    wb: Workbook,
    sheet_name: str,
    table_name: str,
    data: pd.DataFrame,
    require_all_columns: bool = False,  # Changed default to False for more forgiving behavior
    keep_vba: bool = True,
    debug: bool = False,
) -> int:
    """Append a pandas.DataFrame to an existing Excel table.
    
    This function intelligently handles column mismatches:
    - Validates DataFrame has all required columns for the table
    - Uses only columns that exist in both DataFrame and OJS table
    - Warns about extra columns in DataFrame or missing columns in OJS
    - Fills OJS columns not in DataFrame with None
    
    Args:
        wb: The Excel workbook object
        sheet_name: Name of the worksheet containing the table
        table_name: Name of the Excel table to append to
        data: DataFrame containing the data to append
        require_all_columns: If True, require exact column match (legacy mode)
        keep_vba: Placeholder for VBA preservation (not currently used)
        debug: If True, print diagnostic information
        
    Returns:
        Number of rows written to the table
        
    Raises:
        KeyError: If sheet or table is not found
        ValueError: If DataFrame is missing required columns
    """
    if data is None or data.empty:
        print_error(logger, "Attempting to add empty or None DataFrame to table. "
                   f"Sheet: {sheet_name}, Table: {table_name}")
        return 0

    if sheet_name not in wb.sheetnames:
        print_error(
            logger,
            f"Sheet '{sheet_name}' not found in workbook",
            error_type='missing_sheet',
            context={
                'workbook': 'OJS file',
                'sheet_name': sheet_name,
                'available_sheets': list(wb.sheetnames)
            }
        )
    ws = wb[sheet_name]

    if table_name not in ws.tables:
        available_tables = ', '.join(ws.tables.keys()) if ws.tables else 'none'
        print_error(
            logger,
            f"Table '{table_name}' not found on sheet '{sheet_name}'",
            error_type='missing_table',
            context={
                'table_name': table_name,
                'sheet_name': sheet_name,
                'available_tables': list(ws.tables.keys())
            }
        )
    
    logger.debug(f"Adding {len(data)} rows to table '{table_name}' on sheet '{sheet_name}'")
    
    table = ws.tables[table_name]
    table_range = table.ref

    table_head = ws[table_range][0]
    table_data_rows = ws[table_range][1:]
    headers = [
        c.value.strip() if isinstance(c.value, str) else c.value for c in table_head
    ]

    df = data.copy()
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    
    # Validate required columns
    if table_name in REQUIRED_COLUMNS:
        required_cols = REQUIRED_COLUMNS[table_name]
        missing_required = [col for col in required_cols if col not in df.columns]
        if missing_required:
            raise ValueError(
                f"DataFrame is missing required columns for table {table_name!r}: {missing_required}\n"
                f"Required: {required_cols}\n"
                f"DataFrame has: {list(df.columns)}"
            )
        logger.debug(f"✓ DataFrame has all required columns for {table_name}")
    
    # Check for column mismatches (more forgiving approach)
    df_cols_set = set(df.columns)
    table_cols_set = set(headers)
    
    extra_in_df = df_cols_set - table_cols_set
    missing_in_df = table_cols_set - df_cols_set
    matching_cols = df_cols_set & table_cols_set
    
    if extra_in_df:
        logger.warning(
            f"DataFrame has {len(extra_in_df)} column(s) not in OJS table '{table_name}': {sorted(extra_in_df)}"
        )
        logger.warning("These columns will be ignored")
    
    if missing_in_df:
        logger.info(
            f"OJS table '{table_name}' has {len(missing_in_df)} column(s) not in DataFrame: {sorted(missing_in_df)}"
        )
        logger.info("These columns will be filled with None")

    # Legacy strict mode
    if require_all_columns:
        if list(df.columns) != headers:
            raise ValueError(
                f"DataFrame columns do not match table headers for {table_name!r} on sheet {sheet_name!r}.\n"
                f"Table headers: {headers}\nDataFrame columns: {list(df.columns)}"
            )

    start_cell, end_cell = table_range.split(":")
    _, end_row = coordinate_from_string(end_cell)
    start_col_letter, start_row = coordinate_from_string(start_cell)
    start_col_idx = column_index_from_string(start_col_letter)

    rows_written = 0
    df_iter_index = 0
    total_rows = len(df)

    # Fill existing blank rows
    for row_tuple in table_data_rows:
        if df_iter_index >= total_rows:
            break
        if all(
            (cell.value is None)
            or (isinstance(cell.value, str) and cell.value.strip() == "")
            for cell in row_tuple
        ):
            target_row_idx = row_tuple[0].row
            row_values = df.iloc[df_iter_index]
            for j, col_name in enumerate(headers):
                # Only use value if column exists in DataFrame
                val = row_values[col_name] if col_name in df.columns else None
                ws.cell(row=target_row_idx, column=start_col_idx + j).value = val
            df_iter_index += 1
            rows_written += 1

    # Append remaining rows
    current_row = int(end_row)
    appended = False
    while df_iter_index < total_rows:
        current_row += 1
        row_values = df.iloc[df_iter_index]
        for j, col_name in enumerate(headers):
            # Only use value if column exists in DataFrame
            val = row_values[col_name] if col_name in df.columns else None
            ws.cell(row=current_row, column=start_col_idx + j).value = val
        df_iter_index += 1
        rows_written += 1
        appended = True

    if appended:
        end_col_letter = coordinate_from_string(end_cell)[0]
        table.ref = f"{start_cell}:{end_col_letter}{current_row}"

    logger.debug(f"Wrote {rows_written} rows to table '{table_name}'")
    return rows_written
