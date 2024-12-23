#pip install openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles.protection import Protection
import os, sys, re
# pip install print-color
from print_color import print

wb: Workbook = None
def is_worksheet_protected(wb, sheet_name):
    """Checks if a specific worksheet in an Excel workbook is protected."""
    try:
        # print(f'Checking {sheet_name}')
        ws = wb[sheet_name]
        # print("Worksheet is set")
        return ws.protection.sheet is True
    except Exception as e:
        print(f"Error: {e}")
        return False

try:
    directory_list: list[str] = os.listdir()
    print(
        f"Found {len(directory_list)} files in the folder.",
        tag="info",
        tag_color="white",
        color="white",
    )
except Exception as e:
    print(
        f"*-*-* Could not get a directory list. We got this error: {e}",
        tag="error",
        tag_color="red",
        color="red",
    )
    input("Press enter to quit...")
    sys.exit(1)

print("Looking for the OJS spreadsheets", tag="info", tag_color="white", color="white")
xlsm_files: list[str] = [s for s in directory_list if s.endswith(".xlsm")]
worksheet_list: list[str] = ["Results and Rankings", "Robot Game Scores", "Innovation Project Input", "Robot Design Input", "Core Values Input"]

# open wb
for f in xlsm_files:
    print(f"Workbook: {f}")
    wb = load_workbook(f, read_only=False, keep_vba=True)
    edited = False
    for ws_name in worksheet_list:
        ws = wb[ws_name]
        # print(ws)
        if (is_worksheet_protected(wb, ws_name)):
            print(f"{ws_name} is protected")
        else:
            print(f"{ws_name} is not protected. Protecting now.")
            edited = True
            ws.protection.enabled = True
            ws.protection.password = 'skip'
    if edited: 
        print(
            f"Saving {f}",
            tag="complete",
            tag_color="green",
            color="green",
        )
        wb.save(f)
    wb.close()


# select sheet1
# ws = wb["Sheet1"]

# protect sheet

# allow filtering on all cols
# ws.auto_filter.ref = "A1:H1"
# ws.auto_filter.enable = True

# allow editing on all cells in col D
# for row in ws.iter_rows():
#     for cell in row:
#         cell.protection = Protection(locked=(cell.column != "D"))

# save to new file
# wb.save("test9.xlsx")